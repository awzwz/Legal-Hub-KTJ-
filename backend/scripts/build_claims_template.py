"""Создаёт пустой шаблон `claims_registry_template.xlsx` из исходного реестра претензий.

Логика:
1. Загружаем `реестр претензии 25-26гг.xlsx`.
2. Удаляем строки данных (с 4-й строки до конца) на листах с данными.
3. Удаляем мусорный лист `Лист1`.
4. Сохраняем как `backend/templates/claims_registry_template.xlsx`.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


SOURCE = Path(__file__).resolve().parents[2] / "реестр претензии 25-26гг.xlsx"
TARGET = Path(__file__).resolve().parents[1] / "templates" / "claims_registry_template.xlsx"

# Листы с данными — с какой строки начинаются данные (после шапки)
DATA_SHEETS = {
    "претензии 2025": 4,
    "претензии 2026": 4,
}


def main() -> None:
    if not SOURCE.is_file():
        raise FileNotFoundError(f"Source missing: {SOURCE}")

    wb = load_workbook(SOURCE)

    # Удаляем мусорный «Лист1», если есть
    if "Лист1" in wb.sheetnames:
        del wb["Лист1"]

    # На листах с данными — снимаем merges ниже шапки и очищаем строки данных
    for sheet_name, first_data_row in DATA_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # Снимаем все merges, начинающиеся ниже шапки
        for cr in list(ws.merged_cells.ranges):
            if cr.min_row >= first_data_row:
                ws.unmerge_cells(str(cr))
        # Очищаем значения в строках данных
        for row in ws.iter_rows(min_row=first_data_row, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None

    TARGET.parent.mkdir(parents=True, exist_ok=True)
    wb.save(TARGET)
    print(f"Saved template: {TARGET}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
