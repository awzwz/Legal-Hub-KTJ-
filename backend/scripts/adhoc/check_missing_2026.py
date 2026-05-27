import asyncio
from openpyxl import load_workbook
from scripts.import_pir_q1_2026 import parse_sheet, ROLE_BY_SHEET

def test():
    wb = load_workbook("/app/2026.xlsx", data_only=True)
    
    for sheet_name, role in ROLE_BY_SHEET.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        
        parsed_rows = parse_sheet(ws, role)
        parsed_src_rows = {p['src_row'] for p in parsed_rows}
        
        count = 0
        excel_rows = set()
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            b = str(ws.cell(row=r, column=2).value or "").lower().strip()
            c = str(ws.cell(row=r, column=3).value or "").lower().strip()
            
            if b in ("предъявлено", "категория", "кол-во", "итого") or c in ("предъявлено", "категория", "кол-во", "итого"):
                break
                
            if isinstance(a, (int, float)) and int(a) > 0:
                excel_rows.add(r)
                count += 1
                
        print(f"Sheet '{sheet_name}': Excel valid rows={count}, Parsed={len(parsed_rows)}")
        missing = sorted(list(excel_rows - parsed_src_rows))
        if missing:
            print(f"Missing rows in {sheet_name}: {missing}")
            for m in missing:
                row_data = [ws.cell(row=m, column=col).value for col in range(1, 10)]
                print(f"  Row {m}: {row_data}")

test()
