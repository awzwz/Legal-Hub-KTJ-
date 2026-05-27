from openpyxl import load_workbook

wb25 = load_workbook("/Users/awz_wz/Documents/KTJ legalhub/legalhub-project-ktz/Отчет ПИР за 2025г (final).xlsx", data_only=True)
wb26 = load_workbook("/Users/awz_wz/Documents/KTJ legalhub/legalhub-project-ktz/Отчет ПИР за 1 кв.2026г.xlsx", data_only=True)

lawyers = set()

for wb in [wb25, wb26]:
    for sheet_name in ["истец", "ответчик", "3-лицо "]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if isinstance(a, (int, float)) and int(a) > 0:
                val = ws.cell(row=r, column=2).value
                if isinstance(val, str):
                    lawyers.add(val.strip())

print("All unique lawyer names in Excel:")
for l in sorted(list(lawyers)):
    print(repr(l))
