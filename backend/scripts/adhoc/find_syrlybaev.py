from openpyxl import load_workbook

wb = load_workbook("/Users/awz_wz/Documents/KTJ legalhub/legalhub-project-ktz/Отчет ПИР за 1 кв.2026г.xlsx", data_only=True)
found = False

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and "сырлыба" in val.lower():
                print(f"Found in {sheet_name}, row {r}, col {c}: {val}")
                found = True

if not found:
    print("Not found in 2026.")

wb2 = load_workbook("/Users/awz_wz/Documents/KTJ legalhub/Отчет ПИР за 2025г (final).xlsx", data_only=True)
for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and "сырлыба" in val.lower():
                print(f"Found in 2025 {sheet_name}, row {r}, col {c}: {val}")

