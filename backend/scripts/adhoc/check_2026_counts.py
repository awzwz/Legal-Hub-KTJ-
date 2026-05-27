import sys
from openpyxl import load_workbook

wb = load_workbook("/Users/awz_wz/Documents/KTJ legalhub/legalhub-project-ktz/Отчет ПИР за 1 кв.2026г.xlsx", data_only=True)

for sheet_name in ["истец", "ответчик", "3-лицо "]:
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found!")
        continue
    ws = wb[sheet_name]
    
    count = 0
    # count rows where col A is a number (case number)
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        # also if 'Итого' is reached we stop
        b = str(ws.cell(row=r, column=2).value or "").lower().strip()
        c = str(ws.cell(row=r, column=3).value or "").lower().strip()
        
        if b in ("предъявлено", "категория", "кол-во", "итого") or c in ("предъявлено", "категория", "кол-во", "итого"):
            break
            
        if isinstance(a, (int, float)) and int(a) > 0:
            count += 1
            
    print(f"Sheet '{sheet_name}': {count} cases")
