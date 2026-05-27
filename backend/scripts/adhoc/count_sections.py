import openpyxl

wb = openpyxl.load_workbook("/Users/awz_wz/Documents/KTJ legalhub/Отчет ПИР за 2025г (final).xlsx", data_only=True)
ws = wb["Истец"]

counts = {}
current_section = "Закупки / договоры" # Default first section
counts[current_section] = 0

for row in range(5, ws.max_row + 1):
    cell_val = ws.cell(row=row, column=2).value
    # Skip empty or fully merged rows that might not be actual cases
    if cell_val is None:
        continue
        
    val_str = str(cell_val).strip()
    
    # Check if it's a section header
    if val_str.upper() in ["ТРУДОВЫЕ СПОРЫ", "МЕДИАТИВНЫЕ СОГЛАШЕНИЯ", "СПОРЫ С ГОСОРГАНАМИ", "ПЕРЕВОЗОЧНЫЕ", "ИНЫЕ", "ОСТАВЛЕНО БЕЗ РАССМОТРЕНИЯ"]:
        current_section = val_str
        counts[current_section] = 0
        continue
        
    if "ИТОГО ПО ИСКАМ АО «ПАССАЖИРСКИЕ ПЕРЕВОЗКИ»" in val_str.upper() or "ИТОГО" in val_str.upper():
        continue
        
    # It's a case row (if it has a number or text that's not a known header)
    counts[current_section] += 1

print(counts)
