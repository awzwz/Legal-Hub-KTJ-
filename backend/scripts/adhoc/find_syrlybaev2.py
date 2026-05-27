from openpyxl import load_workbook

wb2 = load_workbook("/Users/awz_wz/Documents/KTJ legalhub/legalhub-project-ktz/Отчет ПИР за 2025г (final).xlsx", data_only=True)
for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and "сырлыба" in val.lower():
                print(f"Found in 2025 {sheet_name}, row {r}, col {c}: {val}")
