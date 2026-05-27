import asyncio
from sqlalchemy import select
from openpyxl import load_workbook
from app.db.session import SessionLocal
from app.models.case import Case

async def compare():
    async with SessionLocal() as session:
        result = await session.execute(
            select(Case.case_number)
            .where(Case.filing_date >= '2026-01-01')
        )
        db_cases = {row[0] for row in result.fetchall()}

    wb = load_workbook("/app/2026.xlsx", data_only=True)
    
    excel_cases = set()
    role_map = {"истец": "plaintiff", "ответчик": "defendant", "3-лицо ": "third_party"}
    
    for sheet_name, role in role_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            b = str(ws.cell(row=r, column=2).value or "").lower().strip()
            c = str(ws.cell(row=r, column=3).value or "").lower().strip()
            
            if b in ("предъявлено", "категория", "кол-во", "итого") or c in ("предъявлено", "категория", "кол-во", "итого"):
                break
                
            if isinstance(a, (int, float)) and int(a) > 0:
                excel_cases.add(f"2026-{int(a):03d}")
                
    missing = sorted(list(excel_cases - db_cases))
    print("Missing in DB:", missing)
    
    for sheet_name in role_map.keys():
        ws = wb[sheet_name]
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if isinstance(a, (int, float)) and int(a) > 0:
                cnum = f"2026-{int(a):03d}"
                if cnum in missing:
                    print(f"Row {r} in {sheet_name} missing: Case {cnum}")
                    # Print contents
                    row_data = [ws.cell(row=r, column=col).value for col in range(1, 10)]
                    print(row_data)

asyncio.run(compare())
