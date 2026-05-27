from openpyxl import load_workbook

wb = load_workbook("/Users/awz_wz/Documents/KTJ legalhub/legalhub-project-ktz/Отчет ПИР за 1 кв.2026г.xlsx", data_only=True)
lawyers = set()

for sheet_name in ["истец", "ответчик", "3-лицо "]:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    
    for r in range(1, ws.max_row + 1):
        # Column 2 is lawyer (for plaintiff, maybe different for others?)
        # Let's just collect column 2
        a = ws.cell(row=r, column=1).value
        if isinstance(a, (int, float)) and int(a) > 0:
            val = ws.cell(row=r, column=2).value
            if isinstance(val, str):
                lawyers.add(val.strip())

print(sorted(list(lawyers)))
