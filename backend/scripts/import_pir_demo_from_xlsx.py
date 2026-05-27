#!/usr/bin/env python3
"""
Импорт строк из файла ПИР (xlsx) в БД: листы «истец», «ответчик», «3-лицо »,
«исполнительное производство», «инф по сниже дебит. задолжненно».

Usage (из каталога backend/, venv, DATABASE_URL в .env):

  python scripts/import_pir_demo_from_xlsx.py --xlsx templates/pir_report_2025_template.xlsx
  python scripts/import_pir_demo_from_xlsx.py --xlsx /path/to/PIR_2025.xlsx --purge-cases-first

По умолчанию импортируются все строки до пустой зоны (не более --max-rows-per-sheet на лист).
Колонки «юрист» (2) и «подразделение» (4) сопоставляются с users / branches; при отсутствии филиала
он создаётся, при отсутствии юриста — добавляется пользователь pir-*@import.local (пароль legalhub123).
"""

from __future__ import annotations

import argparse
import asyncio
import re
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.core.security import hash_password
from app.db.session import SessionLocal
from app.models import (
    Branch,
    Case,
    CaseEvent,
    CaseFinance,
    CaseLitigation,
    DebtRecoveryEntry,
    EnforcementProceeding,
    User,
)
from app.domain.pir_excel_fill import (
    DEBT_FIRST_ROW,
    DEBT_SHEET,
    ENFORCEMENT_FIRST_ROW,
    ENFORCEMENT_SHEET,
    FIRST_DATA_ROW,
)

_DEFAULT_COMPANY = "АО «Пассажирские перевозки»"
_DEFAULT_BIN = "020540000922"


_CATEGORY_PATTERNS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("mediation", ("медиатив",)),
    ("transportation", ("перевозочного процесса", "перевозок")),
    ("labor", ("трудовые споры", "трудовой спор")),
    ("procurement", ("закупках", "закупок", "договоров")),
    ("other", ("иные споры", "иные иски", "прочие споры", "иные")),
)


def _detect_category_from_header(text: str) -> str | None:
    """Возвращает категорию по тексту строки-разделителя (или None)."""
    if not text:
        return None
    low = text.lower()
    for cat, needles in _CATEGORY_PATTERNS:
        for n in needles:
            if n in low:
                return cat
    return None


def _build_category_map(ws, start_row: int, max_rows: int) -> dict[int, str]:
    """Сканирует колонку A с строки 1: после строки-разделителя категория
    распространяется на все последующие data-строки до следующего разделителя.
    Возвращает только маппинг для строк ≥ start_row, но категория правильно
    «протекает» из заголовков, расположенных ВЫШЕ start_row.
    """
    result: dict[int, str] = {}
    current = "procurement"
    for r in range(1, start_row + max_rows):
        a = ws.cell(r, 1).value
        if isinstance(a, str):
            cat = _detect_category_from_header(a)
            if cat is not None:
                current = cat
                continue
        if r >= start_row:
            result[r] = current
    return result


def _infer_outcome(category: str, judgment_first: str, judgment_appeal: str, judgment_cassation: str) -> str:
    """Определяет outcome по категории и текстам решений (cassation > appeal > first).

    settled — дела из раздела «Медиативные соглашения».
    fully_satisfied / partially_satisfied / denied — по тексту последнего по уровню решения.
    pending — если ничего не подходит.
    """
    if category == "mediation":
        return "settled"
    # Берём «самое позднее» по инстанции решение
    text = (judgment_cassation or judgment_appeal or judgment_first or "").lower()
    if not text.strip():
        return "pending"
    if "удовлетворен частично" in text or "удовлетворены частично" in text:
        return "partially_satisfied"
    if (
        "удовлетворен в полном объеме" in text
        or "иск удовлетворен" in text
        or "исковые требования удовлетворены" in text
    ):
        # уточняем: «удовлетворены частично» уже отлито выше
        return "fully_satisfied"
    if (
        "в иске отказано" in text
        or "в удовлетворении иска отказано" in text
        or "в удовлетворении искового" in text
        or "в удовлетворении исков" in text
        or "отказано в удовлетворении" in text
        or "в удовлетворении апелляционной жалобы отказано" in text
    ):
        return "denied"
    return "pending"


def _infer_status(outcome: str, execution_proof_note: str, has_collection: bool) -> str:
    """Статус дела: closed для решённых, execution если есть исполнение, иначе active."""
    if outcome in {"fully_satisfied", "partially_satisfied", "settled"}:
        if execution_proof_note.strip() or has_collection:
            return "execution"
        return "closed"
    if outcome == "denied":
        return "closed"
    return "active"


def _infer_court_instance(judgment_first: str, judgment_appeal: str, judgment_cassation: str) -> str:
    if judgment_cassation and judgment_cassation.strip():
        return "cassation"
    if judgment_appeal and judgment_appeal.strip():
        return "appeal"
    return "first"


def _infer_case_type(category: str, status: str) -> str:
    if category == "labor":
        return "labor"
    if status == "execution":
        return "executive"
    return "civil"


def _infer_risk(claim_amount: Decimal) -> str:
    if claim_amount > Decimal("50000000"):
        return "high"
    if claim_amount < Decimal("5000000"):
        return "low"
    return "medium"


def _parse_court_cell(val: object) -> tuple[str, str]:
    s = (str(val) if val is not None else "").strip()
    if not s:
        return "Суд не указан", "—"
    parts = [p.strip() for p in s.replace("\r", "").split("\n") if p.strip()]
    if len(parts) >= 2:
        return parts[0][:512], parts[-1][:255]
    return s[:512], "—"


def _dec(v: object) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    try:
        return Decimal(str(v).replace(" ", "").replace(",", "."))
    except Exception:
        return Decimal("0")


def _txt(v: object, mx: int = 8000) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return s[:mx] if len(s) > mx else s


def _norm_name(s: str) -> str:
    return re.sub(r"\s+", " ", s.replace("\r", "").replace("\n", " ")).strip().casefold()


def _canon_label(s: str) -> str:
    return " ".join(s.split()) if s else ""


_SUMMARY_KEYWORDS = ("предъявлено", "удовлетворено", "отказано", "кол-во дел", "кол-во")


def _is_summary_row(ws, r: int) -> bool:
    """Строки сводного блока (предъявлено / удовлетворено / отказано / кол-во).

    Не путать с строкой-разделителем «Медиативные соглашения», которая идёт
    в колонке A и обозначает раздел.
    """
    a = ws.cell(r, 1).value
    if isinstance(a, str) and _detect_category_from_header(a):
        return False  # это раздел, а не сводка
    for c in range(2, 8):  # сводный блок начинается с колонки D-F
        v = ws.cell(r, c).value
        if isinstance(v, str):
            low = v.strip().lower()
            for kw in _SUMMARY_KEYWORDS:
                if kw in low:
                    return True
    return False


def _is_section_header_row(ws, r: int) -> bool:
    """Строка-разделитель категории: одна непустая ячейка в колонке A."""
    a = ws.cell(r, 1).value
    return isinstance(a, str) and _detect_category_from_header(a) is not None


def _row_nonempty_истец(ws, r: int) -> bool:
    if _is_summary_row(ws, r):
        return False
    # Требуем юриста, контрагента или текст иска
    return (
        _txt(ws.cell(r, 2).value) != ""
        or _txt(ws.cell(r, 5).value) != ""
        or _txt(ws.cell(r, 9).value) != ""
    )


def _row_nonempty_ответчик(ws, r: int) -> bool:
    if _is_summary_row(ws, r):
        return False
    # На листе ответчик колонка 2 (юрист) часто пуста — проверяем филиал/содержание иска
    return (
        _txt(ws.cell(r, 4).value) != ""
        or _txt(ws.cell(r, 5).value) != ""
        or _txt(ws.cell(r, 10).value) != ""
    )


def _row_nonempty_3лицо(ws, r: int) -> bool:
    if _is_summary_row(ws, r):
        return False
    return (
        _txt(ws.cell(r, 2).value) != ""
        or _txt(ws.cell(r, 5).value) != ""
        or _txt(ws.cell(r, 6).value) != ""
        or _txt(ws.cell(r, 11).value) != ""
    )


def _row_nonempty_enf(ws, r: int) -> bool:
    return _txt(ws.cell(r, 2).value) != "" or _txt(ws.cell(r, 4).value) != ""


def _row_nonempty_debt(ws, r: int) -> bool:
    return _txt(ws.cell(r, 2).value) != ""


class _CaseIndex:
    def __init__(self) -> None:
        self.by_defendant: dict[str, list[uuid.UUID]] = {}
        self.by_plaintiff: dict[str, list[uuid.UUID]] = {}
        self.by_bin: dict[str, uuid.UUID] = {}

    def add(self, case_id: uuid.UUID, defendant: str, plaintiff: str, company_bin: str) -> None:
        d = _norm_name(defendant)
        p = _norm_name(plaintiff)
        if d:
            self.by_defendant.setdefault(d, []).append(case_id)
        if p:
            self.by_plaintiff.setdefault(p, []).append(case_id)
        if company_bin and company_bin != _DEFAULT_BIN:
            self.by_bin[company_bin.strip()] = case_id

    def find_for_enforcement(self, debtor_name: str, debtor_bin: str | None) -> uuid.UUID | None:
        if debtor_bin:
            b = debtor_bin.strip()
            if b in self.by_bin:
                return self.by_bin[b]
        dn = _norm_name(debtor_name)
        if not dn:
            return None
        if dn in self.by_defendant:
            return self.by_defendant[dn][0]
        for key, ids in self.by_defendant.items():
            if dn in key or key in dn:
                return ids[0]
        if dn in self.by_plaintiff:
            return self.by_plaintiff[dn][0]
        return None


def _build_case_litigation(
    case_id: uuid.UUID,
    *,
    claim_summary: str = "",
    judgment_first: str = "",
    judgment_appeal: str = "",
    judgment_cassation: str = "",
    damage_recovery_note: str = "",
    writ_request_note: str = "",
    writ_dispatch_note: str = "",
    execution_proof_note: str = "",
) -> CaseLitigation:
    now = datetime.now(timezone.utc)
    return CaseLitigation(
        case_id=case_id,
        claim_summary=claim_summary,
        judgment_first=judgment_first,
        judgment_appeal=judgment_appeal,
        judgment_cassation=judgment_cassation,
        damage_recovery_note=damage_recovery_note,
        writ_request_note=writ_request_note,
        writ_dispatch_note=writ_dispatch_note,
        execution_proof_note=execution_proof_note,
        created_at=now,
        updated_at=now,
    )


async def main_async(
    xlsx: Path,
    *,
    limit: int | None,
    max_rows_per_sheet: int,
    dry_run: bool,
    as_of: date,
    company: str,
    company_bin: str,
    purge_cases_first: bool,
) -> None:
    wb = load_workbook(xlsx, read_only=False, data_only=True)
    try:
        async with SessionLocal() as db:
            if purge_cases_first and not dry_run:
                await db.execute(delete(DebtRecoveryEntry))
                await db.execute(delete(Case))
                await db.commit()

            branches = (await db.execute(select(Branch).order_by(Branch.name))).scalars().all()
            if not branches:
                raise SystemExit("No branches in DB; run seed/migrations first.")
            stub_branch_id = branches[0].id

            lawyer_rows = (await db.execute(select(User).where(User.role == "branch_lawyer"))).scalars().all()
            if not lawyer_rows:
                all_u = (await db.execute(select(User))).scalars().all()
                if not all_u:
                    raise SystemExit("No users in DB.")
                fallback_id = all_u[0].id
            else:
                fallback_id = lawyer_rows[0].id

            branch_by_key: dict[str, uuid.UUID] = {}
            for b in branches:
                k = _canon_label(b.name) or "—"
                branch_by_key[k] = b.id

            user_by_name: dict[str, User] = {}
            for u in (await db.execute(select(User))).scalars().all():
                fk = _canon_label(u.full_name)
                if fk and fk not in user_by_name:
                    user_by_name[fk] = u

            async def branch_id_for(raw: object) -> uuid.UUID:
                label = (_txt(raw, 255) or "—")[:255]
                key = _canon_label(label) or "—"
                if key in branch_by_key:
                    return branch_by_key[key]
                if dry_run:
                    return stub_branch_id
                existing = (await db.execute(select(Branch).where(Branch.name == label))).scalar_one_or_none()
                if existing is not None:
                    branch_by_key[key] = existing.id
                    return existing.id
                nb = Branch(id=uuid.uuid4(), name=label, city=None)
                db.add(nb)
                await db.flush()
                branch_by_key[key] = nb.id
                return nb.id

            async def lawyer_id_for(raw: object, branch_id: uuid.UUID) -> uuid.UUID:
                fn = _txt(raw, 255)
                key = _canon_label(fn)
                if not key:
                    return fallback_id
                u = user_by_name.get(key)
                if u is not None:
                    return u.id
                if dry_run:
                    return fallback_id
                email = f"pir-{uuid.uuid4().hex[:16]}@import.local"
                nu = User(
                    id=uuid.uuid4(),
                    email=email,
                    password_hash=hash_password("legalhub123"),
                    full_name=fn[:255],
                    role="branch_lawyer",
                    branch_id=branch_id,
                    is_active=True,
                )
                db.add(nu)
                await db.flush()
                user_by_name[key] = nu
                return nu.id

            index = _CaseIndex()
            to_add: list[Any] = []
            n_cases = 0

            # --- истец ---
            ws = wb["истец"]
            start = FIRST_DATA_ROW["истец"]
            cat_map_истец = _build_category_map(ws, start, max_rows_per_sheet)
            r = start
            empty_run = 0
            seen_истец = 0
            while r < start + max_rows_per_sheet:
                if limit is not None and n_cases >= limit:
                    break
                if _is_section_header_row(ws, r):
                    empty_run = 0
                    r += 1
                    continue
                if not _row_nonempty_истец(ws, r):
                    empty_run += 1
                    if empty_run >= 5 and seen_истец > 0:
                        break
                    r += 1
                    continue
                empty_run = 0
                seen_истец += 1
                bid = await branch_id_for(ws.cell(r, 4).value)
                lid = await lawyer_id_for(ws.cell(r, 2).value, bid)
                court, judge = _parse_court_cell(ws.cell(r, 3).value)
                defendant = _txt(ws.cell(r, 5).value, 512) or "—"
                case_id = uuid.uuid4()
                case_number = f"PIR-IMP-{case_id.hex[:10].upper()}"
                # Инференс из категории + текстов решений
                disp_cat = cat_map_истец.get(r, "procurement")
                j1 = _txt(ws.cell(r, 10).value)
                ja = _txt(ws.cell(r, 11).value)
                jc = _txt(ws.cell(r, 12).value)
                exec_proof = _txt(ws.cell(r, 18).value)
                rm = _dec(ws.cell(r, 13).value)
                rf = _dec(ws.cell(r, 14).value)
                rsf = _dec(ws.cell(r, 15).value)
                main_debt = _dec(ws.cell(r, 6).value)
                fines = _dec(ws.cell(r, 7).value)
                state_fee = _dec(ws.cell(r, 8).value)
                claim_amount_total = main_debt + fines
                oc = _infer_outcome(disp_cat, j1, ja, jc)
                st = _infer_status(oc, exec_proof, (rm + rf + rsf) > Decimal("0"))
                ci = _infer_court_instance(j1, ja, jc)
                ct = _infer_case_type(disp_cat, st)
                rk = _infer_risk(claim_amount_total)
                c = Case(
                    id=case_id,
                    case_number=case_number,
                    court=court,
                    court_instance=ci,
                    case_type=ct,
                    status=st,
                    outcome=oc,
                    party_role="plaintiff",
                    opponent_type="juridical",
                    plaintiff=company[:512],
                    defendant=defendant,
                    company=defendant[:512],  # фактический контрагент = ответчик
                    company_bin="",
                    dispute_category=disp_cat,
                    city="—",
                    judge=judge,
                    filing_date=as_of,
                    next_hearing=None,
                    payment_deadline=None,
                    last_updated=as_of,
                    days_overdue=0,
                    risk_level=rk,
                    is_archived=False,
                    branch_id=bid,
                    assigned_lawyer_id=lid,
                )
                fin = CaseFinance(
                    case_id=case_id,
                    claim_amount=claim_amount_total,
                    main_debt=main_debt,
                    state_fee=state_fee,
                    fines=fines,
                    rep_expenses=Decimal("0"),
                    other_costs=Decimal("0"),
                    paid_amount=rm + rf + rsf,
                    recovered_main=rm,
                    recovered_fines=rf,
                    recovered_state_fee=rsf,
                )
                lit = _build_case_litigation(
                    case_id,
                    claim_summary=_txt(ws.cell(r, 9).value),
                    judgment_first=j1,
                    judgment_appeal=ja,
                    judgment_cassation=jc,
                    damage_recovery_note=_txt(ws.cell(r, 19).value),
                    writ_request_note=_txt(ws.cell(r, 16).value),
                    writ_dispatch_note=_txt(ws.cell(r, 17).value),
                    execution_proof_note=exec_proof,
                )
                ev = CaseEvent(
                    id=uuid.uuid4(),
                    case_id=case_id,
                    action="Импорт из файла ПИР (лист «истец»)",
                    user_label="import_pir_demo_from_xlsx.py",
                    detail=f"строка xlsx {r}",
                    happened_at=datetime.now(timezone.utc),
                    user_id=None,
                )
                if not dry_run:
                    to_add.extend([c, fin, lit, ev])
                index.add(case_id, c.defendant, c.plaintiff, c.company_bin)
                n_cases += 1
                print(f"истец row {r}: {case_number} — {defendant[:60]}")
                r += 1

            # --- ответчик ---
            ws = wb["ответчик"]
            start = FIRST_DATA_ROW["ответчик"]
            cat_map_отв = _build_category_map(ws, start, max_rows_per_sheet)
            r = start
            empty_run = 0
            seen_ответчик = 0
            while r < start + max_rows_per_sheet:
                if limit is not None and n_cases >= limit:
                    break
                if _is_section_header_row(ws, r):
                    empty_run = 0
                    r += 1
                    continue
                if not _row_nonempty_ответчик(ws, r):
                    empty_run += 1
                    if empty_run >= 5 and seen_ответчик > 0:
                        break
                    r += 1
                    continue
                empty_run = 0
                seen_ответчик += 1
                bid = await branch_id_for(ws.cell(r, 4).value)
                lid = await lawyer_id_for(ws.cell(r, 2).value, bid)
                court, judge = _parse_court_cell(ws.cell(r, 3).value)
                plaintiff = _txt(ws.cell(r, 5).value, 512) or company[:512]
                case_id = uuid.uuid4()
                case_number = f"PIR-IMP-{case_id.hex[:10].upper()}"
                disp_cat = cat_map_отв.get(r, "procurement")
                j1 = _txt(ws.cell(r, 11).value)
                ja = _txt(ws.cell(r, 12).value)
                jc = _txt(ws.cell(r, 13).value)
                def_exec = _txt(ws.cell(r, 18).value)
                main_debt = _dec(ws.cell(r, 6).value)
                fines = _dec(ws.cell(r, 7).value)
                rep_exp = _dec(ws.cell(r, 8).value)
                state_fee = _dec(ws.cell(r, 9).value)
                paid = _dec(ws.cell(r, 15).value)
                claim_amount_total = main_debt + fines + state_fee
                oc = _infer_outcome(disp_cat, j1, ja, jc)
                st = _infer_status(oc, def_exec, paid > Decimal("0"))
                ci = _infer_court_instance(j1, ja, jc)
                ct = _infer_case_type(disp_cat, st)
                rk = _infer_risk(claim_amount_total)
                c = Case(
                    id=case_id,
                    case_number=case_number,
                    court=court,
                    court_instance=ci,
                    case_type=ct,
                    status=st,
                    outcome=oc,
                    party_role="defendant",
                    opponent_type="juridical",
                    plaintiff=plaintiff,
                    defendant=company[:512],
                    company=plaintiff[:512],  # контрагент = истец
                    company_bin="",
                    dispute_category=disp_cat,
                    city="—",
                    judge=judge,
                    filing_date=as_of,
                    next_hearing=None,
                    payment_deadline=None,
                    last_updated=as_of,
                    days_overdue=0,
                    risk_level=rk,
                    is_archived=False,
                    branch_id=bid,
                    assigned_lawyer_id=lid,
                )
                fin = CaseFinance(
                    case_id=case_id,
                    claim_amount=claim_amount_total,
                    main_debt=main_debt,
                    state_fee=state_fee,
                    fines=fines,
                    rep_expenses=rep_exp,
                    other_costs=Decimal("0"),
                    paid_amount=paid,
                    recovered_main=Decimal("0"),
                    recovered_fines=Decimal("0"),
                    recovered_state_fee=Decimal("0"),
                )
                lit = _build_case_litigation(
                    case_id,
                    claim_summary=_txt(ws.cell(r, 10).value),
                    judgment_first=j1,
                    judgment_appeal=ja,
                    judgment_cassation=jc,
                    damage_recovery_note=_txt(ws.cell(r, 19).value),
                )
                # defendant_execution_note хранится отдельно (не в _build_case_litigation сигнатуре)
                if def_exec:
                    lit.defendant_execution_note = def_exec
                ev = CaseEvent(
                    id=uuid.uuid4(),
                    case_id=case_id,
                    action="Импорт из файла ПИР (лист «ответчик»)",
                    user_label="import_pir_demo_from_xlsx.py",
                    detail=f"строка xlsx {r}",
                    happened_at=datetime.now(timezone.utc),
                    user_id=None,
                )
                if not dry_run:
                    to_add.extend([c, fin, lit, ev])
                index.add(case_id, c.defendant, c.plaintiff, c.company_bin)
                n_cases += 1
                print(f"ответчик row {r}: {case_number}")
                r += 1

            # --- 3-лицо ---
            ws = wb["3-лицо "]
            start = FIRST_DATA_ROW["3-лицо "]
            cat_map_3 = _build_category_map(ws, start, max_rows_per_sheet)
            r = start
            empty_run = 0
            seen_3 = 0
            while r < start + max_rows_per_sheet:
                if limit is not None and n_cases >= limit:
                    break
                if _is_section_header_row(ws, r):
                    empty_run = 0
                    r += 1
                    continue
                if not _row_nonempty_3лицо(ws, r):
                    empty_run += 1
                    if empty_run >= 5 and seen_3 > 0:
                        break
                    r += 1
                    continue
                empty_run = 0
                seen_3 += 1
                bid = await branch_id_for(ws.cell(r, 4).value)
                lid = await lawyer_id_for(ws.cell(r, 2).value, bid)
                court, judge = _parse_court_cell(ws.cell(r, 3).value)
                plaintiff = _txt(ws.cell(r, 5).value, 512) or "—"
                defendant = _txt(ws.cell(r, 6).value, 512) or "—"
                case_id = uuid.uuid4()
                case_number = f"PIR-IMP-{case_id.hex[:10].upper()}"
                disp_cat = cat_map_3.get(r, "procurement")
                j1 = _txt(ws.cell(r, 12).value)
                ja = _txt(ws.cell(r, 13).value)
                jc = _txt(ws.cell(r, 14).value)
                main_debt = _dec(ws.cell(r, 7).value)
                fines = _dec(ws.cell(r, 8).value)
                rep_exp = _dec(ws.cell(r, 9).value)
                state_fee = _dec(ws.cell(r, 10).value)
                paid = _dec(ws.cell(r, 16).value)
                claim_amount_total = main_debt + fines + state_fee
                oc = _infer_outcome(disp_cat, j1, ja, jc)
                st = _infer_status(oc, "", paid > Decimal("0"))
                ci = _infer_court_instance(j1, ja, jc)
                ct = _infer_case_type(disp_cat, st)
                rk = _infer_risk(claim_amount_total)
                c = Case(
                    id=case_id,
                    case_number=case_number,
                    court=court,
                    court_instance=ci,
                    case_type=ct,
                    status=st,
                    outcome=oc,
                    party_role="third_party",
                    opponent_type="juridical",
                    plaintiff=plaintiff,
                    defendant=defendant,
                    company=defendant[:512],  # контрагент по делу как ответчик в споре
                    company_bin="",
                    dispute_category=disp_cat,
                    city="—",
                    judge=judge,
                    filing_date=as_of,
                    next_hearing=None,
                    payment_deadline=None,
                    last_updated=as_of,
                    days_overdue=0,
                    risk_level=rk,
                    is_archived=False,
                    branch_id=bid,
                    assigned_lawyer_id=lid,
                )
                fin = CaseFinance(
                    case_id=case_id,
                    claim_amount=claim_amount_total,
                    main_debt=main_debt,
                    state_fee=state_fee,
                    fines=fines,
                    rep_expenses=rep_exp,
                    other_costs=Decimal("0"),
                    paid_amount=paid,
                    recovered_main=Decimal("0"),
                    recovered_fines=Decimal("0"),
                    recovered_state_fee=Decimal("0"),
                )
                lit = _build_case_litigation(
                    case_id,
                    claim_summary=_txt(ws.cell(r, 11).value),
                    judgment_first=j1,
                    judgment_appeal=ja,
                    judgment_cassation=jc,
                    damage_recovery_note=_txt(ws.cell(r, 19).value),
                )
                ev = CaseEvent(
                    id=uuid.uuid4(),
                    case_id=case_id,
                    action="Импорт из файла ПИР (лист «3-лицо»)",
                    user_label="import_pir_demo_from_xlsx.py",
                    detail=f"строка xlsx {r}",
                    happened_at=datetime.now(timezone.utc),
                    user_id=None,
                )
                if not dry_run:
                    to_add.extend([c, fin, lit, ev])
                index.add(case_id, c.defendant, c.plaintiff, c.company_bin)
                n_cases += 1
                print(f"3-лицо row {r}: {case_number}")
                r += 1

            n_enf = 0
            ws = wb[ENFORCEMENT_SHEET]
            r = ENFORCEMENT_FIRST_ROW
            empty_run = 0
            seen_enf = 0
            while r < ENFORCEMENT_FIRST_ROW + max_rows_per_sheet:
                if not _row_nonempty_enf(ws, r):
                    empty_run += 1
                    if empty_run >= 5 and seen_enf > 0:
                        break
                    r += 1
                    continue
                empty_run = 0
                seen_enf += 1
                debtor_name = _txt(ws.cell(r, 2).value, 512)
                debtor_bin = _txt(ws.cell(r, 3).value, 32) or None
                cid = index.find_for_enforcement(debtor_name, debtor_bin)
                if cid is None:
                    case_id = uuid.uuid4()
                    case_number = f"PIR-IMP-ENF-{case_id.hex[:8].upper()}"
                    stub = Case(
                        id=case_id,
                        case_number=case_number,
                        court="—",
                        court_instance="first",
                        case_type="executive",
                        status="execution",
                        outcome="fully_satisfied",
                        party_role="plaintiff",
                        opponent_type="juridical",
                        plaintiff=company[:512],
                        defendant=debtor_name[:512] or "—",
                        company=(debtor_name[:512] or "—"),
                        company_bin=(debtor_bin or "")[:12],
                        dispute_category="other",
                        city="—",
                        judge="—",
                        filing_date=as_of,
                        next_hearing=None,
                        payment_deadline=None,
                        last_updated=as_of,
                        days_overdue=0,
                        risk_level="medium",
                        is_archived=False,
                        branch_id=stub_branch_id,
                        assigned_lawyer_id=fallback_id,
                    )
                    fin = CaseFinance(
                        case_id=case_id,
                        claim_amount=Decimal("0"),
                        main_debt=Decimal("0"),
                        state_fee=Decimal("0"),
                        fines=Decimal("0"),
                        rep_expenses=Decimal("0"),
                        other_costs=Decimal("0"),
                        paid_amount=Decimal("0"),
                        recovered_main=Decimal("0"),
                        recovered_fines=Decimal("0"),
                        recovered_state_fee=Decimal("0"),
                    )
                    lit = _build_case_litigation(case_id)
                    ev = CaseEvent(
                        id=uuid.uuid4(),
                        case_id=case_id,
                        action="Импорт ПИР: дело-заглушка для строки ИП без сопоставления",
                        user_label="import_pir_demo_from_xlsx.py",
                        detail=f"ИП xlsx row {r}, должник: {debtor_name[:200]}",
                        happened_at=datetime.now(timezone.utc),
                        user_id=None,
                    )
                    if not dry_run:
                        to_add.extend([stub, fin, lit, ev])
                    index.add(case_id, stub.defendant, stub.plaintiff, stub.company_bin)
                    cid = case_id
                    n_cases += 1
                    print(f"ИП row {r}: создано дело {case_number} (нет совпадения с реестром)")

                enf = EnforcementProceeding(
                    id=uuid.uuid4(),
                    case_id=cid,
                    debtor_name=debtor_name,
                    debtor_bin=debtor_bin,
                    court_act_summary=_txt(ws.cell(r, 4).value),
                    amount_total=_dec(ws.cell(r, 5).value),
                    amount_main=_dec(ws.cell(r, 6).value),
                    amount_fines=_dec(ws.cell(r, 7).value),
                    amount_fees=_dec(ws.cell(r, 8).value),
                    progress_notes=_txt(ws.cell(r, 9).value),
                    collected_amount=_dec(ws.cell(r, 10).value),
                    collection_doc_ref=_txt(ws.cell(r, 11).value),
                    balance_remaining=_dec(ws.cell(r, 12).value),
                    status_label=_txt(ws.cell(r, 13).value, 255),
                    recorded_at=as_of,
                )
                if not dry_run:
                    to_add.append(enf)
                n_enf += 1
                print(f"  → enforcement row {r}")
                r += 1

            n_debt = 0
            ws = wb[DEBT_SHEET]
            r = DEBT_FIRST_ROW
            empty_run = 0
            seen_debt = 0
            while r < DEBT_FIRST_ROW + max_rows_per_sheet:
                if not _row_nonempty_debt(ws, r):
                    empty_run += 1
                    if empty_run >= 5 and seen_debt > 0:
                        break
                    r += 1
                    continue
                empty_run = 0
                seen_debt += 1
                debtor_name = _txt(ws.cell(r, 2).value, 512)
                cid = index.find_for_enforcement(debtor_name, None)
                debt = DebtRecoveryEntry(
                    id=uuid.uuid4(),
                    case_id=cid,
                    counterparty_bin=None,
                    debtor_name=debtor_name,
                    debtor_status=_txt(ws.cell(r, 3).value, 255),
                    debt_amount=_dec(ws.cell(r, 4).value),
                    paid_amount=_dec(ws.cell(r, 6).value),
                    written_off_amount=_dec(ws.cell(r, 8).value),
                    work_summary=_txt(ws.cell(r, 5).value),
                    recorded_at=as_of,
                )
                if not dry_run:
                    to_add.append(debt)
                n_debt += 1
                print(f"дебиторка row {r}: {debtor_name[:50]}")
                r += 1

            if dry_run:
                print(f"Dry run: {n_cases} дел (включая заглушки ИП), {n_enf} ИП, {n_debt} строк дебиторки — не записано.")
            else:
                db.add_all(to_add)
                await db.commit()
                print(f"Готово: дел {n_cases}, исполнительных {n_enf}, дебиторки {n_debt}.")
    finally:
        wb.close()


def main() -> None:
    p = argparse.ArgumentParser(description="Импорт данных из xlsx ПИР в LegalHub")
    p.add_argument("--xlsx", type=Path, default=Path("templates/pir_report_2025_template.xlsx"))
    p.add_argument(
        "--limit",
        type=int,
        default=None,
        metavar="N",
        help="Максимум импортированных дел с листов истец+ответчик+3-лицо (не считая заглушек ИП). Без аргумента — без лимита.",
    )
    p.add_argument("--max-rows-per-sheet", type=int, default=5000, help="Стоп-лосс по строкам на каждый лист")
    p.add_argument("--dry-run", action="store_true")
    p.add_argument(
        "--purge-cases-first",
        action="store_true",
        help="Удалить все дела и строки дебиторки перед импортом (чтобы реестр совпадал только с файлом).",
    )
    p.add_argument("--as-of", type=lambda s: date.fromisoformat(s), default=None, help="Дата filing/last_updated/ИП (YYYY-MM-DD), по умолчанию сегодня")
    p.add_argument("--company", type=str, default=_DEFAULT_COMPANY)
    p.add_argument("--company-bin", type=str, default=_DEFAULT_BIN, dest="company_bin")
    args = p.parse_args()
    as_of = args.as_of or date.today()
    asyncio.run(
        main_async(
            args.xlsx.resolve(),
            limit=args.limit,
            max_rows_per_sheet=args.max_rows_per_sheet,
            dry_run=args.dry_run,
            as_of=as_of,
            company=args.company,
            company_bin=args.company_bin,
            purge_cases_first=args.purge_cases_first,
        )
    )


if __name__ == "__main__":
    main()
