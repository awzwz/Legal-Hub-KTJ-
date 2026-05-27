"""Find the case that's misclassified: currently fully_satisfied but should be denied."""
from __future__ import annotations
import re, sys
from decimal import Decimal
from pathlib import Path
from typing import Optional
from openpyxl import load_workbook

XLSX = Path(__file__).resolve().parents[1] / "data" / "pir_2025_final.xlsx"

def _to_decimal(v) -> Decimal:
    if v is None: return Decimal("0")
    if isinstance(v, (int, float, Decimal)): return Decimal(str(v))
    s = str(v).strip()
    if not s: return Decimal("0")
    s = s.replace("\xa0", "").replace(" ", "").replace("₸", "").replace(",", ".")
    try: return Decimal(s)
    except: return Decimal("0")

def _txt(v) -> str:
    return str(v).strip() if v else ""

SECTION_KEYS = (("закупках", "procurement"), ("закупок", "procurement"), ("перевозоч", "transportation"),
    ("трудовы", "labor"), ("госорган", "government"), ("иные спор", "other"), ("иные", "other"), ("медиатив", "mediation"))

def _detect_section_in_text(text) -> Optional[str]:
    if not isinstance(text, str): return None
    t = text.strip().lower()
    if not t or len(t) > 200: return None
    for key, cat in SECTION_KEYS:
        if key in t: return cat
    return None

def _detect_section_in_row(ws, r):
    sec = _detect_section_in_text(ws.cell(row=r, column=1).value)
    if sec: return sec
    return _detect_section_in_text(ws.cell(row=r, column=10).value)

def _is_section_only_row(ws, r):
    if _detect_section_in_text(ws.cell(row=r, column=1).value): return True
    c10_sec = _detect_section_in_text(ws.cell(row=r, column=10).value)
    if not c10_sec: return False
    a_val = ws.cell(row=r, column=1).value
    if isinstance(a_val, (int, float)) and int(a_val) > 0: return False
    return not _txt(ws.cell(row=r, column=2).value) and not _txt(ws.cell(row=r, column=5).value)

CMAP = {"lawyer": 2, "court": 3, "branch": 4, "counterparty": 5, "main_debt": 6, "fines": 7, "state_fee": 8,
    "claim_summary": 9, "judgment_first": 10, "judgment_appeal": 11, "judgment_cassation": 12,
    "recovered_main": 13, "recovered_fines": 14, "recovered_state_fee": 15}

_MEDIATION_RE = re.compile(r"(?:медиатив|мирово|утверждено\s+согла\w+.*?(?:медиаци|урегулировани))")
_APPEAL_INDICATOR_RE = re.compile(r"(?:апелляционн\w+|кассационн\w+|апеляционн\w+|апеляцон\w+)\s+(?:жалоб\w*|представлени\w*)")
_APPEAL_DISMISSED_RE = re.compile(r"(?:в\s+(?:удовлетворении\s+)?(?:апелляционн\w+|кассационн\w+|апеляционн\w+|апеляцон\w+)\s+(?:жалоб\w*|представлени\w*).*?отказ|отказ\w+\s+в\s+(?:удовлетворении\s+)?(?:апелляционн\w+|кассационн\w+|апеляционн\w+|апеляцон\w+)|оставлен[оаы]?\s+без\s+изменени|оставить\s+без\s+изменени|решени\w*\s+оставлен\w*\s+в\s+силе)")
_DENIED_RE = re.compile(r"(?:в\s+(?:удовлетворении\s+)?иск\w*.*?отказ|отказ\w+\s+в\s+иск|отказ\w+\s+от\s+иск|иск\w*.*?отозван|истц\w+\s+принят\w+\s+решени\w*\s+об\s+отказе|прекра[щт]\w+)")
_PARTIAL_RE = re.compile(r"(?:удовлетвор\w+\s+част|удовлетвор\w+\s+в\s+част|част\w+\s+удовлетвор)")
_RETURN_POS_RE = re.compile(r"возвра[щт]\w*.*?(?:оплат|исполнен|до\s+принятия|ввиду)")
_RETURN_NEG_RE = re.compile(r"возвра[щт]\w*")
_LEFT_NO_REVIEW_RE = re.compile(r"оставлен\w*\s+без\s+рассмотрени")

def _classify_instance(t):
    if not t: return "pending"
    tl = " ".join(t.lower().split())
    if _MEDIATION_RE.search(tl): return "settled"
    if _APPEAL_INDICATOR_RE.search(tl):
        if not (re.search(r"решени\w*\s+отменен", tl) or re.search(r"акт\w*\s+отменен", tl) or re.search(r"жалоб\w*\s+удовлетвор", tl) or re.search(r"представлени\w+\s+удовлетвор", tl)):
            return "appeal_dismissed"
    if _APPEAL_DISMISSED_RE.search(tl): return "appeal_dismissed"
    if _RETURN_POS_RE.search(tl): return "fully"
    if _LEFT_NO_REVIEW_RE.search(tl) and "удерж" in tl: return "fully"
    if _DENIED_RE.search(tl): return "denied"
    if _RETURN_NEG_RE.search(tl): return "denied"
    if _LEFT_NO_REVIEW_RE.search(tl): return "denied"
    if _PARTIAL_RE.search(tl): return "partial"
    if "удовлетвор" in tl: return "fully"
    return "pending"

_I2O = {"fully": "fully_satisfied", "partial": "partially_satisfied", "settled": "settled", "denied": "denied", "pending": "pending", "appeal_dismissed": "denied"}

def _infer_outcome(category, j1, ja, jc):
    if category == "mediation": return "settled"
    rc = _classify_instance(jc)
    if rc not in ("pending", "appeal_dismissed"): return _I2O[rc]
    ra = _classify_instance(ja)
    if ra not in ("pending", "appeal_dismissed"): return _I2O[ra]
    r1 = _classify_instance(j1)
    if r1 != "pending": return _I2O[r1]
    return "pending"

def main():
    print("Loading...")
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["истец"]
    
    # Parse
    cases = []
    cur_section = "procurement"
    cur_case = None
    summary_labels = ("предъявлено", "категория", "кол-во")
    
    for r in range(7, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = str(ws.cell(row=r, column=2).value or "").lower().strip()
        c = str(ws.cell(row=r, column=3).value or "").lower().strip()
        if b in summary_labels or c in summary_labels:
            break
        if _is_section_only_row(ws, r):
            sec = _detect_section_in_row(ws, r)
            if sec: cur_section = sec
            continue
        if isinstance(a, str) and not a.strip().isdigit():
            continue
        is_new = isinstance(a, (int, float)) and int(a) > 0
        cp_val = ws.cell(row=r, column=CMAP["counterparty"]).value
        is_cont = a is None and isinstance(cp_val, str) and cp_val.strip()
        if is_new:
            if cur_case: cases.append(cur_case)
            cur_case = {"section": cur_section, "src_row": r, "counterparty": _txt(ws.cell(row=r, column=CMAP["counterparty"]).value),
                "main_debt": _to_decimal(ws.cell(row=r, column=CMAP["main_debt"]).value),
                "fines": _to_decimal(ws.cell(row=r, column=CMAP["fines"]).value),
                "state_fee": _to_decimal(ws.cell(row=r, column=CMAP["state_fee"]).value),
                "judgment_first": _txt(ws.cell(row=r, column=CMAP["judgment_first"]).value),
                "judgment_appeal": _txt(ws.cell(row=r, column=CMAP["judgment_appeal"]).value),
                "judgment_cassation": _txt(ws.cell(row=r, column=CMAP["judgment_cassation"]).value),
                "recovered_main": _to_decimal(ws.cell(row=r, column=CMAP["recovered_main"]).value),
                "recovered_fines": _to_decimal(ws.cell(row=r, column=CMAP["recovered_fines"]).value),
                "recovered_state_fee": _to_decimal(ws.cell(row=r, column=CMAP["recovered_state_fee"]).value)}
        elif is_cont and cur_case:
            for f in ("main_debt", "fines", "state_fee", "recovered_main", "recovered_fines", "recovered_state_fee"):
                cur_case[f] += _to_decimal(ws.cell(row=r, column=CMAP[f]).value)
            for f in ("judgment_first", "judgment_appeal", "judgment_cassation"):
                extra = _txt(ws.cell(row=r, column=CMAP[f]).value)
                if extra and extra not in cur_case[f]:
                    cur_case[f] = (cur_case[f] + "\n" + extra).strip()
    if cur_case: cases.append(cur_case)
    
    # Find cases where recovered=0 but outcome is fully_satisfied or partially_satisfied (suspicious)
    print(f"\n=== Cases classified as SATISFIED but with ZERO recovered (suspicious) ===")
    for i, c in enumerate(cases):
        cat = c["section"]
        outcome = _infer_outcome(cat, c["judgment_first"], c["judgment_appeal"], c["judgment_cassation"])
        if outcome == "settled": cat = "mediation"
        recovered = c["recovered_main"] + c["recovered_fines"] + c["recovered_state_fee"]
        claim = c["main_debt"] + c["fines"] + c["state_fee"]
        
        if outcome in ("fully_satisfied", "partially_satisfied") and recovered == 0:
            print(f"\n  #{i+1} row={c['src_row']} cat={cat} outcome={outcome}")
            print(f"  counterparty: {c['counterparty']}")
            print(f"  claim: {float(claim):,.2f}  recovered: {float(recovered):,.2f}")
            print(f"  J1: {c['judgment_first'][:200]}")
            print(f"  JA: {c['judgment_appeal'][:200]}")
            print(f"  JC: {c['judgment_cassation'][:200]}")
    
    # Also show ALL non-mediation satisfied cases with detail
    print(f"\n=== ALL SATISFIED cases (non-mediation) for review ===")
    satisfied_count = 0
    for i, c in enumerate(cases):
        cat = c["section"]
        outcome = _infer_outcome(cat, c["judgment_first"], c["judgment_appeal"], c["judgment_cassation"])
        if outcome == "settled": cat = "mediation"
        if cat != "mediation" and outcome in ("fully_satisfied", "partially_satisfied"):
            satisfied_count += 1
            recovered = c["recovered_main"] + c["recovered_fines"] + c["recovered_state_fee"]
            claim = c["main_debt"] + c["fines"] + c["state_fee"]
            j1_cls = _classify_instance(c["judgment_first"])
            ja_cls = _classify_instance(c["judgment_appeal"])
            jc_cls = _classify_instance(c["judgment_cassation"])
            print(f"  #{satisfied_count:>2} (case {i+1}) row={c['src_row']:>3} {outcome:<22} claim={float(claim):>15,.2f} recov={float(recovered):>15,.2f} j1={j1_cls:<8} ja={ja_cls:<8} jc={jc_cls:<8} {c['counterparty'][:40]}")
    print(f"  Total satisfied (non-mediation): {satisfied_count}")
    
    # Show all denied
    print(f"\n=== ALL DENIED cases for review ===")
    denied_count = 0
    for i, c in enumerate(cases):
        cat = c["section"]
        outcome = _infer_outcome(cat, c["judgment_first"], c["judgment_appeal"], c["judgment_cassation"])
        if outcome == "settled": cat = "mediation"
        if cat != "mediation" and outcome in ("denied", "dismissed"):
            denied_count += 1
            claim = c["main_debt"] + c["fines"] + c["state_fee"]
            j1_cls = _classify_instance(c["judgment_first"])
            ja_cls = _classify_instance(c["judgment_appeal"])
            jc_cls = _classify_instance(c["judgment_cassation"])
            print(f"  #{denied_count:>2} (case {i+1}) row={c['src_row']:>3} {outcome:<22} claim={float(claim):>15,.2f} j1={j1_cls:<18} ja={ja_cls:<18} jc={jc_cls}")
            print(f"       J1: {c['judgment_first'][:150]}")
            if c['judgment_appeal']: print(f"       JA: {c['judgment_appeal'][:150]}")
            if c['judgment_cassation']: print(f"       JC: {c['judgment_cassation'][:150]}")
    print(f"  Total denied: {denied_count}")

if __name__ == "__main__":
    main()
