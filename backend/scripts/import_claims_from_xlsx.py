"""Импорт реестра претензий из xlsx-исходника в БД (`claims`).

Использование (внутри контейнера svc-legal):
    python /app/scripts/import_claims_from_xlsx.py --xlsx /data/claims_source.xlsx [--purge-first]

Алгоритм:
1. Читаем листы `претензии 2025` и `претензии 2026` (структура A..H).
2. Forward-fill для колонки B (Контрагент) — в исходнике пустые ячейки означают «то же, что выше».
3. Парсим дату «09.01.2025г.» → date.
4. Парсим сумму (пробелы, неразрывный пробел, запятая → точка) → Decimal.
5. Нормализуем статус по подстрокам:
     'удержано в безакцептном' → offset
     'перерасчёт' / 'перерасчет' → recalculation
     'не взыска'  → not_collected
     'взыскано'   → collected
   Деталь статуса (всё что после первой строки или в скобках) → status_detail.
6. Колонка H+ → notes.
7. Сматчим counterparty_bin со справочником `BIN_IIN_kompanii.xlsx` по нормализованному имени.
8. INSERT всех строк в одной транзакции.
"""
from __future__ import annotations

import argparse
import asyncio
import re
import sys
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

sys.path.insert(0, "/app")

from app.db.session import SessionLocal
from app.models import Claim


DATE_RE = re.compile(r"(\d{1,2})\.(\d{1,2})\.(\d{4})")
SHEETS = ("претензии 2025", "претензии 2026")
DATA_FIRST_ROW = 4

BIN_REGISTRY_PATH = Path("/data/BIN_IIN_kompanii.xlsx")


def _norm(text: str | None) -> str:
    if not text:
        return ""
    s = text.lower()
    s = re.sub(r"[\"«»“”\(\)\[\]\.,\-_/\\]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace("ё", "е")
    return re.sub(r"\s", "", s)


def _load_bin_registry() -> dict[str, str]:
    """Возвращает map normalized_name -> BIN."""
    out: dict[str, str] = {}
    if not BIN_REGISTRY_PATH.is_file():
        return out
    wb = load_workbook(BIN_REGISTRY_PATH, data_only=True)
    ws = wb["БИН компаний"]
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        bin_v = ws.cell(row=r, column=3).value
        if not name or not bin_v:
            continue
        bin_str = "".join(c for c in str(bin_v) if c.isdigit())
        if len(bin_str) == 12:
            out[_norm(str(name))] = bin_str
    return out


def _parse_date(value) -> Iterable | None:
    from datetime import date as _date, datetime as _datetime

    if value is None:
        return None
    if isinstance(value, _datetime):
        return value.date()
    if isinstance(value, _date):
        return value
    s = str(value).strip()
    m = DATE_RE.search(s)
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return _date(y, mo, d)
    except ValueError:
        return None


def _parse_amount(value) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    s = str(value).strip()
    if not s:
        return None
    # Убираем пробелы, неразрывный пробел, валютные символы
    s = s.replace("\xa0", "").replace(" ", "").replace("₸", "").replace("тенге", "")
    # Запятая = десятичный разделитель в исходнике
    s = s.replace(",", ".")
    try:
        return Decimal(s)
    except Exception:
        return None


def _normalize_status(raw: str | None) -> tuple[str, str | None]:
    """Возвращает (canonical_status, status_detail)."""
    if not raw:
        return "not_collected", None
    s = str(raw).strip()
    s_low = s.lower()
    if "удержан" in s_low and "безакцепт" in s_low:
        canonical = "offset"
    elif "перерасч" in s_low:
        canonical = "recalculation"
    elif "не взыска" in s_low:
        canonical = "not_collected"
    elif "взыскано" in s_low or "взыскан" in s_low:
        canonical = "collected"
    else:
        canonical = "not_collected"

    # Деталь: всё что после первой строки или внутри скобок после ключевого слова
    detail = None
    if "\n" in s:
        detail = s.split("\n", 1)[1].strip() or None
    else:
        m = re.search(r"\(([^)]+)\)", s)
        if m:
            detail = m.group(1).strip() or None
    return canonical, detail


async def _import_sheet(db: AsyncSession, sheet, bin_map: dict[str, str], claims: list[Claim]) -> int:
    """Импортирует один лист, добавляет Claim-объекты в claims-список. Возвращает кол-во строк."""
    cur_counterparty = ""
    added = 0
    for r in range(DATA_FIRST_ROW, sheet.max_row + 1):
        a = sheet.cell(row=r, column=1).value  # №
        b = sheet.cell(row=r, column=2).value  # Контрагент
        c = sheet.cell(row=r, column=3).value  # ИСХ.№
        d = sheet.cell(row=r, column=4).value  # Дата
        e = sheet.cell(row=r, column=5).value  # Сущность
        f = sheet.cell(row=r, column=6).value  # Сумма
        g = sheet.cell(row=r, column=7).value  # Статус
        h = sheet.cell(row=r, column=8).value  # Примечание

        # Пропускаем пустые строки и подзаголовки (типа «Итого»)
        if not c and not d and not e and not f:
            if b:
                cur_counterparty = str(b).strip()
            continue

        # Forward-fill контрагента
        if b:
            cur_counterparty = str(b).strip()

        outgoing = (str(c).strip() if c else "") or None
        claim_date = _parse_date(d)
        subject = (str(e).strip() if e else "") or None
        amount = _parse_amount(f)
        status, status_detail = _normalize_status(str(g) if g else None)
        notes = (str(h).strip() if h else None)

        # Минимально требуемые поля
        if not outgoing or claim_date is None or not subject or amount is None:
            print(f"  ! Пропуск строки R{r}: outgoing={outgoing}, date={claim_date}, subject={subject}, amount={amount}")
            continue

        # Матчинг БИН
        bin_value = bin_map.get(_norm(cur_counterparty))

        claims.append(Claim(
            counterparty_name=cur_counterparty or "(не указан)",
            counterparty_bin=bin_value,
            outgoing_number=outgoing,
            claim_date=claim_date,
            subject=subject,
            amount=amount,
            status=status,
            status_detail=status_detail,
            notes=notes,
            # ЦА по умолчанию (юрист перепривяжет вручную)
            branch_id=None,
            assigned_lawyer_id=None,
            case_id=None,
        ))
        added += 1
    return added


async def main_async(xlsx_path: str, purge_first: bool) -> None:
    src = Path(xlsx_path)
    if not src.is_file():
        raise SystemExit(f"Source xlsx not found: {src}")

    wb = load_workbook(src, data_only=True)
    bin_map = _load_bin_registry()
    print(f"BIN registry loaded: {len(bin_map)} entries")

    claims: list[Claim] = []
    async with SessionLocal() as db:  # type: AsyncSession
        if purge_first:
            res = await db.execute(delete(Claim))
            print(f"Deleted existing claims: {res.rowcount}")

        for sheet_name in SHEETS:
            if sheet_name not in wb.sheetnames:
                print(f"Sheet missing: {sheet_name}")
                continue
            ws = wb[sheet_name]
            print(f"\n=== {sheet_name} ===")
            count = await _import_sheet(db, ws, bin_map, claims)
            print(f"  parsed: {count} rows")

        # Подставим branch_id ЦА всем по умолчанию.
        # ID берётся из миграции normalize_branches.sql.
        for c in claims:
            c.branch_id = c.branch_id  # noqa: keeps NULL — пользователь распределит

        db.add_all(claims)
        await db.commit()
        print(f"\nImported total: {len(claims)} claims")
        # Сводка по статусам
        from collections import Counter
        cnt = Counter(c.status for c in claims)
        print("By status:")
        for k, v in cnt.items():
            print(f"  {k}: {v}")
        bin_count = sum(1 for c in claims if c.counterparty_bin)
        print(f"With BIN matched: {bin_count}/{len(claims)}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--purge-first", action="store_true", help="delete all existing claims before import")
    args = parser.parse_args()
    asyncio.run(main_async(args.xlsx, args.purge_first))


if __name__ == "__main__":
    main()
