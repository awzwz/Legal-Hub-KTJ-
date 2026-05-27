"""Делает пустой ПИР-шаблон из реального отчёта Q1 2026.

Логика:
1. Загружаем `Отчет ПИР за 1 кв.2026г.xlsx`.
2. На листах истец/ответчик/3-лицо: снимаем все merges ниже шапки и очищаем строки с данными.
3. Сохраняем как `backend/templates/pir_report_2025_template.xlsx` (имя сохраняем
   старое — pir_excel_fill.py его читает по этой константе).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

ROOT = Path(__file__).resolve().parents[2]
SOURCE = ROOT / "Отчет ПИР за 1 кв.2026г.xlsx"
TARGET = ROOT / "backend" / "templates" / "pir_report_2025_template.xlsx"

# Первая data-row для каждого листа (после шапки)
DATA_SHEETS = {
    "истец": 7,        # секция procurement в R7, дела с R8
    "ответчик": 6,     # секция procurement в R6, дела с R7
    "3-лицо ": 7,
}


def main() -> None:
    if not SOURCE.is_file():
        raise FileNotFoundError(f"Source missing: {SOURCE}")

    wb = load_workbook(SOURCE)

    for sheet_name, first_data_row in DATA_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # Снимаем все merges ниже шапки
        for cr in list(ws.merged_cells.ranges):
            if cr.min_row >= first_data_row:
                ws.unmerge_cells(str(cr))
        # Очищаем все строки данных
        for row in ws.iter_rows(min_row=first_data_row, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None

    TARGET.parent.mkdir(parents=True, exist_ok=True)
    wb.save(TARGET)
    print(f"Saved template: {TARGET}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
