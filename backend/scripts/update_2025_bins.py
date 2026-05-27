import asyncio
import os
import sys

sys.path.append(os.path.join(os.path.dirname(__file__), ".."))

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
import openpyxl

from app.db.session import SessionLocal
from app.models import Case

import re

def clean_name(name: str) -> str:
    if not name:
        return ""
    name = str(name).lower()
    # Normalize unicode spaces/dashes
    name = name.replace("–", "").replace("-", "").replace("—", "")
    # Remove corporate prefixes/suffixes
    forms = ["товарищество с ограниченной ответственностью", "акционерное общество", 
             "индивидуальный предприниматель", "республиканское государственное учреждение",
             "государственное учреждение", "тоо", "ао", "ип", "гу", "ргу", "оо", "чси", "пк", "филиал"]
    for f in forms:
        name = re.sub(rf'\b{f}\b', '', name)
    
    # Remove all non-alphanumeric characters (keep cyrillic and latin characters)
    name = re.sub(r'[^a-z0-9а-яёәғқңөұүһі]', '', name)
    return name.strip()

def load_bin_mapping(excel_path: str) -> dict:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    mapping = {}
    for row in range(1, ws.max_row + 1):
        name = ws.cell(row=row, column=2).value
        bin_val = ws.cell(row=row, column=3).value
        if name and bin_val:
            bin_str = str(bin_val).strip()
            bin_str = "".join(c for c in bin_str if c.isdigit())
            if len(bin_str) >= 10:
                mapping[clean_name(name)] = bin_str
    return mapping

async def update_bins(excel_paths: list[str]):
    mapping = {}
    for path in excel_paths:
        if os.path.exists(path):
            file_map = load_bin_mapping(path)
            mapping.update(file_map)
            print(f"Loaded {len(file_map)} mappings from {os.path.basename(path)}")
        else:
            print(f"Skipping {os.path.basename(path)}: file not found")
    
    async with SessionLocal() as db:
        result = await db.execute(select(Case))
        cases = result.scalars().all()
        
        updated_count = 0
        missing = set()
        
        for case in cases:
            if case.company:
                db_clean = clean_name(case.company)
                if not db_clean:
                    continue
                
                # Check exact match
                if db_clean in mapping:
                    case.company_bin = mapping[db_clean]
                    updated_count += 1
                else:
                    # Check substring match
                    found = False
                    for k, bin_val in mapping.items():
                        if len(db_clean) >= 4 and len(k) >= 4:
                            if db_clean in k or k in db_clean:
                                case.company_bin = bin_val
                                updated_count += 1
                                found = True
                                break
                    if not found:
                        missing.add(case.company.strip())
                    
        await db.commit()
        print(f"Updated {updated_count} cases with new BINs.")
        print(f"Could not find BIN for {len(missing)} distinct companies.")
        if missing:
            print("Examples of missing:", sorted(list(missing))[:10])

if __name__ == "__main__":
    excel_paths = [
        "/app/BIN_IIN_Отчет_ПИР_2025_final.xlsx",
        "/app/BIN_IIN_kompanii.xlsx",
        "/app/BIN_IIN_1Q2026_companies.xlsx"
    ]
    asyncio.run(update_bins(excel_paths))
