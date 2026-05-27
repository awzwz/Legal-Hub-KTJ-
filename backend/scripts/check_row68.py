"""Show full judgment text for row 68 (ЧСИ Бижаскынов А.К.)"""
from openpyxl import load_workbook
from pathlib import Path

XLSX = Path(__file__).resolve().parents[1] / "data" / "pir_2025_final.xlsx"
print("Loading...")
wb = load_workbook(XLSX, data_only=True)
ws = wb["истец"]

for r in (67, 68, 69, 70):
    print(f"\n=== Row {r} ===")
    for c in range(1, 20):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            print(f"  Col {c}: {repr(v)[:300]}")

