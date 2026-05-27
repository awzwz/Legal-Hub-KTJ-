"""Одноразовый фикс: пересчитать dispute_category для дел 2025 года
строго по секциям оригинального Excel (без content-detection override).

Маппинг: case_number 2025-NNN ↔ позиция в parsed-списке (определяется тем же
parse_sheet что и оригинальный импорт, поэтому индексы совпадают).
"""
from __future__ import annotations

import asyncio
import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

sys.path.insert(0, "/app")

from app.db.session import SessionLocal
from app.models import Case

# Используем тот же parse_sheet из import_pir_2025
sys.path.insert(0, "/app/scripts")
from import_pir_2025 import parse_sheet, ROLE_BY_SHEET  # noqa: E402


async def main_async(xlsx_path: str) -> None:
    src = Path(xlsx_path)
    if not src.is_file():
        raise SystemExit(f"Xlsx not found: {src}")
    wb = load_workbook(src, data_only=True)

    # Собираем все дела в том же порядке что делал import_pir_2025
    parsed_all = []
    for sheet_name, role in ROLE_BY_SHEET.items():
        if sheet_name not in wb.sheetnames:
            continue
        rows = parse_sheet(wb[sheet_name], role)
        parsed_all.extend(rows)
        print(f"  {sheet_name}: {len(rows)} cases")
    print(f"\nTotal parsed: {len(parsed_all)}")

    updates: dict[str, str] = {}
    for i, c in enumerate(parsed_all):
        case_number = f"2025-{i+1:03d}"
        # ВАЖНО: используем именно cur_section из parse_sheet,
        # БЕЗ content-detection override.
        section_category = c["dispute_category"]
        # Если outcome=settled, экспортный шаблон ожидает категорию mediation.
        # Импортный скрипт делал так — оставляем для согласованности.
        # (Но settled-кейсы обычно уже в секции mediation.)
        updates[case_number] = section_category

    async with SessionLocal() as db:
        # Получим текущее состояние
        from collections import Counter
        before = Counter()
        after = Counter()
        changes = []

        rows = (await db.execute(
            select(Case.case_number, Case.dispute_category, Case.party_role)
            .where(Case.case_number.like("2025-%"))
        )).all()
        cur_map: dict[str, tuple[str, str]] = {cn: (cat, role) for cn, cat, role in rows}

        for cn, new_cat in updates.items():
            old = cur_map.get(cn)
            if not old:
                continue
            old_cat, role = old
            before[(role, old_cat)] += 1
            after[(role, new_cat)] += 1
            if old_cat != new_cat:
                changes.append((cn, role, old_cat, new_cat))

        print(f"\nИзменений: {len(changes)}")
        print("Примеры (первые 10):")
        for cn, role, ocat, ncat in changes[:10]:
            print(f"  {cn} ({role}): {ocat} → {ncat}")

        # Применяем
        for cn, role, ocat, ncat in changes:
            await db.execute(
                update(Case).where(Case.case_number == cn).values(dispute_category=ncat)
            )
        await db.commit()
        print(f"\nApplied {len(changes)} updates.")

        # Сравнение
        print("\n=== ДО ===")
        for (role, cat), n in sorted(before.items()):
            print(f"  {role:12} {cat:14} {n}")
        print("\n=== ПОСЛЕ ===")
        for (role, cat), n in sorted(after.items()):
            print(f"  {role:12} {cat:14} {n}")


def main() -> None:
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default="/app/data/pir_2025_final.xlsx")
    args = parser.parse_args()
    asyncio.run(main_async(args.xlsx))


if __name__ == "__main__":
    main()
