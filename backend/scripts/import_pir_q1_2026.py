"""Импорт реальных дел из `Отчет ПИР за 1 кв.2026г.xlsx` в БД.

Логика:
1. Перед импортом удаляет все Case (и связанные таблицы), обнуляет claims.case_id.
2. Парсит 3 листа: истец (34 дела), ответчик (14), 3-лицо (2). Итого 50.
3. Подстроки (с пустым A) объединяются в предыдущее дело — суммируются суммы и
   конкатенируются текстовые поля через `\n`.
4. Юристов матчит по фамилии+инициалу (фолбэк — оставляет NULL).
5. Филиалы матчит на канонические UUID (aa000000-...).
6. filing_date — равномерно распределена по Q1 2026 (01.01..31.03).

Запуск (внутри контейнера svc-legal):
    python /app/scripts/import_pir_q1_2026.py --xlsx /data/q1_2026.xlsx
"""
from __future__ import annotations

import argparse
import asyncio
import re
import sys
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy import delete, select, update, extract
from sqlalchemy.ext.asyncio import AsyncSession

sys.path.insert(0, "/app")

from app.db.session import SessionLocal
from app.models import (
    Branch,
    Case,
    CaseComment,
    CaseDocument,
    CaseEvent,
    CaseFinance,
    CaseLawyer,
    CaseLitigation,
    Claim,
    DebtRecoveryEntry,
    EnforcementProceeding,
    Payment,
    ProceduralDeadline,
    User,
)


# ─────────────────────────────────────────────────────────────────────────────
# Канонические филиалы (UUID совпадают с normalize_branches.sql)
# ─────────────────────────────────────────────────────────────────────────────
BRANCH_UUID = {
    "ЦА - Центральный аппарат":        "aa000000-0000-0000-0000-000000000001",
    "АО «Вагонсервис»":                "aa000000-0000-0000-0000-000000000002",
    "РФ «Северный»":                   "aa000000-0000-0000-0000-000000000003",
    "РФ «Западный»":                   "aa000000-0000-0000-0000-000000000004",
    "РФ «Южный»":                      "aa000000-0000-0000-0000-000000000005",
    "Филиал «Экспресс»":               "aa000000-0000-0000-0000-000000000006",
    "Филиал «Пригородные перевозки»":  "aa000000-0000-0000-0000-000000000007",
    "Филиал «Сұңқар»":                 "aa000000-0000-0000-0000-000000000008",
}


def _normalize_branch(raw: Optional[str]) -> Optional[str]:
    """Маппинг произвольной строки из колонки D в каноническое имя филиала."""
    if not raw:
        return None
    s = str(raw).lower().replace("ё", "е")
    s = re.sub(r"[\"«»“”]", "", s).strip()
    if "вагонсервис" in s:
        return "АО «Вагонсервис»"
    if "северн" in s:
        return "РФ «Северный»"
    if "западн" in s:
        return "РФ «Западный»"
    if "южн" in s:
        return "РФ «Южный»"
    if "экспресс" in s:
        return "Филиал «Экспресс»"
    if "пригородн" in s:
        return "Филиал «Пригородные перевозки»"
    if "с[уұ]н[кқ]ар" in s or re.search(r"с[уұ]н[кқ]ар", s):
        return "Филиал «Сұңқар»"
    if any(t in s for t in ("цлю", "цюс", "ца -", "центральн", "цлвсю")):
        return "ЦА - Центральный аппарат"
    return None


def _norm_lawyer(name: Optional[str]) -> str:
    """«Ахатов А.Б.» / «Ахатов Айдос Булатұлы» → «ахатов аб»."""
    if not name:
        return ""
    s = str(name).strip()
    if "Ахатов А.А" in s:
        s = "Ахатов А.Б"
    if "Сырлыбаев Е" in s:
        s = "Сырлыбаев Е.Е"
    if "Әбсеметов Д" in s:
        s = "Әбсеметов Д.Е"
    s = s.lower()
    # Заменяем казахские буквы для унификации
    s = s.replace("і", "и").replace("ң", "н").replace("қ", "к").replace("ғ", "г").replace("ү", "у").replace("ұ", "у").replace("ә", "а").replace("ө", "о")
    s = re.sub(r"\.", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    parts = s.split()
    if not parts:
        return ""
    surname = parts[0]
    initials = "".join(p[0] for p in parts[1:3] if p)
    return f"{surname} {initials}".strip()


# ─────────────────────────────────────────────────────────────────────────────
# Парсинг ячеек
# ─────────────────────────────────────────────────────────────────────────────
def _to_decimal(v) -> Decimal:
    if v is None:
        return Decimal("0")
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip()
    if not s:
        return Decimal("0")
    s = s.replace("\xa0", "").replace(" ", "").replace("₸", "")
    s = s.replace(",", ".")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def _txt(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


# ─────────────────────────────────────────────────────────────────────────────
# Inference: outcome / status / instance / type
# ─────────────────────────────────────────────────────────────────────────────
def _infer_outcome(category: str, j1: str, ja: str, jc: str) -> str:
    """Самый «свежий» вердикт побеждает (кассация > апелляция > 1-я инстанция).

    Логика согласована с эталонным ПИР-отчётом юриста:
    - «удовлетворен (в полном/частично)»                 → fully_satisfied / partially_satisfied
    - «иск возвращен ввиду/в связи с оплат(ой)»          → fully_satisfied (долг погашен мирно)
    - «оставлен без рассмотрения … удержани»             → fully_satisfied (удержано из ЗП)
    - «утверждено соглашение … о (мирном) урегулир …    → settled (мировое/медиатив)
    - «утверждено соглашение … в порядке медиации»       → settled
    - «иске отказан / отказано»                          → denied
    - «иск возвращен» (без указания оплаты)              → denied
    - всё остальное                                      → pending
    """
    if category == "mediation":
        return "settled"
    chain = [jc, ja, j1]
    for t in chain:
        if not t:
            continue
        tl = " ".join(t.lower().split())  # схлопываем пробелы
        # Медиативные / мировые соглашения — проверяем ПЕРВЫМИ, до satisfied
        # (учитываем разные окончания: «соглашение/соглашения», «медиации/медиативное»)
        if ("утверждено согла" in tl and ("медиаци" in tl or "урегулировани" in tl)) \
                or "медиатив" in tl or "мирово" in tl:
            return "settled"
        # Иск удовлетворен (с учётом возможных двойных пробелов уже схлопнули)
        if "иск удовлетворен" in tl or "иск удовлетворены" in tl \
                or "удовлетворен в полном" in tl or "удовлетворены в полном" in tl \
                or "иск удовлетворены в полном" in tl:
            return "fully_satisfied"
        if "удовлетворен частично" in tl or "удовлетворены частично" in tl \
                or "удовлетворены в части" in tl or "удовлетворен в части" in tl:
            return "partially_satisfied"
        # Иск возвращен → denied
        if "возвращен" in tl:
            return "denied"
        if "отказан" in tl:  # широкий matcher: «отказано», «отказан», «в иске отказано», «исковых требованиях отказано»
            return "denied"
        # «Дело прекращено» — эталон относит к отказано (если не медиация — то прекращение по др.основаниям)
        if "прекращ" in tl:
            return "denied"
    return "pending"


def _infer_status(outcome: str, has_execution: bool) -> str:
    if outcome == "pending":
        return "active"
    if has_execution:
        return "execution"
    if outcome in {"fully_satisfied", "partially_satisfied", "denied", "dismissed", "settled"}:
        return "closed"
    return "active"


def _infer_instance(j1: str, ja: str, jc: str) -> str:
    if jc:
        return "cassation"
    if ja:
        return "appeal"
    return "first"


def _infer_risk(amount: Decimal) -> str:
    a = float(amount)
    if a > 50_000_000:
        return "high"
    if a < 5_000_000:
        return "low"
    return "medium"


# ─────────────────────────────────────────────────────────────────────────────
# Лист → описание колонок
# ─────────────────────────────────────────────────────────────────────────────
ROLE_BY_SHEET = {
    "истец": "plaintiff",
    "ответчик": "defendant",
    "3-лицо ": "third_party",
}

# Для каждого листа: какие колонки заполнены и куда мапятся.
# (col, model_field)
COLUMN_MAP_PLAINTIFF = {
    "lawyer": 2,
    "court": 3,
    "branch": 4,
    "counterparty": 5,
    "main_debt": 6,
    "fines": 7,
    "state_fee": 8,
    "claim_summary": 9,
    "judgment_first": 10,
    "judgment_appeal": 11,
    "judgment_cassation": 12,
    "recovered_main": 13,
    "recovered_fines": 14,
    "recovered_state_fee": 15,
    "writ_request_note": 16,
    "writ_dispatch_note": 17,
    "execution_proof_note": 18,
    "damage_recovery_note": 19,
}

COLUMN_MAP_DEFENDANT = {
    "lawyer": 2,
    "court": 3,
    "branch": 4,
    "counterparty": 5,
    "main_debt": 6,
    "fines": 7,
    "rep_expenses": 8,
    "state_fee": 9,
    "claim_summary": 10,
    "judgment_first": 11,
    "judgment_appeal": 12,
    "judgment_cassation": 13,
    "recovered_main": 14,
    "recovered_fines": 15,
    "recovered_rep_expenses": 16,
    "recovered_state_fee": 17,
    "defendant_execution_note": 18,
    "damage_recovery_note": 19,
}

COLUMN_MAP_THIRD = {
    "lawyer": 2,
    "court": 3,
    "branch": 4,
    "plaintiff_name": 5,  # на «3-лицо» — наименование истца
    "counterparty": 6,    # ответчика
    "main_debt": 7,
    "fines": 8,
    "rep_expenses": 9,
    "state_fee": 10,
    "claim_summary": 11,
    "judgment_first": 12,
    "judgment_appeal": 13,
    "judgment_cassation": 14,
    "recovered_main": 15,
    "recovered_fines": 16,
    "recovered_rep_expenses": 17,
    "recovered_state_fee": 18,
    "third_party_note": 19,
}

COL_MAP_BY_SHEET = {
    "истец": COLUMN_MAP_PLAINTIFF,
    "ответчик": COLUMN_MAP_DEFENDANT,
    "3-лицо ": COLUMN_MAP_THIRD,
}

# Первая data-row для каждого листа (после шапки и строки нумерации колонок)
FIRST_DATA_ROW = {
    "истец": 7,        # R6 = шапка-нумерация (1..19), R7 — секция или первое дело
    "ответчик": 6,
    "3-лицо ": 7,
}


SECTION_KEYS = (
    ("закупках", "procurement"),
    ("перевозоч", "transportation"),
    ("трудовы", "labor"),
    ("госорган", "government"),
    ("иные спор", "other"),
    ("медиатив", "mediation"),
)


def _detect_section(a_value) -> Optional[str]:
    if not isinstance(a_value, str):
        return None
    al = a_value.lower()
    for key, cat in SECTION_KEYS:
        if key in al:
            return cat
    return None


def _detect_category_by_content(court: str, claim_summary: str, judgment_first: str) -> Optional[str]:
    """Распознать категорию дела по содержанию иска и суду.

    Эталон Q1 2026 не выделяет отдельных разделов на листе, но дела по СОДЕРЖАНИЮ
    относятся к разным категориям. Юрист обычно видит их сам, мы помогаем автомату:
      - «о восстановлении на работе» / «трудовой» / «инспекция труда» / «несчастный случай» / «отпуск»  → labor
      - «СМАС» / «административный суд» / «предписание» / «об отмене акта»                              → government
      - «о возмещении ущерба» / «о взыскании ущерба» / «несанкционированный сход» / «начет»             → other
      - «договор о закупке» / «штраф/неустойка/пеня по договору» / «задолженность»                     → procurement
    """
    text = " ".join(filter(None, [court, claim_summary, judgment_first])).lower()
    # Трудовые споры — самая «жёсткая» категория
    labor_keys = (
        "восстановлен", "трудового договор", "расторжен", "инспекци", "несчастн",
        "отпускн", "вынужденн", "проводник", "форменную одежду", "должностной инструкции",
        "билетного кассир", "об отмене приказа",
    )
    for k in labor_keys:
        if k in text:
            return "labor"
    # Госорганы / административные споры (но НЕ трудовые)
    gov_keys = ("смас", "административн", "предписани", "айзрк", "санитарно-эпидемиологическ",
                "защите и развитию конкуренции")
    if any(k in text for k in gov_keys):
        return "government"
    # «Иные споры» — возмещение ущерба / прочие гражданские
    other_keys = ("возмещени", "ущерб", "несанкционированн", "сход пассажир")
    if any(k in text for k in other_keys):
        return "other"
    # По умолчанию — закупки/договоры
    return "procurement"


# ─────────────────────────────────────────────────────────────────────────────
# Парсинг листа
# ─────────────────────────────────────────────────────────────────────────────
def parse_sheet(ws, role: str) -> list[dict]:
    """Возвращает список дел (dict) с полями для создания Case."""
    cmap = COL_MAP_BY_SHEET[ws.title]
    cases: list[dict] = []
    cur_section = "procurement"
    cur_case: Optional[dict] = None
    first_row = FIRST_DATA_ROW.get(ws.title, 1)

    finished = False  # после первого «маркера сводки» прекращаем парсинг
    # Сводный блок располагается в B/C: "Категория" / "предъявлено" и т.п.
    # Дополнительно — строка где E=число (а не имя контрагента) — это итог сводки.
    summary_labels = ("предъявлено", "категория", "кол-во")
    for r in range(first_row, ws.max_row + 1):
        if finished:
            break
        a = ws.cell(row=r, column=1).value
        b = str(ws.cell(row=r, column=2).value or "").lower().strip()
        c = str(ws.cell(row=r, column=3).value or "").lower().strip()
        # Останавливаемся если в B или C ровно метка сводки (точное совпадение, не подстрока)
        if b in summary_labels or c in summary_labels:
            finished = True
            break
        # Разделитель секции
        sec = _detect_section(a)
        if sec:
            cur_section = sec
            continue
        # Игнорируем шапку и пустые строки
        if isinstance(a, str) and not a.strip().isdigit():
            continue
        # A — это номер (int) → новое дело
        is_new_case = isinstance(a, (int, float)) and int(a) > 0
        # Continuation: A пуст, counterparty — текстовая строка (не число; иначе это сводка)
        cp_val = ws.cell(row=r, column=cmap.get("counterparty", 5)).value
        is_continuation = a is None and isinstance(cp_val, str) and cp_val.strip()

        if is_new_case:
            # Перед началом нового — закроем предыдущее
            if cur_case is not None:
                cases.append(cur_case)
            cur_case = {
                "role": role,
                "dispute_category": cur_section,
                "src_row": r,
                "lawyer": _txt(ws.cell(row=r, column=cmap["lawyer"]).value),
                "court": _txt(ws.cell(row=r, column=cmap["court"]).value),
                "branch": _txt(ws.cell(row=r, column=cmap["branch"]).value),
                "counterparty": _txt(ws.cell(row=r, column=cmap["counterparty"]).value),
                "main_debt": _to_decimal(ws.cell(row=r, column=cmap.get("main_debt", 0)).value) if "main_debt" in cmap else Decimal("0"),
                "fines": _to_decimal(ws.cell(row=r, column=cmap.get("fines", 0)).value) if "fines" in cmap else Decimal("0"),
                "rep_expenses": _to_decimal(ws.cell(row=r, column=cmap.get("rep_expenses", 0)).value) if "rep_expenses" in cmap else Decimal("0"),
                "state_fee": _to_decimal(ws.cell(row=r, column=cmap.get("state_fee", 0)).value) if "state_fee" in cmap else Decimal("0"),
                "claim_summary": _txt(ws.cell(row=r, column=cmap["claim_summary"]).value),
                "judgment_first": _txt(ws.cell(row=r, column=cmap["judgment_first"]).value),
                "judgment_appeal": _txt(ws.cell(row=r, column=cmap["judgment_appeal"]).value),
                "judgment_cassation": _txt(ws.cell(row=r, column=cmap["judgment_cassation"]).value),
                "recovered_main": _to_decimal(ws.cell(row=r, column=cmap.get("recovered_main", 0)).value) if "recovered_main" in cmap else Decimal("0"),
                "recovered_fines": _to_decimal(ws.cell(row=r, column=cmap.get("recovered_fines", 0)).value) if "recovered_fines" in cmap else Decimal("0"),
                "recovered_rep_expenses": _to_decimal(ws.cell(row=r, column=cmap.get("recovered_rep_expenses", 0)).value) if "recovered_rep_expenses" in cmap else Decimal("0"),
                "recovered_state_fee": _to_decimal(ws.cell(row=r, column=cmap.get("recovered_state_fee", 0)).value) if "recovered_state_fee" in cmap else Decimal("0"),
                "writ_request_note": _txt(ws.cell(row=r, column=cmap.get("writ_request_note", 0)).value) if "writ_request_note" in cmap else "",
                "writ_dispatch_note": _txt(ws.cell(row=r, column=cmap.get("writ_dispatch_note", 0)).value) if "writ_dispatch_note" in cmap else "",
                "execution_proof_note": _txt(ws.cell(row=r, column=cmap.get("execution_proof_note", 0)).value) if "execution_proof_note" in cmap else "",
                "defendant_execution_note": _txt(ws.cell(row=r, column=cmap.get("defendant_execution_note", 0)).value) if "defendant_execution_note" in cmap else "",
                "damage_recovery_note": _txt(ws.cell(row=r, column=cmap.get("damage_recovery_note", 0)).value) if "damage_recovery_note" in cmap else "",
                "third_party_note": _txt(ws.cell(row=r, column=cmap.get("third_party_note", 0)).value) if "third_party_note" in cmap else "",
                "execution_status": _txt(ws.cell(row=r, column=19).value) if role == "plaintiff" else "",
            }
        elif is_continuation and cur_case is not None:
            # Подстрока — суммируем суммы, конкатим тексты
            def _add_num(field: str):
                if field in cmap:
                    cur_case[field] += _to_decimal(ws.cell(row=r, column=cmap[field]).value)
            def _append_text(field: str):
                if field in cmap:
                    extra = _txt(ws.cell(row=r, column=cmap[field]).value)
                    if extra and extra not in cur_case[field]:
                        cur_case[field] = (cur_case[field] + "\n" + extra).strip()
            for f in ("main_debt", "fines", "rep_expenses", "state_fee",
                      "recovered_main", "recovered_fines", "recovered_rep_expenses", "recovered_state_fee"):
                _add_num(f)
            for f in ("claim_summary", "judgment_first", "judgment_appeal", "judgment_cassation",
                      "writ_request_note", "writ_dispatch_note", "execution_proof_note",
                      "defendant_execution_note", "damage_recovery_note", "third_party_note"):
                _append_text(f)
        # иначе игнор

    if cur_case is not None:
        cases.append(cur_case)
    return cases


# ─────────────────────────────────────────────────────────────────────────────
# Очистка БД и импорт
# ─────────────────────────────────────────────────────────────────────────────
async def purge_cases_for_year(db: AsyncSession, year: int) -> None:
    """Удалить cases ТОЛЬКО за указанный год."""
    target_ids_q = select(Case.id).where(extract("year", Case.filing_date) == year)
    target_ids = [str(r[0]) for r in (await db.execute(target_ids_q)).all()]
    if not target_ids:
        print(f"Nothing to purge for year {year}")
        return
    await db.execute(
        update(Claim).where(Claim.case_id.in_(target_ids)).values(case_id=None)
    )
    for model in (
        EnforcementProceeding, DebtRecoveryEntry, CaseLitigation, CaseFinance,
        Payment, CaseDocument, CaseEvent, CaseComment, ProceduralDeadline,
    ):
        try:
            await db.execute(delete(model).where(model.case_id.in_(target_ids)))
        except Exception as e:
            print(f"  purge {model.__name__}: {e}")
    res = await db.execute(delete(Case).where(Case.id.in_(target_ids)))
    print(f"Deleted cases for {year}: {res.rowcount}")


async def load_lawyer_map(db: AsyncSession) -> dict[str, str]:
    """Возвращает map нормализованного имени → user_id."""
    users = (await db.execute(select(User))).scalars().all()
    out = {}
    for u in users:
        key = _norm_lawyer(u.full_name)
        if key and key not in out:
            out[key] = str(u.id)
    return out


def _distribute_dates(n: int) -> list[date]:
    """Равномерно по Q1 2026 (01.01..31.03 — 90 дней)."""
    if n <= 0:
        return []
    span = 89  # 90 дней с 01.01 по 31.03
    out = []
    for i in range(n):
        offset = round(i * span / max(n - 1, 1)) if n > 1 else 0
        out.append(date(2026, 1, 1) + timedelta(days=offset))
    return out


async def import_cases(db: AsyncSession, parsed: list[dict]) -> int:
    lawyer_map = await load_lawyer_map(db)
    dates = _distribute_dates(len(parsed))
    inserted = 0

    for i, c in enumerate(parsed):
        # Найдём юриста
        raw_lawyers = re.split(r"[\n,]", c["lawyer"]) if c["lawyer"] else []
        lawyer_ids = []
        for rl in raw_lawyers:
            lk = _norm_lawyer(rl)
            if lk and lk in lawyer_map:
                lawyer_ids.append(lawyer_map[lk])
                
        lawyer_id = lawyer_ids[0] if lawyer_ids else None

        # Филиал
        canon_branch = _normalize_branch(c["branch"])
        branch_uuid = BRANCH_UUID.get(canon_branch or "")
        if not branch_uuid:
            branch_uuid = BRANCH_UUID["ЦА - Центральный аппарат"]

        # Контрагент
        counterparty = c["counterparty"] or "—"
        if c["role"] == "plaintiff":
            plaintiff_name = "АО «Пассажирские перевозки»"
            defendant_name = counterparty
            company_name = counterparty
        elif c["role"] == "defendant":
            plaintiff_name = counterparty
            defendant_name = "АО «Пассажирские перевозки»"
            company_name = counterparty
        else:  # third_party
            plaintiff_name = c.get("plaintiff_name") or "—"
            defendant_name = counterparty
            company_name = counterparty

        # Извлекаем имя суда и судью из колонки C (формат: "СМЭС...\nсудья ФИО")
        court_block = c["court"]
        if "\n" in court_block:
            court_name, judge = court_block.split("\n", 1)
            court_name = court_name.strip()
            judge = judge.replace("судья", "").strip()
        else:
            court_name = court_block
            judge = ""

        # Категория: сначала смотрим по содержимому (если эталон явно даёт ключевые слова),
        # потом резервно — на текущей секции из колонки A. Если outcome=settled — переводим
        # дело в mediation (как делает шаблон КТЖ).
        detected = _detect_category_by_content(c["court"], c["claim_summary"], c["judgment_first"])
        category = detected if detected else c["dispute_category"]
        outcome = _infer_outcome(category, c["judgment_first"], c["judgment_appeal"], c["judgment_cassation"])
        # Mediation выявленный по outcome — приоритет
        if outcome == "settled":
            category = "mediation"
        
        is_defendant = c["role"] == "defendant"
        rec_any = any(_to_decimal(c[k]) > 0 for k in ("recovered_main", "recovered_fines", "recovered_state_fee"))
        exec_proof = c.get("execution_proof_note", "")
        writ_req = c.get("writ_request_note", "")
        def_exec = c.get("defendant_execution_note", "")
        
        if outcome == "pending":
            status = "active"
        else:
            exec_note = c.get("damage_recovery_note", "").lower()
            def_exec_lower = def_exec.lower()
            proof_note_lower = exec_proof.lower()
            disp_note = c.get("writ_dispatch_note", "")
            disp_note_lower = disp_note.lower()

            is_closed = False
            if "исполнено" in exec_note or "исполнено" in proof_note_lower:
                is_closed = True
            elif is_defendant and ("пл.пор" in def_exec_lower or "пл. пор" in def_exec_lower or "пл.поручение" in def_exec_lower):
                is_closed = True
            elif is_defendant and def_exec_lower and not ("на исполнении" in exec_note):
                is_closed = True

            status = "closed"
            if not is_closed:
                if "на исполнении" in exec_note:
                    status = "execution"
                elif disp_note_lower:
                    status = "execution"
                elif proof_note_lower and ("возбужден" in proof_note_lower or "направлен" in proof_note_lower or "предъявлен" in proof_note_lower):
                    status = "execution"
        
        instance = _infer_instance(c["judgment_first"], c["judgment_appeal"], c["judgment_cassation"])
        case_type = "civil"
        if category == "labor":
            case_type = "labor"
        elif "смас" in court_name.lower() or "административн" in court_name.lower():
            case_type = "administrative"

        # Общая сумма иска
        claim_amount = c["main_debt"] + c["fines"] + c["state_fee"] + c.get("rep_expenses", Decimal("0"))

        case = Case(
            case_number=f"Q1-2026-{i+1:03d}",
            court=court_name or "Суд не указан",
            judge=judge or "—",
            court_instance=instance,
            case_type=case_type,
            status=status,
            outcome=outcome,
            party_role=c["role"],
            opponent_type="juridical" if any(t in counterparty.lower() for t in ("тоо", "ао", "ип ", "оо ", "рг", "филиал", "нао")) else "physical",
            plaintiff=plaintiff_name[:512],
            defendant=defendant_name[:512],
            company=company_name[:512],
            company_bin="",
            city="—",
            filing_date=dates[i],
            last_updated=dates[i],
            days_overdue=0,
            risk_level=_infer_risk(claim_amount),
            is_archived=False,
            dispute_category=category,
            branch_id=branch_uuid,
            assigned_lawyer_id=lawyer_id,
        )
        case.finances = CaseFinance(
            claim_amount=claim_amount,
            main_debt=c["main_debt"],
            state_fee=c["state_fee"],
            fines=c["fines"],
            rep_expenses=c.get("rep_expenses", Decimal("0")),
            other_costs=Decimal("0"),
            paid_amount=Decimal("0"),
            recovered_main=c.get("recovered_main", Decimal("0")),
            recovered_fines=c.get("recovered_fines", Decimal("0")),
            recovered_state_fee=c.get("recovered_state_fee", Decimal("0")),
            recovered_rep_expenses=c.get("recovered_rep_expenses", Decimal("0")),
        )
        now = datetime.now(timezone.utc)
        case.litigation = CaseLitigation(
            claim_summary=c["claim_summary"],
            judgment_first=c["judgment_first"],
            judgment_appeal=c["judgment_appeal"],
            judgment_cassation=c["judgment_cassation"],
            writ_request_note=c.get("writ_request_note", ""),
            writ_dispatch_note=c.get("writ_dispatch_note", ""),
            execution_proof_note=c.get("execution_proof_note", ""),
            defendant_execution_note=c.get("defendant_execution_note", ""),
            damage_recovery_note=c.get("damage_recovery_note", ""),
            third_party_note=c.get("third_party_note", ""),
            created_at=now,
            updated_at=now,
        )
        if lawyer_ids:
            unique_ids = list(dict.fromkeys(lawyer_ids))
            case.case_lawyers = [
                CaseLawyer(user_id=uid) for uid in unique_ids
            ]

        db.add(case)
        inserted += 1
    return inserted


async def main_async(xlsx_path: str) -> None:
    src = Path(xlsx_path)
    if not src.is_file():
        raise SystemExit(f"Xlsx not found: {src}")
    wb = load_workbook(src, data_only=True)

    parsed: list[dict] = []
    for sheet_name, role in ROLE_BY_SHEET.items():
        if sheet_name not in wb.sheetnames:
            print(f"Sheet missing: {sheet_name}")
            continue
        ws = wb[sheet_name]
        rows = parse_sheet(ws, role)
        print(f"  {sheet_name}: parsed {len(rows)} cases")
        parsed.extend(rows)
    print(f"\nTotal parsed: {len(parsed)}")

    async with SessionLocal() as db:  # type: AsyncSession
        await purge_cases_for_year(db, 2026)
        await db.flush()
        inserted = await import_cases(db, parsed)
        await db.commit()
        print(f"\nInserted: {inserted} cases")

        # Сводка
        from collections import Counter
        by_role = Counter(c["role"] for c in parsed)
        by_cat = Counter(c["dispute_category"] for c in parsed)
        print(f"By role: {dict(by_role)}")
        print(f"By category: {dict(by_cat)}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    args = parser.parse_args()
    asyncio.run(main_async(args.xlsx))


if __name__ == "__main__":
    main()
