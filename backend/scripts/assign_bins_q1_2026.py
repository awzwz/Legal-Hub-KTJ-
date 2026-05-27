"""Разнести БИНы из `BIN_IIN_1Q2026_companies.xlsx` по делам Q1 2026.

Логика та же что в assign_bins.py — нормализованное матчинг по имени компании.
Для физлиц (где БИН отсутствует) — company_bin остаётся пустым.

Запуск (внутри контейнера svc-legal):
    python /app/scripts/assign_bins_q1_2026.py --xlsx /data/bins_q1_2026.xlsx
"""
from __future__ import annotations

import argparse
import asyncio
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

sys.path.insert(0, "/app")

from app.db.session import SessionLocal
from app.models import Case


def _norm(s: str | None) -> str:
    if not s:
        return ""
    s = s.lower()
    s = re.sub(r"[\"«»“”\(\)\[\]\.,\-_/\\]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace("ё", "е")
    return re.sub(r"\s", "", s)


def load_bin_map(xlsx_path: Path) -> dict[str, str]:
    """Map normalized company name → BIN (или ИИН для ИП)."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["БИН_ИИН"]
    # Колонки: A=№, B=Компания из файла, C=БИН/ИИН, ...
    out: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        bin_v = ws.cell(row=r, column=3).value
        if not name or not bin_v:
            continue
        bin_str = "".join(c for c in str(bin_v) if c.isdigit())
        if len(bin_str) != 12:
            continue
        key = _norm(str(name))
        if key and key not in out:
            out[key] = bin_str
    return out


async def main_async(xlsx_path: str) -> None:
    src = Path(xlsx_path)
    if not src.is_file():
        raise SystemExit(f"Xlsx not found: {src}")
    bin_map = load_bin_map(src)
    print(f"Loaded {len(bin_map)} (company → BIN) mappings")

    matched = 0
    unmatched = 0
    unmatched_names: list[str] = []

    async with SessionLocal() as db:  # type: AsyncSession
        cases = (await db.execute(select(Case))).scalars().all()
        for case in cases:
            company = case.company or ""
            key = _norm(company)
            # Точный матч
            bin_value = bin_map.get(key)
            # Подстрочный матч (если key содержит ключ из мапы или наоборот)
            if not bin_value:
                for k, v in bin_map.items():
                    if k and (k in key or key in k):
                        bin_value = v
                        break
            if bin_value:
                case.company_bin = bin_value
                matched += 1
            else:
                unmatched += 1
                unmatched_names.append(company)
        await db.commit()

    print(f"\nMatched: {matched}")
    print(f"Unmatched: {unmatched}")
    if unmatched_names:
        print("Unmatched companies (физлица или нет в справочнике):")
        for n in sorted(set(unmatched_names))[:30]:
            print(f"  - {n!r}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    args = parser.parse_args()
    asyncio.run(main_async(args.xlsx))


if __name__ == "__main__":
    main()
