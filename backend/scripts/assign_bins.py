"""Распределить компании из BIN_IIN_kompanii.xlsx по cases.

- Если company уже похож на имя из списка БИНов → проставить company_bin и канонизировать company.
- Если company = «АО Пассажирские перевозки» (мы ответчики) → назначить псевдо-случайного истца из списка
  и обновить plaintiff/company/company_bin одновременно.
"""
from __future__ import annotations

import asyncio
import random
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

sys.path.insert(0, "/app")

from app.db.session import SessionLocal
from app.models import Case

OWN_COMPANY_PATTERNS = (
    "пассажирскиеперевозки",
    "ао пассажирские",
)

BIN_XLSX = Path("/data/BIN_IIN_kompanii.xlsx")


def _norm(s: str | None) -> str:
    if not s:
        return ""
    s = s.lower()
    s = re.sub(r"[\"«»“”\(\)\[\]\.,\-_/\\]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace("ё", "е")
    s = re.sub(r"\s", "", s)
    return s


def load_bin_list() -> list[tuple[str, str]]:
    wb = load_workbook(BIN_XLSX, data_only=True)
    ws = wb["БИН компаний"]
    out: list[tuple[str, str]] = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        bin_v = ws.cell(row=r, column=3).value
        if name and bin_v:
            bin_str = str(bin_v).strip()
            if len(bin_str) == 12 and bin_str.isdigit():
                out.append((str(name).strip(), bin_str))
    return out


async def main() -> None:
    bin_list = load_bin_list()
    print(f"Loaded {len(bin_list)} canonical (company, BIN) pairs")

    name_lookup: dict[str, tuple[str, str]] = {}
    for canon_name, bin_v in bin_list:
        name_lookup[_norm(canon_name)] = (canon_name, bin_v)

    rng = random.Random(42)
    async with SessionLocal() as db:  # type: AsyncSession
        rows = (await db.execute(select(Case.id, Case.company, Case.plaintiff, Case.defendant, Case.party_role))).all()

        matched = 0
        substituted = 0
        unmatched: list[tuple[str, str]] = []

        for case_id, company, plaintiff, defendant, party_role in rows:
            company_norm = _norm(company)
            is_own = any(p in company_norm for p in (_norm("Пассажирские перевозки"),))

            if is_own:
                # Назначаем случайного истца из списка БИНов
                canon_name, new_bin = rng.choice(bin_list)
                await db.execute(
                    update(Case).where(Case.id == case_id).values(
                        plaintiff=canon_name,
                        company=canon_name,
                        company_bin=new_bin,
                    )
                )
                substituted += 1
            else:
                hit = name_lookup.get(company_norm)
                if hit is None:
                    for nk, val in name_lookup.items():
                        if nk and (nk in company_norm or company_norm in nk):
                            hit = val
                            break
                if hit:
                    canon_name, new_bin = hit
                    new_defendant = defendant
                    new_plaintiff = plaintiff
                    if party_role == "plaintiff":
                        new_defendant = canon_name
                    elif party_role == "defendant":
                        new_plaintiff = canon_name
                    await db.execute(
                        update(Case).where(Case.id == case_id).values(
                            plaintiff=new_plaintiff,
                            defendant=new_defendant,
                            company=canon_name,
                            company_bin=new_bin,
                        )
                    )
                    matched += 1
                else:
                    # Не нашли — назначаем случайный
                    canon_name, new_bin = rng.choice(bin_list)
                    new_defendant = defendant
                    new_plaintiff = plaintiff
                    if party_role == "plaintiff":
                        new_defendant = canon_name
                    elif party_role == "defendant":
                        new_plaintiff = canon_name
                    await db.execute(
                        update(Case).where(Case.id == case_id).values(
                            plaintiff=new_plaintiff,
                            defendant=new_defendant,
                            company=canon_name,
                            company_bin=new_bin,
                        )
                    )
                    unmatched.append((str(company), canon_name))

        await db.commit()

    print(f"Exact/substring matches: {matched}")
    print(f"«Своя компания» substituted: {substituted}")
    print(f"Unmatched (randomly assigned): {len(unmatched)}")
    if unmatched:
        print("First 20 unmatched:")
        for orig, new in unmatched[:20]:
            print(f"  {orig!r} -> {new!r}")


if __name__ == "__main__":
    asyncio.run(main())
