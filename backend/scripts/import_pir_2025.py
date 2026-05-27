"""Импорт реальных дел из ПИР-отчёта за 2025 год в БД.

Отличия от ``import_pir_q1_2026.py``:
1. **Year-scoped purge** — удаляем ТОЛЬКО дела с filing_date в 2025, оставляя 2026.
   Это позволяет хранить оба года рядом и переключаться через год-селектор на дашборде.
2. **Section-header detection** — в 2025-файле часть заголовков (Трудовые / Перевозочный /
   Иные / Иные споры) находится в колонке 10 (judgment_first), а не A. _detect_section
   сканирует обе колонки.
3. **Date distribution** — равномерно по всему 2025 (01.01..31.12 = 365 дней).
4. **Case number prefix** — ``2025-001``, ``2025-002`` ... (без квартала).

Запуск (внутри контейнера svc-legal):
    docker compose exec svc-legal python /app/scripts/import_pir_2025.py \
        --xlsx /app/data/pir_2025_final.xlsx
"""
from __future__ import annotations

import argparse
import asyncio
import re
import sys
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy import delete, extract, select, update
from sqlalchemy.ext.asyncio import AsyncSession

sys.path.insert(0, "/app")

from app.db.session import SessionLocal
from app.models import (
    Branch,
    Case,
    CaseComment,
    CaseDocument,
    CaseEvent,
    CaseFinance,
    CaseLawyer,
    CaseLitigation,
    Claim,
    DebtRecoveryEntry,
    EnforcementProceeding,
    Payment,
    ProceduralDeadline,
    User,
)


YEAR = 2025  # таргет-год импорта, по умолчанию

# ─────────────────────────────────────────────────────────────────────────────
# Канонические филиалы (UUID совпадают с normalize_branches.sql)
# ─────────────────────────────────────────────────────────────────────────────
BRANCH_UUID = {
    "ЦА - Центральный аппарат":        "aa000000-0000-0000-0000-000000000001",
    "АО «Вагонсервис»":                "aa000000-0000-0000-0000-000000000002",
    "РФ «Северный»":                   "aa000000-0000-0000-0000-000000000003",
    "РФ «Западный»":                   "aa000000-0000-0000-0000-000000000004",
    "РФ «Южный»":                      "aa000000-0000-0000-0000-000000000005",
    "Филиал «Экспресс»":               "aa000000-0000-0000-0000-000000000006",
    "Филиал «Пригородные перевозки»":  "aa000000-0000-0000-0000-000000000007",
    "Филиал «Сұңқар»":                 "aa000000-0000-0000-0000-000000000008",
}


def _normalize_branch(raw: Optional[str]) -> Optional[str]:
    if not raw:
        return None
    s = str(raw).lower().replace("ё", "е")
    s = re.sub(r"[\"«»“”]", "", s).strip()
    if "вагонсервис" in s:
        return "АО «Вагонсервис»"
    if "северн" in s:
        return "РФ «Северный»"
    if "западн" in s:
        return "РФ «Западный»"
    if "южн" in s:
        return "РФ «Южный»"
    if "экспресс" in s:
        return "Филиал «Экспресс»"
    if "пригородн" in s:
        return "Филиал «Пригородные перевозки»"
    if "с[уұ]н[кқ]ар" in s or re.search(r"с[уұ]н[кқ]ар", s):
        return "Филиал «Сұңқар»"
    if any(t in s for t in ("цлю", "цюс", "ца -", "центральн", "цлвсю")):
        return "ЦА - Центральный аппарат"
    return None


def _norm_lawyer(name: Optional[str]) -> str:
    if not name:
        return ""
    s = str(name).strip()
    if "Ахатов А.А" in s:
        s = "Ахатов А.Б"
    if "Сырлыбаев Е" in s:
        s = "Сырлыбаев Е.Е"
    if "Әбсеметов Д" in s:
        s = "Әбсеметов Д.Е"
    s = s.lower()
    # Заменяем казахские буквы для унификации
    s = s.replace("і", "и").replace("ң", "н").replace("қ", "к").replace("ғ", "г").replace("ү", "у").replace("ұ", "у").replace("ә", "а").replace("ө", "о")
    s = re.sub(r"\.", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    parts = s.split()
    if not parts:
        return ""
    surname = parts[0]
    initials = "".join(p[0] for p in parts[1:3] if p)
    return f"{surname} {initials}".strip()


def _to_decimal(v) -> Decimal:
    if v is None:
        return Decimal("0")
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip()
    if not s:
        return Decimal("0")
    s = s.replace("\xa0", "").replace(" ", "").replace("₸", "")
    s = s.replace(",", ".")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def _txt(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


_APPEAL_INDICATOR_RE = re.compile(
    r"(?:апелляционн\w+|кассационн\w+|апеляционн\w+|апеляцон\w+)\s+(?:жалоб\w*|представлени\w*)"
)
_APPEAL_DISMISSED_RE = re.compile(
    r"(?:"
    r"в\s+(?:удовлетворении\s+)?(?:апелляционн\w+|кассационн\w+|апеляционн\w+|апеляцон\w+)\s+(?:жалоб\w*|представлени\w*).*?отказ"
    r"|отказ\w+\s+в\s+(?:удовлетворении\s+)?(?:апелляционн\w+|кассационн\w+|апеляционн\w+|апеляцон\w+)"
    r"|оставлен[оаы]?\s+без\s+изменени"
    r"|оставить\s+без\s+изменени"
    r"|решени\w*\s+оставлен\w*\s+в\s+силе"
    r")"
)
_DENIED_RE = re.compile(
    r"(?:"
    r"в\s+(?:удовлетворении\s+)?иск\w*.*?отказ"   # в удовлетворении иска отказано / в иске отказано / в исковых требованиях отказано
    r"|отказ\w+\s+в\s+иск"                         # отказано в иске
    r"|отказ\w+\s+от\s+иск"                        # отказе от иска
    r"|иск\w*.*?отозван"                           # иск был отозван / иск отозван
    r"|истц\w+\s+принят\w+\s+решени\w*\s+об\s+отказе"
    r"|прекра[щт]\w+"                              # прекращено / прекратил производство
    r")"
)
_PARTIAL_RE = re.compile(
    r"(?:"
    r"удовлетвор\w+\s+част"
    r"|удовлетвор\w+\s+в\s+част"
    r"|част\w+\s+удовлетвор"
    r")"
)

_RETURN_NEG_RE = re.compile(r"возвра[щт]\w*")
_LEFT_NO_REVIEW_RE = re.compile(r"оставлен\w*\s+без\s+рассмотрени")
_MEDIATION_RE = re.compile(
    r"(?:медиатив|мирово|утверждено\s+согла\w+.*?(?:медиаци|урегулировани))"
)


def _classify_instance(t: str) -> str:
    """Классифицирует исход одной инстанции.

    Возвращает: 'settled' | 'fully' | 'partial' | 'denied' | 'appeal_dismissed' | 'pending'.

    «appeal_dismissed» — апелляция/кассация отклонена, т.е. решение нижестоящего
    суда остаётся в силе. Caller должен опуститься на предыдущую инстанцию.
    """
    if not t:
        return "pending"
    tl = " ".join(t.lower().split())
    # 1. Медиация / мировое
    if _MEDIATION_RE.search(tl):
        return "settled"
    # 2. Особый случай: текст про апелляцию/кассацию. Если упомянута «апелляционная/
    # кассационная жалоба» — обрабатываем отдельно:
    #   - «решение отменено» или «жалоба удовлетворена» → апелляция ПЕРЕСМОТРЕЛА решение,
    #     используем содержание этой инстанции (падаем дальше к общему классификатору).
    #   - иначе → appeal_dismissed (жалоба не повлияла на исход).
    if _APPEAL_INDICATOR_RE.search(tl):
        appeal_granted = (
            re.search(r"решени\w*\s+отменен", tl)
            or re.search(r"акт\w*\s+отменен", tl)
            or re.search(r"жалоб\w*\s+удовлетвор", tl)
            or re.search(r"представлени\w+\s+удовлетвор", tl)
        )
        if not appeal_granted:
            return "appeal_dismissed"
        # Жалоба удовлетворена — падаем дальше, классификатор определит исход
    # 3. Прямое appeal_dismissed (без явного указателя — например «оставлено без изменения»)
    if _APPEAL_DISMISSED_RE.search(tl):
        return "appeal_dismissed"
    # 3. Явный отказ в иске
    if _DENIED_RE.search(tl):
        return "denied"
    # 4. Иск возвращён → denied
    if _RETURN_NEG_RE.search(tl):
        return "denied"
    # 5. Оставлено без рассмотрения → denied
    if _LEFT_NO_REVIEW_RE.search(tl):
        return "denied"
    # 8. Частично удовлетворен
    if _PARTIAL_RE.search(tl):
        return "partial"
    # 9. Удовлетворен (общий matcher — после явного «отказа в удовлетворении»)
    if "удовлетвор" in tl:
        return "fully"
    return "pending"


_INSTANCE_TO_OUTCOME = {
    "fully": "fully_satisfied",
    "partial": "partially_satisfied",
    "settled": "settled",
    "denied": "denied",
    "pending": "pending",
    # appeal_dismissed на 1-й инстанции (нет нижестоящего суда) трактуем как
    # подтверждение существующего решения. Но обычно такое не встречается на j1.
    "appeal_dismissed": "denied",
}


def _infer_outcome(category: str, j1: str, ja: str, jc: str) -> str:
    """Финальный исход — по самой свежей инстанции, кроме случаев когда
    апелляция/кассация только отклонила жалобу (тогда решение нижестоящего суда).
    """
    if category == "mediation":
        return "settled"
    # Кассация
    rc = _classify_instance(jc)
    if rc not in ("pending", "appeal_dismissed"):
        return _INSTANCE_TO_OUTCOME[rc]
    # Апелляция
    ra = _classify_instance(ja)
    if ra not in ("pending", "appeal_dismissed"):
        return _INSTANCE_TO_OUTCOME[ra]
    # 1-я инстанция
    r1 = _classify_instance(j1)
    if r1 != "pending":
        return _INSTANCE_TO_OUTCOME[r1]
    return "pending"


def _infer_status(outcome: str, has_execution: bool) -> str:
    if outcome == "pending":
        return "active"
    if has_execution:
        return "execution"
    if outcome in {"fully_satisfied", "partially_satisfied", "denied", "dismissed", "settled"}:
        return "closed"
    return "active"


def _infer_instance(j1: str, ja: str, jc: str) -> str:
    if jc:
        return "cassation"
    if ja:
        return "appeal"
    return "first"


def _infer_risk(amount: Decimal) -> str:
    a = float(amount)
    if a > 50_000_000:
        return "high"
    if a < 5_000_000:
        return "low"
    return "medium"


# ─────────────────────────────────────────────────────────────────────────────
# Sheet -> role + column maps
# ─────────────────────────────────────────────────────────────────────────────
ROLE_BY_SHEET = {
    "истец": "plaintiff",
    "ответчик": "defendant",
    "3-лицо ": "third_party",
}

COLUMN_MAP_PLAINTIFF = {
    "lawyer": 2, "court": 3, "branch": 4, "counterparty": 5,
    "main_debt": 6, "fines": 7, "state_fee": 8,
    "claim_summary": 9,
    "judgment_first": 10, "judgment_appeal": 11, "judgment_cassation": 12,
    "recovered_main": 13, "recovered_fines": 14, "recovered_state_fee": 15,
    "writ_request_note": 16, "writ_dispatch_note": 17, "execution_proof_note": 18,
    "damage_recovery_note": 19,
}

COLUMN_MAP_DEFENDANT = {
    "lawyer": 2, "court": 3, "branch": 4, "counterparty": 5,
    "main_debt": 6, "fines": 7, "rep_expenses": 8, "state_fee": 9,
    "claim_summary": 10,
    "judgment_first": 11, "judgment_appeal": 12, "judgment_cassation": 13,
    "recovered_main": 14, "recovered_fines": 15,
    "recovered_rep_expenses": 16, "recovered_state_fee": 17,
    "defendant_execution_note": 18, "damage_recovery_note": 19,
}

COLUMN_MAP_THIRD = {
    "lawyer": 2, "court": 3, "branch": 4,
    "plaintiff_name": 5, "counterparty": 6,
    "main_debt": 7, "fines": 8, "rep_expenses": 9, "state_fee": 10,
    "claim_summary": 11,
    "judgment_first": 12, "judgment_appeal": 13, "judgment_cassation": 14,
    "recovered_main": 15, "recovered_fines": 16,
    "recovered_rep_expenses": 17, "recovered_state_fee": 18,
    "third_party_note": 19,
}

COL_MAP_BY_SHEET = {
    "истец": COLUMN_MAP_PLAINTIFF,
    "ответчик": COLUMN_MAP_DEFENDANT,
    "3-лицо ": COLUMN_MAP_THIRD,
}

FIRST_DATA_ROW = {
    "истец": 7,
    "ответчик": 6,
    "3-лицо ": 7,
}


# Расширенный список ключевых слов: 2025-файл использует «иные» (короткое) и
# «иные споры» в разных листах.
SECTION_KEYS = (
    ("закупках", "procurement"),
    ("закупок", "procurement"),
    ("перевозоч", "transportation"),
    ("трудовы", "labor"),
    ("госорган", "government"),
    ("иные спор", "other"),
    ("иные", "other"),          # короткий заголовок на истце 2025
    ("медиатив", "mediation"),
)


def _detect_section_in_text(text) -> Optional[str]:
    """Распознать секционный заголовок в произвольной строке."""
    if not isinstance(text, str):
        return None
    t = text.strip().lower()
    if not t:
        return None
    # Section header — обычно короткий, без числовых данных
    if len(t) > 200:
        return None
    for key, cat in SECTION_KEYS:
        if key in t:
            return cat
    return None


def _detect_section_in_row(ws, r: int) -> Optional[str]:
    """В 2025-файле заголовки секций могут быть в col A ИЛИ в col 10.

    Проверяем оба места. Возвращает категорию или None.
    """
    # Сначала col A
    sec = _detect_section_in_text(ws.cell(row=r, column=1).value)
    if sec:
        return sec
    # Затем col 10 (там в 2025 встроены подзаголовки)
    sec = _detect_section_in_text(ws.cell(row=r, column=10).value)
    if sec:
        return sec
    return None


def _is_section_only_row(ws, r: int) -> bool:
    """True если в строке ТОЛЬКО заголовок секции (нет данных дела).

    Используется для решения «consume этот ряд как заголовок» vs «парсить как case».
    Логика: если col A — заголовок секции, ИЛИ col A пустой/нет номера И col 10
    содержит ключевое слово И col B-C-D-E пустые (нет contractor data), это
    section-only row.
    """
    a_sec = _detect_section_in_text(ws.cell(row=r, column=1).value)
    if a_sec:
        return True
    # col 10 содержит заголовок?
    c10_sec = _detect_section_in_text(ws.cell(row=r, column=10).value)
    if not c10_sec:
        return False
    a_val = ws.cell(row=r, column=1).value
    # Если в col A число (data-row), то col 10 не section-header а часть данных
    if isinstance(a_val, (int, float)) and int(a_val) > 0:
        return False
    # col A пуст или текст — но не число. Проверим что col B/E (counterparty) пустые
    b_val = _txt(ws.cell(row=r, column=2).value)
    e_val = _txt(ws.cell(row=r, column=5).value)
    return not b_val and not e_val


def parse_sheet(ws, role: str) -> list[dict]:
    cmap = COL_MAP_BY_SHEET[ws.title]
    cases: list[dict] = []
    cur_section = "procurement"
    cur_case: Optional[dict] = None
    first_row = FIRST_DATA_ROW.get(ws.title, 1)

    finished = False
    summary_labels = ("предъявлено", "категория", "кол-во")
    for r in range(first_row, ws.max_row + 1):
        if finished:
            break
        a = ws.cell(row=r, column=1).value
        b = str(ws.cell(row=r, column=2).value or "").lower().strip()
        c = str(ws.cell(row=r, column=3).value or "").lower().strip()
        if b in summary_labels or c in summary_labels:
            finished = True
            break

        # Заголовок секции (col A или col 10) — переключаем категорию и идём дальше
        if _is_section_only_row(ws, r):
            sec = _detect_section_in_row(ws, r)
            if sec:
                cur_section = sec
            continue

        if isinstance(a, str) and not a.strip().isdigit():
            continue

        is_new_case = isinstance(a, (int, float)) and int(a) > 0
        cp_val = ws.cell(row=r, column=cmap.get("counterparty", 5)).value
        is_continuation = a is None and isinstance(cp_val, str) and cp_val.strip()

        if is_new_case:
            if cur_case is not None:
                cases.append(cur_case)
            cur_case = {
                "role": role,
                "dispute_category": cur_section,
                "src_row": r,
                "lawyer": _txt(ws.cell(row=r, column=cmap["lawyer"]).value),
                "court": _txt(ws.cell(row=r, column=cmap["court"]).value),
                "branch": _txt(ws.cell(row=r, column=cmap["branch"]).value),
                "counterparty": _txt(ws.cell(row=r, column=cmap["counterparty"]).value),
                "main_debt": _to_decimal(ws.cell(row=r, column=cmap.get("main_debt", 0)).value) if "main_debt" in cmap else Decimal("0"),
                "fines": _to_decimal(ws.cell(row=r, column=cmap.get("fines", 0)).value) if "fines" in cmap else Decimal("0"),
                "rep_expenses": _to_decimal(ws.cell(row=r, column=cmap.get("rep_expenses", 0)).value) if "rep_expenses" in cmap else Decimal("0"),
                "state_fee": _to_decimal(ws.cell(row=r, column=cmap.get("state_fee", 0)).value) if "state_fee" in cmap else Decimal("0"),
                "claim_summary": _txt(ws.cell(row=r, column=cmap["claim_summary"]).value),
                "judgment_first": _txt(ws.cell(row=r, column=cmap["judgment_first"]).value),
                "judgment_appeal": _txt(ws.cell(row=r, column=cmap["judgment_appeal"]).value),
                "judgment_cassation": _txt(ws.cell(row=r, column=cmap["judgment_cassation"]).value),
                "recovered_main": _to_decimal(ws.cell(row=r, column=cmap.get("recovered_main", 0)).value) if "recovered_main" in cmap else Decimal("0"),
                "recovered_fines": _to_decimal(ws.cell(row=r, column=cmap.get("recovered_fines", 0)).value) if "recovered_fines" in cmap else Decimal("0"),
                "recovered_rep_expenses": _to_decimal(ws.cell(row=r, column=cmap.get("recovered_rep_expenses", 0)).value) if "recovered_rep_expenses" in cmap else Decimal("0"),
                "recovered_state_fee": _to_decimal(ws.cell(row=r, column=cmap.get("recovered_state_fee", 0)).value) if "recovered_state_fee" in cmap else Decimal("0"),
                "writ_request_note": _txt(ws.cell(row=r, column=cmap.get("writ_request_note", 0)).value) if "writ_request_note" in cmap else "",
                "writ_dispatch_note": _txt(ws.cell(row=r, column=cmap.get("writ_dispatch_note", 0)).value) if "writ_dispatch_note" in cmap else "",
                "execution_proof_note": _txt(ws.cell(row=r, column=cmap.get("execution_proof_note", 0)).value) if "execution_proof_note" in cmap else "",
                "defendant_execution_note": _txt(ws.cell(row=r, column=cmap.get("defendant_execution_note", 0)).value) if "defendant_execution_note" in cmap else "",
                "damage_recovery_note": _txt(ws.cell(row=r, column=cmap.get("damage_recovery_note", 0)).value) if "damage_recovery_note" in cmap else "",
                "third_party_note": _txt(ws.cell(row=r, column=cmap.get("third_party_note", 0)).value) if "third_party_note" in cmap else "",
                "execution_status": _txt(ws.cell(row=r, column=19).value) if role == "plaintiff" else "",
            }
        elif is_continuation and cur_case is not None:
            def _add_num(field: str):
                if field in cmap:
                    cur_case[field] += _to_decimal(ws.cell(row=r, column=cmap[field]).value)
            def _append_text(field: str):
                if field in cmap:
                    extra = _txt(ws.cell(row=r, column=cmap[field]).value)
                    if extra and extra not in cur_case[field]:
                        cur_case[field] = (cur_case[field] + "\n" + extra).strip()
            for f in ("main_debt", "fines", "rep_expenses", "state_fee",
                      "recovered_main", "recovered_fines", "recovered_rep_expenses", "recovered_state_fee"):
                _add_num(f)
            for f in ("claim_summary", "judgment_first", "judgment_appeal", "judgment_cassation",
                      "writ_request_note", "writ_dispatch_note", "execution_proof_note",
                      "defendant_execution_note", "damage_recovery_note", "third_party_note"):
                _append_text(f)

    if cur_case is not None:
        cases.append(cur_case)
    return cases


# ─────────────────────────────────────────────────────────────────────────────
# Year-scoped purge
# ─────────────────────────────────────────────────────────────────────────────
async def purge_cases_for_year(db: AsyncSession, year: int) -> None:
    """Удаляет дела ТОЛЬКО за указанный год (по filing_date)."""
    target_ids_q = select(Case.id).where(extract("year", Case.filing_date) == year)
    target_ids = [str(r[0]) for r in (await db.execute(target_ids_q)).all()]
    if not target_ids:
        print(f"Nothing to purge for year {year}")
        return
    # Обнулим claims.case_id для затронутых дел
    await db.execute(
        update(Claim).where(Claim.case_id.in_(target_ids)).values(case_id=None)
    )
    # Каскадно зависимые таблицы (cascade в моделях должен сработать, но явно надёжнее)
    for model in (
        EnforcementProceeding, DebtRecoveryEntry, CaseLitigation, CaseFinance,
        Payment, CaseDocument, CaseEvent, CaseComment, ProceduralDeadline,
    ):
        try:
            await db.execute(delete(model).where(model.case_id.in_(target_ids)))
        except Exception as e:
            print(f"  purge {model.__name__}: {e}")
    res = await db.execute(delete(Case).where(Case.id.in_(target_ids)))
    print(f"Deleted cases for {year}: {res.rowcount}")


async def load_lawyer_map(db: AsyncSession) -> dict[str, str]:
    users = (await db.execute(select(User))).scalars().all()
    out = {}
    for u in users:
        key = _norm_lawyer(u.full_name)
        if key and key not in out:
            out[key] = str(u.id)
    return out


def _distribute_dates_full_year(n: int, year: int) -> list[date]:
    """Равномерно по всему году (01.01..31.12 — 365 или 366 дней)."""
    if n <= 0:
        return []
    is_leap = (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0))
    span = 365 if is_leap else 364  # offset от 1 января
    out = []
    for i in range(n):
        offset = round(i * span / max(n - 1, 1)) if n > 1 else 0
        out.append(date(year, 1, 1) + timedelta(days=offset))
    return out


def _parse_manual_adjustments(ws) -> dict[str, Decimal]:
    """Парсит сводный блок листа на предмет ручных корректировок Малики.

    У ВПР Малики бывают строки вида «заключены медиативные соглашения» с числом
    в правой колонке (J = 10 на ИСТЦЕ) и пояснением (K = 11), напр.
    «добровольно выплаченная сумма ТОО «Жарас»». Возвращаем {company_lower: amount}.
    """
    adjustments: dict[str, Decimal] = {}
    company_re = re.compile(
        r"(?:тоо|ао|ип|оо|нао|рг)\s*[«\"]([^»\"]+)[»\"]",
        flags=re.IGNORECASE,
    )
    # Сканируем строки 70..max_row (сводный блок обычно ниже последних данных)
    for r in range(70, ws.max_row + 1):
        # Ищем «медиативн» в любой из первых колонок (B/C/D)
        is_med_row = False
        for c in (2, 3, 4):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "медиативн" in v.lower():
                is_med_row = True
                break
        if not is_med_row:
            continue
        # В этой строке: ищем колонку с числовым значением > 0 (adjustment)
        # и соседнюю колонку с текстом — это компания, к которой относится коррекция.
        for c in range(5, 15):
            val = ws.cell(r, c).value
            if isinstance(val, (int, float)) and float(val) >= 1_000_000:
                # Это потенциальное adjustment-значение. Ищем company в соседних колонках.
                for cc in (c + 1, c - 1, c + 2):
                    txt = ws.cell(r, cc).value
                    if isinstance(txt, str):
                        m = company_re.search(txt)
                        if m:
                            company = m.group(1).strip().lower()
                            adjustments[company] = Decimal(str(val))
                            break
                if adjustments:
                    break
    return adjustments


def _detect_category_by_content(court: str, claim_summary: str, judgment_first: str) -> Optional[str]:
    text = " ".join(filter(None, [court, claim_summary, judgment_first])).lower()
    labor_keys = (
        "восстановлен", "трудового договор", "расторжен", "инспекци", "несчастн",
        "отпускн", "вынужденн", "проводник", "форменную одежду", "должностной инструкции",
        "билетного кассир", "об отмене приказа",
    )
    for k in labor_keys:
        if k in text:
            return "labor"
    gov_keys = ("смас", "административн", "предписани", "айзрк", "санитарно-эпидемиологическ",
                "защите и развитию конкуренции")
    if any(k in text for k in gov_keys):
        return "government"
    other_keys = ("возмещени", "ущерб", "несанкционированн", "сход пассажир")
    if any(k in text for k in other_keys):
        return "other"
    return "procurement"


async def import_cases(db: AsyncSession, parsed: list[dict], year: int, manual_adjustments: dict[str, Decimal] | None = None) -> int:
    lawyer_map = await load_lawyer_map(db)
    dates = _distribute_dates_full_year(len(parsed), year)
    adjustments = manual_adjustments or {}
    inserted = 0

    for i, c in enumerate(parsed):
        raw_lawyers = re.split(r"[\n,]", c["lawyer"]) if c["lawyer"] else []
        lawyer_ids = []
        for rl in raw_lawyers:
            lk = _norm_lawyer(rl)
            if lk and lk in lawyer_map:
                lawyer_ids.append(lawyer_map[lk])
                
        lawyer_id = lawyer_ids[0] if lawyer_ids else None

        canon_branch = _normalize_branch(c["branch"])
        branch_uuid = BRANCH_UUID.get(canon_branch or "")
        if not branch_uuid:
            branch_uuid = BRANCH_UUID["ЦА - Центральный аппарат"]

        counterparty = c["counterparty"] or "—"
        if c["role"] == "plaintiff":
            plaintiff_name = "АО «Пассажирские перевозки»"
            defendant_name = counterparty
            company_name = counterparty
        elif c["role"] == "defendant":
            plaintiff_name = counterparty
            defendant_name = "АО «Пассажирские перевозки»"
            company_name = counterparty
        else:
            plaintiff_name = c.get("plaintiff_name") or "—"
            defendant_name = counterparty
            company_name = counterparty

        court_block = c["court"]
        if "\n" in court_block:
            court_name, judge = court_block.split("\n", 1)
            court_name = court_name.strip()
            judge = judge.replace("судья", "").strip()
        else:
            court_name = court_block
            judge = ""

        # В файле 2025 секции явно проставлены через col A и col 10 — доверяем
        # секционному заголовку, content-detection НЕ применяем (он перекрашивал
        # procurement-дела в labor/government по ключевым словам в тексте акта,
        # что давало расхождение с эталоном Малики).
        category = c["dispute_category"]
        outcome = _infer_outcome(category, c["judgment_first"], c["judgment_appeal"], c["judgment_cassation"])
        if outcome == "settled":
            category = "mediation"
        is_defendant = c["role"] == "defendant"
        rec_any = any(_to_decimal(c[k]) > 0 for k in ("recovered_main", "recovered_fines", "recovered_state_fee"))
        exec_proof = c.get("execution_proof_note", "")
        writ_req = c.get("writ_request_note", "")
        def_exec = c.get("defendant_execution_note", "")
        
        if outcome == "pending":
            status = "active"
        else:
            exec_note = c.get("damage_recovery_note", "").lower()
            def_exec_lower = def_exec.lower()
            proof_note_lower = exec_proof.lower()
            disp_lower = writ_req.lower() if isinstance(writ_req, str) else "" 
            # Note: writ_req logic was wrong in the old script, it was actually using writ_dispatch_note?
            # c.get("writ_dispatch_note", "") is better.
            disp_note = c.get("writ_dispatch_note", "")
            disp_note_lower = disp_note.lower()

            is_closed = False
            if "исполнено" in exec_note or "исполнено" in proof_note_lower:
                is_closed = True
            elif is_defendant and ("пл.пор" in def_exec_lower or "пл. пор" in def_exec_lower or "пл.поручение" in def_exec_lower):
                is_closed = True
            elif is_defendant and def_exec_lower and not ("на исполнении" in exec_note):
                is_closed = True

            status = "closed"
            if not is_closed:
                if "на исполнении" in exec_note:
                    status = "execution"
                elif disp_note_lower:
                    status = "execution"
                elif proof_note_lower and ("возбужден" in proof_note_lower or "направлен" in proof_note_lower or "предъявлен" in proof_note_lower):
                    status = "execution"
        
        instance = _infer_instance(c["judgment_first"], c["judgment_appeal"], c["judgment_cassation"])
        case_type = "civil"
        if category == "labor":
            case_type = "labor"
        elif "смас" in court_name.lower() or "административн" in court_name.lower():
            case_type = "administrative"

        claim_amount = c["main_debt"] + c["fines"] + c["state_fee"] + c.get("rep_expenses", Decimal("0"))

        # Ручная корректировка Малики: для добровольно выплаченных сумм
        # (записанных в сводном блоке Excel) — пишем в paid_amount, PIR-экспорт
        # учтёт это в Медиативной сумме.
        cp_lower = (counterparty or "").lower()
        paid_amount = Decimal("0")
        for adj_company, adj_value in adjustments.items():
            if adj_company in cp_lower or cp_lower in adj_company:
                paid_amount = adj_value
                break

        case = Case(
            case_number=f"{year}-{i+1:03d}",
            court=court_name or "Суд не указан",
            judge=judge or "—",
            court_instance=instance,
            case_type=case_type,
            status=status,
            outcome=outcome,
            party_role=c["role"],
            opponent_type="juridical" if any(t in counterparty.lower() for t in ("тоо", "ао", "ип ", "оо ", "рг", "филиал", "нао")) else "physical",
            plaintiff=plaintiff_name[:512],
            defendant=defendant_name[:512],
            company=company_name[:512],
            company_bin="",
            city="—",
            filing_date=dates[i],
            last_updated=dates[i],
            days_overdue=0,
            risk_level=_infer_risk(claim_amount),
            is_archived=False,
            dispute_category=category,
            branch_id=branch_uuid,
            assigned_lawyer_id=lawyer_id,
        )
        case.finances = CaseFinance(
            claim_amount=claim_amount,
            main_debt=c["main_debt"],
            state_fee=c["state_fee"],
            fines=c["fines"],
            rep_expenses=c.get("rep_expenses", Decimal("0")),
            other_costs=Decimal("0"),
            paid_amount=paid_amount,
            recovered_main=c.get("recovered_main", Decimal("0")),
            recovered_fines=c.get("recovered_fines", Decimal("0")),
            recovered_state_fee=c.get("recovered_state_fee", Decimal("0")),
            recovered_rep_expenses=c.get("recovered_rep_expenses", Decimal("0")),
        )
        now = datetime.now(timezone.utc)
        case.litigation = CaseLitigation(
            claim_summary=c["claim_summary"],
            judgment_first=c["judgment_first"],
            judgment_appeal=c["judgment_appeal"],
            judgment_cassation=c["judgment_cassation"],
            writ_request_note=c.get("writ_request_note", ""),
            writ_dispatch_note=c.get("writ_dispatch_note", ""),
            execution_proof_note=c.get("execution_proof_note", ""),
            defendant_execution_note=c.get("defendant_execution_note", ""),
            damage_recovery_note=c.get("damage_recovery_note", ""),
            third_party_note=c.get("third_party_note", ""),
            created_at=now,
            updated_at=now,
        )
        if lawyer_ids:
            unique_ids = list(dict.fromkeys(lawyer_ids))
            case.case_lawyers = [
                CaseLawyer(user_id=uid) for uid in unique_ids
            ]

        db.add(case)
        inserted += 1
    return inserted


async def main_async(xlsx_path: str, year: int) -> None:
    src = Path(xlsx_path)
    if not src.is_file():
        raise SystemExit(f"Xlsx not found: {src}")
    wb = load_workbook(src, data_only=True)

    parsed: list[dict] = []
    manual_adjustments: dict[str, Decimal] = {}
    for sheet_name, role in ROLE_BY_SHEET.items():
        if sheet_name not in wb.sheetnames:
            print(f"Sheet missing: {sheet_name}")
            continue
        ws = wb[sheet_name]
        rows = parse_sheet(ws, role)
        print(f"  {sheet_name}: parsed {len(rows)} cases")
        parsed.extend(rows)
        # Парсим ручные корректировки сводного блока (например «добровольно выплаченная сумма ТОО Жарас»)
        sheet_adj = _parse_manual_adjustments(ws)
        for company, amount in sheet_adj.items():
            manual_adjustments[company] = amount
            print(f"  [manual adjustment] {sheet_name}: {company} → {amount}")
    print(f"\nTotal parsed: {len(parsed)}")

    async with SessionLocal() as db:
        await purge_cases_for_year(db, year)
        await db.flush()
        inserted = await import_cases(db, parsed, year, manual_adjustments)
        await db.commit()
        print(f"\nInserted: {inserted} cases for year {year}")

        from collections import Counter
        by_role = Counter(c["role"] for c in parsed)
        by_cat = Counter(c["dispute_category"] for c in parsed)
        print(f"By role: {dict(by_role)}")
        print(f"By section-category: {dict(by_cat)}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--year", type=int, default=YEAR)
    args = parser.parse_args()
    asyncio.run(main_async(args.xlsx, args.year))


if __name__ == "__main__":
    main()
