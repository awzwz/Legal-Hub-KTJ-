"""Fill PIR xlsx from in-memory Case rows (no FastAPI/JWT imports — testable in isolation).

Раскладка строго следует шаблону КТЖ:

* Лист **«истец»** — категории «Закупки/договоры», «Трудовые», «Иные», «Медиативные».
  Перевозочные споры на этом листе не предусмотрены — сваливаются в «Иные».
* Лист **«ответчик»** — добавляется «Иски, вытекающие из перевозочного процесса».
* Лист **«в качестве 3 лица»** — реестр дел, где КТЖ-ГП выступает третьим лицом.
* Лист **«3-лицо »** (с пробелом, КТЖ — истец c третьими лицами) — дублирует строки
  третьего лица для совместимости с шаблоном.
* Лист **«исполнительное производство»** — построчные записи `enforcement_proceedings`.
* Лист **«инф по сниже дебит. задолжненно»** — `debt_recovery_entries`.

Колонки внутри листов соответствуют шапке шаблона: см. `_write_*_row` ниже. Нумерация
дел идёт **отдельно для каждой категории** (как в эталонном файле).
"""

from __future__ import annotations

import math
import re
import zipfile
from copy import copy
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.models import Case, DebtRecoveryEntry, EnforcementProceeding

TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "templates" / "pir_report_2025_template.xlsx"

# party_role → лист, в который пишется первичная запись
SHEET_BY_ROLE = {
    "plaintiff": "истец",
    "defendant": "ответчик",
    "third_party": "в качестве 3 лица",
}

# Резервный лист для роли "third_party" (КТЖ — истец, но дело с 3-ми лицами):
# дублируем строки сюда, чтобы шаблонный реестр не оставался пустым.
THIRD_PARTY_MIRROR_SHEET = "3-лицо "

ENFORCEMENT_SHEET = "исполнительное производство"
ENFORCEMENT_FIRST_ROW = 6

DEBT_SHEET = "инф по сниже дебит. задолжненно"
DEBT_FIRST_ROW = 7

# Старые константы (используются в тестах/интеграциях): первая строка данных
# первой категории на основном листе. Сохраняем имена ради совместимости.
FIRST_DATA_ROW = {
    "истец": 8,
    "ответчик": 7,
    "3-лицо ": 7,
    "в качестве 3 лица": 7,
}

# Где заканчивается «фиксированная» шапка — выше этой строки openpyxl не должен
# трогать ни значения, ни merge.
HEADER_LAST_ROW = {
    "истец": 6,
    "ответчик": 5,
    "3-лицо ": 6,
    "в качестве 3 лица": 6,
}

MAIN_MAX_COL = {
    "истец": 19,
    "ответчик": 19,
    "3-лицо ": 19,
    "в качестве 3 лица": 19,
}

# Порядок категорий внутри каждого «ролевого» листа: совпадает с шаблоном КТЖ.
# Названия идентичны эталону (опечатка «законадательства» в «ответчик» исправлена;
# юристы предпочли корректное написание).
PLAINTIFF_CATEGORY_ORDER: list[tuple[str, str]] = [
    ("procurement", "Иски, связанные с нарушением законодательства о закупках и вытекающие из договоров"),
    ("labor", "Трудовые споры"),
    ("other", "Иные споры"),
    ("mediation", "Медиативные соглашения"),
]
DEFENDANT_CATEGORY_ORDER: list[tuple[str, str]] = [
    ("procurement", "Иски, связанные с нарушением законодательства о закупках и вытекающие из договоров"),
    ("transportation", "Иски, вытекающие из перевозочного процесса"),
    ("labor", "Трудовые споры"),
    ("other", "Иные споры"),
    ("mediation", "Медиативные соглашения"),
]
THIRD_PARTY_CATEGORY_ORDER: list[tuple[str, str]] = [
    ("procurement", "Иски, связанные с нарушением законодательства о закупках и вытекающие из договоров"),
    ("labor", "Трудовые споры"),
    ("other", "Иные споры"),
    ("mediation", "Медиативные соглашения"),
]

CATEGORY_ORDER_BY_SHEET: dict[str, list[tuple[str, str]]] = {
    "истец": PLAINTIFF_CATEGORY_ORDER,
    "ответчик": DEFENDANT_CATEGORY_ORDER,
    "3-лицо ": THIRD_PARTY_CATEGORY_ORDER,
    "в качестве 3 лица": THIRD_PARTY_CATEGORY_ORDER,
}


def _allowed_categories(sheet_name: str) -> set[str]:
    return {cat for cat, _ in CATEGORY_ORDER_BY_SHEET.get(sheet_name, [])}


def _normalize_category(case: Case, sheet_name: str) -> str:
    """Приводит ``case.dispute_category`` к одной из категорий целевого листа.

    Категории, которых нет на листе (например, «transportation» на «истец»), сваливаются в «other».
    """
    raw = (getattr(case, "dispute_category", None) or "").strip().lower()
    allowed = _allowed_categories(sheet_name)
    if raw in allowed:
        return raw
    if raw == "transportation" and "other" in allowed:
        return "other"
    return "other" if "other" in allowed else (next(iter(allowed)) if allowed else "procurement")


_ILLEGAL_XML_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_CALCCHAIN_REL_RE = re.compile(
    r'<Relationship\b[^>]*?(?:relationships/calcChain|calcChain\.xml)[^>]*/>\s*',
    re.IGNORECASE,
)
_CT_CALCCHAIN_RE = re.compile(
    r'<Override\s+PartName="/xl/calcChain\.xml"[^>]*/>\s*',
    re.IGNORECASE,
)
_DRAWING_REL_RE = re.compile(
    r'<Relationship\b[^>]*?(?:relationships/drawing|/drawings/drawing)[^>]*/>\s*',
    re.IGNORECASE,
)
_CT_DRAWING_RE = re.compile(
    r'<Override\s+PartName="/xl/drawings/[^"]+"[^>]*/>\s*',
    re.IGNORECASE,
)
_CT_SHAREDSTRINGS_RE = re.compile(
    r'<Override\s+PartName="/xl/sharedStrings\.xml"[^>]*/>\s*',
    re.IGNORECASE,
)
_SS_REL_RE = re.compile(
    r'<Relationship\b[^>]*?(?:relationships/sharedStrings|sharedStrings\.xml)[^>]*/>\s*',
    re.IGNORECASE,
)


def _safe_txt(value: str | None) -> str:
    if not value:
        return ""
    return _ILLEGAL_XML_CHARS.sub("", str(value))


def _fnum(value: object) -> float:
    try:
        v = float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return 0.0
    if math.isnan(v) or math.isinf(v):
        return 0.0
    return v


def _excel_num(value: object) -> int | float | None:
    """Как в типовом ПИР-xlsx: пустая ячейка вместо нуля для сумм (импорт даёт Decimal(0))."""
    v = _fnum(value)
    if v == 0.0:
        return None
    if abs(v - int(v)) < 1e-9:
        return int(v)
    return round(v, 6)


def _outcome_bucket(case: Case) -> str:
    """Куда дело попадает в сводном блоке внизу листа."""
    cat = (getattr(case, "dispute_category", None) or "").strip().lower()
    out = (case.outcome or "").strip().lower()
    if cat == "mediation" or out == "settled":
        return "mediation"
    if out in {"fully_satisfied", "partially_satisfied"}:
        return "satisfied"
    if out in {"denied", "dismissed"}:
        return "denied"
    return "filed_only"


def _claim_total(case: Case) -> Decimal:
    fin = case.finances
    if fin is None:
        return Decimal("0")
    val = fin.claim_amount
    return val if isinstance(val, Decimal) else Decimal(str(val or 0))


def _unmerge_in_range(ws: Worksheet, start_row: int, end_row: int | None = None) -> None:
    """Снять merge у диапазонов, целиком или частично попадающих в [start_row..end_row]."""
    for cr in list(ws.merged_cells.ranges):
        if cr.max_row < start_row:
            continue
        if end_row is not None and cr.min_row > end_row:
            continue
        ws.unmerge_cells(str(cr))


def _clear_rows(ws: Worksheet, start_row: int, end_row: int, max_col: int) -> None:
    """Очистить значения и убрать рамки/заливку у каждой ячейки в диапазоне."""
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _court_block(case: Case) -> str:
    court = _safe_txt(case.court)
    j = _safe_txt(case.judge).strip()
    if not court:
        return ""
    if not j or j in ("—", "–", "-"):
        return court
    return _safe_txt(f"{court}\n{j}".strip())


# ---------------------------------------------------------------------------
# Стили строк-разделителей категорий и шапки авто-сводки
# ---------------------------------------------------------------------------

_CATEGORY_FILL = PatternFill(fill_type="solid", start_color="FFE599", end_color="FFE599")
_CATEGORY_FONT = Font(bold=True, size=11, name="Times New Roman")
_CATEGORY_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
_THIN_SIDE = Side(style="thin", color="000000")
_DATA_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_DATA_ALIGN_TEXT = Alignment(horizontal="left", vertical="top", wrap_text=True)
_DATA_ALIGN_NUM = Alignment(horizontal="right", vertical="center", wrap_text=False)
_DATA_FONT = Font(size=10, name="Times New Roman")
_SUMMARY_HEADER_FILL = PatternFill(fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")
_SUMMARY_HEADER_FONT = Font(bold=True, size=11, name="Times New Roman")
_SUMMARY_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_category_row(ws: Worksheet, row_idx: int, max_col: int, text: str) -> None:
    cell = ws.cell(row=row_idx, column=1)
    cell.value = text
    cell.font = _CATEGORY_FONT
    cell.fill = _CATEGORY_FILL
    cell.alignment = _CATEGORY_ALIGN
    cell.border = _DATA_BORDER
    if max_col >= 2:
        merge_ref = f"A{row_idx}:{ws.cell(row=row_idx, column=max_col).coordinate}"
        ws.merge_cells(merge_ref)
    ws.row_dimensions[row_idx].height = 28


def _apply_data_cell(cell, value, *, numeric: bool) -> None:
    cell.value = value
    cell.font = _DATA_FONT
    cell.alignment = _DATA_ALIGN_NUM if numeric else _DATA_ALIGN_TEXT
    cell.border = _DATA_BORDER


# ---------------------------------------------------------------------------
# Запись строк по ролям
# ---------------------------------------------------------------------------


@dataclass
class _RowCtx:
    ws: Worksheet
    row_idx: int
    seq: int
    case: Case


def _write_истец_row(ctx: _RowCtx) -> None:
    case = ctx.case
    fin = case.finances
    lit = case.litigation
    cells: list[tuple[int, object, bool]] = [
        (1, ctx.seq, True),
        (2, _safe_txt(case.assigned_lawyer.full_name if case.assigned_lawyer else ""), False),
        (3, _court_block(case), False),
        (4, _safe_txt(case.branch.name if case.branch else ""), False),
        (5, _safe_txt(case.defendant), False),
        (6, _excel_num(fin.main_debt) if fin else None, True),
        (7, _excel_num(fin.fines) if fin else None, True),
        (8, _excel_num(fin.state_fee) if fin else None, True),
        (9, _safe_txt((lit.claim_summary if lit else "") or ""), False),
        (10, _safe_txt((lit.judgment_first if lit else "") or ""), False),
        (11, _safe_txt((lit.judgment_appeal if lit else "") or ""), False),
        (12, _safe_txt((lit.judgment_cassation if lit else "") or ""), False),
        (13, _excel_num(fin.recovered_main) if fin else None, True),
        (14, _excel_num(fin.recovered_fines) if fin else None, True),
        (15, _excel_num(fin.recovered_state_fee) if fin else None, True),
        (16, _safe_txt((lit.writ_request_note if lit else "") or ""), False),
        (17, _safe_txt((lit.writ_dispatch_note if lit else "") or ""), False),
        (18, _safe_txt((lit.execution_proof_note if lit else "") or ""), False),
        (19, _safe_txt((lit.damage_recovery_note if lit else "") or ""), False),
    ]
    for col, value, numeric in cells:
        _apply_data_cell(ctx.ws.cell(row=ctx.row_idx, column=col), value, numeric=numeric)


def _write_ответчик_row(ctx: _RowCtx) -> None:
    case = ctx.case
    fin = case.finances
    lit = case.litigation
    cells: list[tuple[int, object, bool]] = [
        (1, ctx.seq, True),
        (2, _safe_txt(case.assigned_lawyer.full_name if case.assigned_lawyer else ""), False),
        (3, _court_block(case), False),
        (4, _safe_txt(case.branch.name if case.branch else ""), False),
        (5, _safe_txt(case.plaintiff), False),
        (6, _excel_num(fin.main_debt) if fin else None, True),
        (7, _excel_num(fin.fines) if fin else None, True),
        (8, _excel_num(fin.rep_expenses) if fin else None, True),
        (9, _excel_num(fin.state_fee) if fin else None, True),
        (10, _safe_txt((lit.claim_summary if lit else "") or ""), False),
        (11, _safe_txt((lit.judgment_first if lit else "") or ""), False),
        (12, _safe_txt((lit.judgment_appeal if lit else "") or ""), False),
        (13, _safe_txt((lit.judgment_cassation if lit else "") or ""), False),
        # ВЗЫСКАННАЯ сумма: основная / штрафы / представительские / госпошлина
        (14, _excel_num(fin.recovered_main) if fin else None, True),
        (15, _excel_num(fin.recovered_fines) if fin else None, True),
        (16, _excel_num(fin.recovered_rep_expenses) if fin else None, True),
        (17, _excel_num(fin.recovered_state_fee) if fin else None, True),
        (18, _safe_txt((lit.defendant_execution_note if lit else "") or ""), False),
        (19, _safe_txt((lit.damage_recovery_note if lit else "") or ""), False),
    ]
    for col, value, numeric in cells:
        _apply_data_cell(ctx.ws.cell(row=ctx.row_idx, column=col), value, numeric=numeric)


def _write_third_party_row(ctx: _RowCtx) -> None:
    """Общий формат для листов «3-лицо » и «в качестве 3 лица»."""
    case = ctx.case
    fin = case.finances
    lit = case.litigation
    cells: list[tuple[int, object, bool]] = [
        (1, ctx.seq, True),
        (2, _safe_txt(case.assigned_lawyer.full_name if case.assigned_lawyer else ""), False),
        (3, _court_block(case), False),
        (4, _safe_txt(case.branch.name if case.branch else ""), False),
        (5, _safe_txt(case.plaintiff), False),
        (6, _safe_txt(case.defendant), False),
        (7, _excel_num(fin.main_debt) if fin else None, True),
        (8, _excel_num(fin.fines) if fin else None, True),
        (9, _excel_num(fin.rep_expenses) if fin else None, True),
        (10, _excel_num(fin.state_fee) if fin else None, True),
        (11, _safe_txt((lit.claim_summary if lit else "") or ""), False),
        (12, _safe_txt((lit.judgment_first if lit else "") or ""), False),
        (13, _safe_txt((lit.judgment_appeal if lit else "") or ""), False),
        (14, _safe_txt((lit.judgment_cassation if lit else "") or ""), False),
        # ВЗЫСКАННАЯ сумма
        (15, _excel_num(fin.recovered_main) if fin else None, True),
        (16, _excel_num(fin.recovered_fines) if fin else None, True),
        (17, _excel_num(fin.recovered_rep_expenses) if fin else None, True),
        (18, _excel_num(fin.recovered_state_fee) if fin else None, True),
        (19, _safe_txt((lit.third_party_note if lit else "") or ""), False),
    ]
    for col, value, numeric in cells:
        _apply_data_cell(ctx.ws.cell(row=ctx.row_idx, column=col), value, numeric=numeric)


_ROW_WRITER_BY_SHEET = {
    "истец": _write_истец_row,
    "ответчик": _write_ответчик_row,
    "3-лицо ": _write_third_party_row,
    "в качестве 3 лица": _write_third_party_row,
}


# ---------------------------------------------------------------------------
# Сводный блок (предъявлено / удовлетворено / медиативные / отказано)
# ---------------------------------------------------------------------------


def _write_summary_block(ws: Worksheet, start_row: int, max_col: int, cases: Sequence[Case]) -> int:
    """Пишет сводку и возвращает первую свободную строку после неё."""
    if start_row + 6 > ws.max_row:
        # ничего страшного: openpyxl расширит лист по мере записи

        pass

    counts = {"filed": 0, "satisfied": 0, "mediation": 0, "denied": 0}
    sums = {"filed": Decimal("0"), "satisfied": Decimal("0"), "mediation": Decimal("0"), "denied": Decimal("0")}
    for c in cases:
        bucket = _outcome_bucket(c)
        amount = _claim_total(c)
        counts["filed"] += 1
        sums["filed"] += amount
        if bucket == "satisfied":
            counts["satisfied"] += 1
            sums["satisfied"] += amount
        elif bucket == "mediation":
            counts["mediation"] += 1
            sums["mediation"] += amount
        elif bucket == "denied":
            counts["denied"] += 1
            sums["denied"] += amount

    title_row = start_row + 1
    title_cell = ws.cell(row=title_row, column=2)
    title_cell.value = "Итог по делам листа"
    title_cell.font = _SUMMARY_HEADER_FONT
    title_cell.fill = _SUMMARY_HEADER_FILL
    title_cell.alignment = _SUMMARY_HEADER_ALIGN
    title_cell.border = _DATA_BORDER
    end_col_letter = ws.cell(row=title_row, column=max_col).coordinate
    ws.merge_cells(f"B{title_row}:{end_col_letter}")

    header_row = title_row + 1
    headers = ("Категория", "Кол-во дел", "Сумма, тенге")
    for offset, text in enumerate(headers):
        cell = ws.cell(row=header_row, column=2 + offset)
        cell.value = text
        cell.font = _SUMMARY_HEADER_FONT
        cell.fill = _SUMMARY_HEADER_FILL
        cell.alignment = _SUMMARY_HEADER_ALIGN
        cell.border = _DATA_BORDER

    rows = (
        ("Предъявлено", counts["filed"], sums["filed"]),
        ("Удовлетворено", counts["satisfied"], sums["satisfied"]),
        ("Заключены медиативные соглашения", counts["mediation"], sums["mediation"]),
        ("Отказано", counts["denied"], sums["denied"]),
    )
    for i, (label, count, total) in enumerate(rows, start=1):
        r = header_row + i
        c1 = ws.cell(row=r, column=2)
        c1.value = label
        c1.font = _DATA_FONT
        c1.alignment = _DATA_ALIGN_TEXT
        c1.border = _DATA_BORDER
        c2 = ws.cell(row=r, column=3)
        c2.value = count or None
        c2.font = _DATA_FONT
        c2.alignment = _DATA_ALIGN_NUM
        c2.border = _DATA_BORDER
        c3 = ws.cell(row=r, column=4)
        c3.value = _excel_num(total)
        c3.font = _DATA_FONT
        c3.alignment = _DATA_ALIGN_NUM
        c3.border = _DATA_BORDER

    return header_row + len(rows) + 1


# ---------------------------------------------------------------------------
# Основная отрисовка ролевого листа
# ---------------------------------------------------------------------------


def _render_role_sheet(wb: Workbook, sheet_name: str, cases: Iterable[Case]) -> None:
    """Перерисовать ролевой лист: шапку оставляем, тело и сводку строим заново.

    ``FIRST_DATA_ROW`` указывает на первую строку **данных** первой категории —
    непосредственно над ней размещается строка-разделитель ``first_data_row - 1``
    (как в шаблоне КТЖ).
    """
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    max_col = MAIN_MAX_COL[sheet_name]
    header_last = HEADER_LAST_ROW[sheet_name]
    first_data_row = FIRST_DATA_ROW[sheet_name]
    writer = _ROW_WRITER_BY_SHEET[sheet_name]
    cases_list = list(cases)

    clear_until = max(ws.max_row, first_data_row + len(cases_list) + 25)
    _unmerge_in_range(ws, header_last + 1, None)
    _clear_rows(ws, header_last + 1, clear_until, max_col)

    by_category: dict[str, list[Case]] = {}
    for c in cases_list:
        cat = _normalize_category(c, sheet_name)
        by_category.setdefault(cat, []).append(c)

    # Стабильный порядок дел внутри категории — по case_number, затем по filing_date
    for items in by_category.values():
        items.sort(key=lambda x: (x.case_number or "", x.filing_date or date.min))

    # Первая строка-разделитель категории идёт на (first_data_row - 1), как в шаблоне.
    row_idx = first_data_row - 1
    for cat, title in CATEGORY_ORDER_BY_SHEET[sheet_name]:
        items = by_category.get(cat, [])
        if not items:
            continue
        _apply_category_row(ws, row_idx, max_col, title)
        row_idx += 1
        for seq, case in enumerate(items, start=1):
            writer(_RowCtx(ws=ws, row_idx=row_idx, seq=seq, case=case))
            row_idx += 1

    # Сводка ниже данных (через одну пустую строку)
    if cases_list:
        _write_summary_block(ws, row_idx + 1, max_col, cases_list)


def _horizontalize_main_sheet_headers(wb: Workbook) -> None:
    """После round-trip openpyxl+merge шапка основных листов часто становится «вертикальной».

    В эталонном xlsx заголовки колонок горизонтальны; выравниваем шапку до первой строки данных.
    """
    hdr_align = Alignment(text_rotation=0, horizontal="center", vertical="center", wrap_text=True)
    for sheet_name, last_header in HEADER_LAST_ROW.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        max_col = MAIN_MAX_COL.get(sheet_name, 19)
        for r in range(1, last_header + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                if cell.alignment is None:
                    cell.alignment = hdr_align
                else:
                    cell.alignment = copy(hdr_align)


def _fill_main_sheets(wb: Workbook, cases: list[Case]) -> None:
    plaintiff_cases: list[Case] = []
    defendant_cases: list[Case] = []
    third_party_cases: list[Case] = []
    for c in cases:
        role = (c.party_role or "").strip().lower()
        if role == "plaintiff":
            plaintiff_cases.append(c)
        elif role == "defendant":
            defendant_cases.append(c)
        elif role == "third_party":
            third_party_cases.append(c)

    _render_role_sheet(wb, "истец", plaintiff_cases)
    _render_role_sheet(wb, "ответчик", defendant_cases)
    # third_party дублируется в оба листа — выбран пользователем сознательно.
    _render_role_sheet(wb, "в качестве 3 лица", third_party_cases)
    _render_role_sheet(wb, THIRD_PARTY_MIRROR_SHEET, third_party_cases)

    _horizontalize_main_sheet_headers(wb)


# ---------------------------------------------------------------------------
# Листы исполнительного производства и дебиторки (без категорий)
# ---------------------------------------------------------------------------


def _enforcement_rows_in_period(case: Case, date_from: date, date_to: date) -> list[EnforcementProceeding]:
    return [e for e in case.enforcement_proceedings if date_from <= e.recorded_at <= date_to]


def _debt_rows_in_period(case: Case, date_from: date, date_to: date) -> list[DebtRecoveryEntry]:
    return [
        e for e in case.debt_recovery_entries if e.case_id == case.id and date_from <= e.recorded_at <= date_to
    ]


def _fill_enforcement_sheet(wb: Workbook, cases: list[Case], date_from: date, date_to: date) -> None:
    ws = wb[ENFORCEMENT_SHEET]
    start = ENFORCEMENT_FIRST_ROW
    _unmerge_in_range(ws, start, None)
    _clear_rows(ws, start, max(ws.max_row, start + 50), max_col=14)

    rows: list[EnforcementProceeding] = []
    for c in cases:
        rows.extend(_enforcement_rows_in_period(c, date_from, date_to))
    rows.sort(key=lambda x: (x.recorded_at, str(x.id)))

    for i, e in enumerate(rows, start=1):
        r = start + i - 1
        cells = [
            (1, i, True),
            (2, _safe_txt(e.debtor_name or ""), False),
            (3, _safe_txt(e.debtor_bin or ""), False),
            (4, _safe_txt(e.court_act_summary or ""), False),
            (5, _excel_num(e.amount_total), True),
            (6, _excel_num(e.amount_main), True),
            (7, _excel_num(e.amount_fines), True),
            (8, _excel_num(e.amount_fees), True),
            (9, _safe_txt(e.progress_notes or ""), False),
            (10, _excel_num(e.collected_amount), True),
            (11, _safe_txt(e.collection_doc_ref or ""), False),
            (12, _excel_num(e.balance_remaining), True),
            (13, _safe_txt(e.status_label or ""), False),
        ]
        for col, value, numeric in cells:
            _apply_data_cell(ws.cell(row=r, column=col), value, numeric=numeric)


def _unhide_report_sheets(wb: Workbook) -> None:
    """Шаблон КТЖ часто помечает листы как hidden; в выгрузке все вкладки должны быть видимы."""
    for name in wb.sheetnames:
        ws = wb[name]
        st = getattr(ws, "sheet_state", "visible") or "visible"
        if st != "visible":
            ws.sheet_state = "visible"


def _fill_debt_sheet(wb: Workbook, cases: list[Case], date_from: date, date_to: date) -> None:
    ws = wb[DEBT_SHEET]
    start = DEBT_FIRST_ROW
    _unmerge_in_range(ws, start, None)
    _clear_rows(ws, start, max(ws.max_row, start + 50), max_col=15)

    rows: list[DebtRecoveryEntry] = []
    for c in cases:
        rows.extend(_debt_rows_in_period(c, date_from, date_to))
    rows.sort(key=lambda x: (x.recorded_at, str(x.id)))

    for i, d in enumerate(rows, start=1):
        r = start + i - 1
        cells = [
            (1, i, True),
            (2, _safe_txt(d.debtor_name or ""), False),
            (3, _safe_txt(d.debtor_status or ""), False),
            (4, _excel_num(d.debt_amount), True),
            (5, _safe_txt(d.work_summary or ""), False),
            (6, _excel_num(d.paid_amount), True),
            (8, _excel_num(d.written_off_amount), True),
        ]
        for col, value, numeric in cells:
            _apply_data_cell(ws.cell(row=r, column=col), value, numeric=numeric)


# ---------------------------------------------------------------------------
# Сборка zip-пакета (фикс для Apple Numbers)
# ---------------------------------------------------------------------------


def _worksheets_use_shared_string_type(parts: dict[str, bytes]) -> bool:
    """True if any sheet cell uses shared-string index (t=\"s\")."""
    for key, payload in parts.items():
        if not key.startswith("xl/worksheets/sheet") or not key.endswith(".xml"):
            continue
        if "/_rels/" in key:
            continue
        if b't="s"' in payload or b"t='s'" in payload:
            return True
    return False


def _strip_orphan_shared_strings_package(parts: dict[str, bytes]) -> None:
    if _worksheets_use_shared_string_type(parts):
        return
    parts.pop("xl/sharedStrings.xml", None)
    rk = "xl/_rels/workbook.xml.rels"
    if rk in parts:
        txt = parts[rk].decode("utf-8")
        parts[rk] = _SS_REL_RE.sub("", txt).encode("utf-8")
    ct = "[Content_Types].xml"
    if ct in parts:
        txt = parts[ct].decode("utf-8")
        parts[ct] = _CT_SHAREDSTRINGS_RE.sub("", txt).encode("utf-8")


def _strip_legacy_drawings(parts: dict[str, bytes]) -> None:
    for k in [k for k in parts if k.startswith("xl/drawings/")]:
        del parts[k]

    for key in list(parts):
        if key.startswith("xl/worksheets/_rels/") and key.endswith(".rels"):
            txt = parts[key].decode("utf-8")
            parts[key] = _DRAWING_REL_RE.sub("", txt).encode("utf-8")

    ct = "[Content_Types].xml"
    if ct in parts:
        txt = parts[ct].decode("utf-8")
        parts[ct] = _CT_DRAWING_RE.sub("", txt).encode("utf-8")


def _strip_calc_chain_package(parts: dict[str, bytes]) -> None:
    parts.pop("xl/calcChain.xml", None)
    rk = "xl/_rels/workbook.xml.rels"
    if rk in parts:
        txt = parts[rk].decode("utf-8")
        parts[rk] = _CALCCHAIN_REL_RE.sub("", txt).encode("utf-8")
    ct = "[Content_Types].xml"
    if ct in parts:
        txt = parts[ct].decode("utf-8")
        parts[ct] = _CT_CALCCHAIN_RE.sub("", txt).encode("utf-8")


def _merge_openpyxl_into_template_package(template_path: Path, openpyxl_bytes: bytes) -> bytes:
    if not template_path.is_file():
        return openpyxl_bytes
    with zipfile.ZipFile(template_path, "r") as ztpl:
        template_order = ztpl.namelist()
        template_meta = {zi.filename: zi for zi in ztpl.infolist()}
        parts: dict[str, bytes] = {n: ztpl.read(n) for n in template_order}
        app_xml = parts.get("docProps/app.xml")

    with zipfile.ZipFile(BytesIO(openpyxl_bytes), "r") as zmod:
        mod = {n: zmod.read(n) for n in zmod.namelist()}

    for name, payload in mod.items():
        if name.startswith("xl/worksheets/"):
            parts[name] = payload
        elif name in ("xl/styles.xml", "xl/sharedStrings.xml", "xl/workbook.xml", "xl/_rels/workbook.xml.rels"):
            parts[name] = payload
        elif name == "docProps/core.xml":
            parts[name] = payload

    if app_xml is not None:
        parts["docProps/app.xml"] = app_xml

    _strip_calc_chain_package(parts)
    _strip_legacy_drawings(parts)
    _strip_orphan_shared_strings_package(parts)

    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for name in template_order:
            if name not in parts:
                continue
            data = parts[name]
            src = template_meta.get(name)
            info = zipfile.ZipInfo(filename=name)
            if src is not None:
                info.compress_type = src.compress_type
                info.external_attr = src.external_attr
                info.date_time = src.date_time
            else:
                info.compress_type = zipfile.ZIP_DEFLATED
            zout.writestr(info, data)

    raw = buf.getvalue()
    with zipfile.ZipFile(BytesIO(raw), "r") as zcheck:
        bad = zcheck.testzip()
    if bad is not None:
        raise RuntimeError(f"PIR xlsx failed ZIP integrity check: {bad}")
    return raw


def build_pir_workbook_bytes(cases: list[Case], date_from: date, date_to: date) -> bytes:
    if not TEMPLATE_PATH.is_file():
        raise FileNotFoundError(f"PIR template missing: {TEMPLATE_PATH}")
    wb = load_workbook(TEMPLATE_PATH)
    _unhide_report_sheets(wb)
    _fill_main_sheets(wb, cases)
    _fill_enforcement_sheet(wb, cases, date_from, date_to)
    _fill_debt_sheet(wb, cases, date_from, date_to)
    buf = BytesIO()
    wb.save(buf)
    return _merge_openpyxl_into_template_package(TEMPLATE_PATH, buf.getvalue())
