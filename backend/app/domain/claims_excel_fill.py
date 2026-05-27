"""Заполнение xlsx-реестра претензий по шаблону `claims_registry_template.xlsx`.

Структура шаблона:
- лист `претензии 2025` — данные с 4-й строки
- лист `претензии 2026` — данные с 4-й строки
- листы `в качестве 3 лица`, `исполнительное производство`, `инф по сниже дебит. задолжненно`
  оставлены как пустые шаблоны (соответствуют эталону юриста)

В каждом «годовом» листе колонки A–G:
  A — № п/п
  B — Контрагент (merge по серии претензий с одним контрагентом, как в эталоне)
  C — ИСХ.№
  D — Дата (форматируется как «09.01.2025г.»)
  E — Сущность претензии
  F — Сумма (Тенге)
  G — Статус (каноническое название + детализация через \n, если есть)

После заполнения openpyxl-частей делаем тот же ZIP-merge, что и в pir_excel_fill.py,
чтобы сохранить точную графическую разметку шаблона (шрифты, ширины колонок, рамки).
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.models import Claim
from app.domain.pir_package_xml import merge_openpyxl_into_template_package

TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "templates" / "claims_registry_template.xlsx"

DATA_SHEET_FIRST_ROW = 4
DATA_SHEET_MAX_COL = 7  # A..G

YEAR_SHEET_BY_YEAR = {
    2025: "претензии 2025",
    2026: "претензии 2026",
}

STATUS_LABELS = {
    "collected": "взыскано",
    "not_collected": "не взыскано",
    "offset": "удержано в безакцептном порядке",
    "recalculation": "перерасчёт",
}

_THIN_SIDE = Side(style="thin", color="000000")
_DATA_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_DATA_FONT = Font(size=10, name="Times New Roman")
_ALIGN_TEXT = Alignment(horizontal="left", vertical="top", wrap_text=True)
_ALIGN_NUM = Alignment(horizontal="right", vertical="center", wrap_text=False)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _format_date_ru(d: date) -> str:
    return d.strftime("%d.%m.%Yг.")


def _excel_num(value: object) -> int | float | None:
    """Конвертирует Decimal/число в int/float; пустую/0 оставляет None — как в эталоне."""
    if value is None:
        return None
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    if v == 0.0:
        return None
    if abs(v - int(v)) < 1e-9:
        return int(v)
    return round(v, 2)


def _format_status(claim: Claim) -> str:
    label = STATUS_LABELS.get(claim.status, claim.status or "")
    if claim.status_detail:
        return f"{label}\n{claim.status_detail}".strip()
    return label


def _apply_cell(cell, value, *, numeric: bool = False, center: bool = False) -> None:
    if isinstance(cell, MergedCell):
        return
    cell.value = value
    cell.font = _DATA_FONT
    if center:
        cell.alignment = _ALIGN_CENTER
    elif numeric:
        cell.alignment = _ALIGN_NUM
    else:
        cell.alignment = _ALIGN_TEXT
    cell.border = _DATA_BORDER


def _clear_data_rows(ws: Worksheet) -> None:
    """Снимает merge и очищает все строки начиная с DATA_SHEET_FIRST_ROW."""
    for cr in list(ws.merged_cells.ranges):
        if cr.min_row >= DATA_SHEET_FIRST_ROW:
            ws.unmerge_cells(str(cr))
    last_row = max(ws.max_row, DATA_SHEET_FIRST_ROW)
    for r in range(DATA_SHEET_FIRST_ROW, last_row + 1):
        for c in range(1, DATA_SHEET_MAX_COL + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.border = Border()


def _group_consecutive(rows: list[Claim], key) -> Iterable[tuple[int, int, list[Claim]]]:
    """Группирует подряд идущие записи по ключу. Возвращает (start_row_idx, length, group)."""
    if not rows:
        return
    cur_key = key(rows[0])
    start = 0
    for i in range(1, len(rows)):
        k = key(rows[i])
        if k != cur_key:
            yield start, i - start, rows[start:i]
            start = i
            cur_key = k
    yield start, len(rows) - start, rows[start:]


def _render_year_sheet(ws: Worksheet, claims: list[Claim]) -> None:
    _clear_data_rows(ws)

    # Сортируем по контрагенту, потом по дате
    claims_sorted = sorted(claims, key=lambda c: (c.counterparty_name or "", c.claim_date))

    # Группировка по контрагенту для merge B-колонки
    counter_groups = list(_group_consecutive(claims_sorted, key=lambda c: c.counterparty_name or ""))

    seq = 1
    for offset_idx, group_len, group in counter_groups:
        first_row = DATA_SHEET_FIRST_ROW + offset_idx
        last_row = first_row + group_len - 1

        # Колонка B — контрагент только в первой строке, остальные пустые, потом merge
        _apply_cell(ws.cell(row=first_row, column=2), group[0].counterparty_name or "")
        if group_len > 1:
            for r in range(first_row + 1, last_row + 1):
                _apply_cell(ws.cell(row=r, column=2), None)
            ws.merge_cells(start_row=first_row, end_row=last_row, start_column=2, end_column=2)

        # Внутри группы — статусы тоже могут объединяться по подгруппам с одинаковым статусом + detail
        sub_groups = list(_group_consecutive(
            group, key=lambda c: (c.status, c.status_detail or "")
        ))
        for sub_offset, sub_len, sub_group in sub_groups:
            sub_first = first_row + sub_offset
            sub_last = sub_first + sub_len - 1
            _apply_cell(ws.cell(row=sub_first, column=7), _format_status(sub_group[0]))
            if sub_len > 1:
                for r in range(sub_first + 1, sub_last + 1):
                    _apply_cell(ws.cell(row=r, column=7), None)
                ws.merge_cells(start_row=sub_first, end_row=sub_last, start_column=7, end_column=7)

        # Построчные поля (№, ИСХ.№, дата, сущность, сумма)
        for i, claim in enumerate(group):
            r = first_row + i
            _apply_cell(ws.cell(row=r, column=1), seq, center=True)
            _apply_cell(ws.cell(row=r, column=3), claim.outgoing_number or "")
            _apply_cell(ws.cell(row=r, column=4), _format_date_ru(claim.claim_date), center=True)
            _apply_cell(ws.cell(row=r, column=5), claim.subject or "")
            amount = _excel_num(claim.amount)
            num_cell = ws.cell(row=r, column=6)
            _apply_cell(num_cell, amount, numeric=True)
            if amount is not None:
                num_cell.number_format = "#,##0.00"
            seq += 1


def build_claims_workbook_bytes(claims: list[Claim], date_from: date, date_to: date) -> bytes:
    """Главная точка входа: формирует bytes XLSX-файла со всеми листами как в эталоне."""
    if not TEMPLATE_PATH.is_file():
        raise FileNotFoundError(f"Claims template missing: {TEMPLATE_PATH}")

    wb: Workbook = load_workbook(TEMPLATE_PATH)

    # Фильтр по периоду
    in_period = [
        c for c in claims
        if date_from <= c.claim_date <= date_to
    ]

    # Разнесём по годам
    by_year: dict[int, list[Claim]] = {}
    for c in in_period:
        by_year.setdefault(c.claim_date.year, []).append(c)

    for year, sheet_name in YEAR_SHEET_BY_YEAR.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        _render_year_sheet(ws, by_year.get(year, []))

    # Сохраним результат и пропустим через ZIP-merge для сохранения форматирования
    buf = BytesIO()
    wb.save(buf)
    return merge_openpyxl_into_template_package(TEMPLATE_PATH, buf.getvalue())
