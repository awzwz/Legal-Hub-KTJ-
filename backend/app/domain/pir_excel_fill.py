"""Fill PIR xlsx from in-memory Case rows (no FastAPI/JWT imports — testable in isolation).

Раскладка строго следует шаблону КТЖ:

* Лист **«истец»** — категории «Закупки/договоры», «Трудовые», «Иные», «Медиативные».
  Перевозочные споры на этом листе не предусмотрены — сваливаются в «Иные».
* Лист **«ответчик»** — добавляется «Иски, вытекающие из перевозочного процесса».
* Лист **«в качестве 3 лица»** — реестр дел, где КТЖ-ГП выступает третьим лицом.
* Лист **«3-лицо »** (с пробелом, КТЖ — истец c третьими лицами) — дублирует строки
  третьего лица для совместимости с шаблоном.
* Лист **«исполнительное производство»** — построчные записи `enforcement_proceedings`.
* Лист **«инф по сниже дебит. задолжненно»** — `debt_recovery_entries`.

Колонки внутри листов соответствуют шапке шаблона: см. `_write_*_row` ниже. Нумерация
дел идёт **отдельно для каждой категории** (как в эталонном файле).
"""

from __future__ import annotations

import math
import re
from copy import copy
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Iterable, Sequence

from app.domain.pir_package_xml import merge_openpyxl_into_template_package

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.models import Case, DebtRecoveryEntry, EnforcementProceeding

TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "templates" / "pir_report_2025_template.xlsx"

# party_role → лист, в который пишется первичная запись
SHEET_BY_ROLE = {
    "plaintiff": "истец",
    "defendant": "ответчик",
    "third_party": "в качестве 3 лица",
}

# Резервный лист для роли "third_party" (КТЖ — истец, но дело с 3-ми лицами):
# дублируем строки сюда, чтобы шаблонный реестр не оставался пустым.
THIRD_PARTY_MIRROR_SHEET = "3-лицо "

ENFORCEMENT_SHEET = "исполнительное производство"
ENFORCEMENT_FIRST_ROW = 6

DEBT_SHEET = "инф по сниже дебит. задолжненно"
DEBT_FIRST_ROW = 7

# Старые константы (используются в тестах/интеграциях): первая строка данных
# первой категории на основном листе. Сохраняем имена ради совместимости.
FIRST_DATA_ROW = {
    "истец": 8,
    "ответчик": 7,
    "3-лицо ": 7,
    "в качестве 3 лица": 7,
}

# Где заканчивается «фиксированная» шапка — выше этой строки openpyxl не должен
# трогать ни значения, ни merge.
HEADER_LAST_ROW = {
    "истец": 6,
    "ответчик": 5,
    "3-лицо ": 6,
    "в качестве 3 лица": 6,
}

MAIN_MAX_COL = {
    "истец": 19,
    "ответчик": 19,
    "3-лицо ": 19,
    "в качестве 3 лица": 19,
}

# Порядок категорий внутри каждого «ролевого» листа: совпадает с шаблоном КТЖ.
# Названия идентичны эталону (опечатка «законадательства» в «ответчик» исправлена;
# юристы предпочли корректное написание).
PLAINTIFF_CATEGORY_ORDER: list[tuple[str, str]] = [
    ("procurement", "Иски, связанные с нарушением законодательства о закупках и вытекающие из договоров"),
    ("labor", "Трудовые споры"),
    ("government", "Иски, вытекающие из споров с госорганами"),
    ("other", "Иные споры"),
    ("mediation", "Медиативные соглашения"),
]
DEFENDANT_CATEGORY_ORDER: list[tuple[str, str]] = [
    ("procurement", "Иски, связанные с нарушением законодательства о закупках и вытекающие из договоров"),
    ("labor", "Трудовые споры"),
    ("transportation", "Иски, вытекающие из перевозочного процесса"),
    ("other", "Иные споры"),
    ("mediation", "Медиативные соглашения"),
]
THIRD_PARTY_CATEGORY_ORDER: list[tuple[str, str]] = [
    ("procurement", "Иски, связанные с нарушением законодательства о закупках и вытекающие из договоров"),
    ("labor", "Трудовые споры"),
    ("government", "Иски, вытекающие из споров с госорганами"),
    ("other", "Иные споры"),
    ("mediation", "Медиативные соглашения"),
]

CATEGORY_ORDER_BY_SHEET: dict[str, list[tuple[str, str]]] = {
    "истец": PLAINTIFF_CATEGORY_ORDER,
    "ответчик": DEFENDANT_CATEGORY_ORDER,
    "3-лицо ": THIRD_PARTY_CATEGORY_ORDER,
    "в качестве 3 лица": THIRD_PARTY_CATEGORY_ORDER,
}


def _allowed_categories(sheet_name: str) -> set[str]:
    return {cat for cat, _ in CATEGORY_ORDER_BY_SHEET.get(sheet_name, [])}


def _normalize_category(case: Case, sheet_name: str) -> str:
    """Приводит ``case.dispute_category`` к одной из категорий целевого листа.

    Шаблон КТЖ имеет 5 секций на каждый лист (procurement, labor,
    government/transportation, other, mediation). Маршрутизация:
      - settled outcome или mediation → «Медиативные соглашения»
      - transportation на «истец»/«3-лицо» → «Иные споры» (transportation — это
        defendant-only секция «перевозочный процесс»)
      - government на «ответчик» → «Иные споры» (нет секции «госорганы» у ответчика)
      - other остаётся в «Иные споры» — есть на всех листах.
    """
    raw = (getattr(case, "dispute_category", None) or "").strip().lower()
    out = (getattr(case, "outcome", None) or "").strip().lower()
    allowed = _allowed_categories(sheet_name)
    # Settled / mediation outcome → раздел «Медиативные соглашения»
    if (out == "settled" or raw == "mediation") and "mediation" in allowed:
        return "mediation"
    if raw in allowed:
        return raw
    # Перевозочный процесс есть только на «ответчик»; на других листах — к «Иные споры»
    if raw == "transportation":
        if "transportation" in allowed:
            return "transportation"
        if "other" in allowed:
            return "other"
    # «government» на «ответчик» — нет такой секции, относим к «Иные споры»
    if raw == "government" and "government" not in allowed:
        if "other" in allowed:
            return "other"
        if "procurement" in allowed:
            return "procurement"
    # Fallback — «Иные споры», иначе procurement
    if "other" in allowed:
        return "other"
    if "procurement" in allowed:
        return "procurement"
    return next(iter(allowed)) if allowed else "procurement"


_ILLEGAL_XML_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _safe_txt(value: str | None) -> str:
    if not value:
        return ""
    return _ILLEGAL_XML_CHARS.sub("", str(value))


def _fnum(value: object) -> float:
    try:
        v = float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return 0.0
    if math.isnan(v) or math.isinf(v):
        return 0.0
    return v


def _excel_num(value: object) -> int | float | None:
    """Как в типовом ПИР-xlsx: пустая ячейка вместо нуля для сумм (импорт даёт Decimal(0))."""
    v = _fnum(value)
    if v == 0.0:
        return None
    if abs(v - int(v)) < 1e-9:
        return int(v)
    return round(v, 6)


def _outcome_bucket(case: Case) -> str:
    """Куда дело попадает в сводном блоке внизу листа."""
    cat = (getattr(case, "dispute_category", None) or "").strip().lower()
    out = (case.outcome or "").strip().lower()
    if cat == "mediation" or out == "settled":
        return "mediation"
    if out in {"fully_satisfied", "partially_satisfied"}:
        # Неденежные иски (claim=0, recovered=0) — например, обжалование действий ЧСИ —
        # юристы КТЖ не считают «удовлетворёнными» в сводке для ответчиков и третьих лиц, относят к «отказано».
        claim = _claim_total(case)
        recovered = _recovered_total(case)
        role = (case.party_role or "").strip().lower()
        if role != "plaintiff" and claim == 0 and recovered == 0:
            return "denied"
        return "satisfied"
    if out in {"denied", "dismissed"}:
        return "denied"
    return "filed_only"


def _claim_total(case: Case) -> Decimal:
    fin = case.finances
    if fin is None:
        return Decimal("0")
    val = fin.claim_amount
    return val if isinstance(val, Decimal) else Decimal(str(val or 0))


def _recovered_total(case: Case) -> Decimal:
    """Сумма «взысканная» — main + fines + state_fee (+ rep_expenses для ответчика).

    В эталоне юрист считает категории «удовлетворено» / «медиативные» в сводке через
    `=SUM(M..O)` — взысканную сумму, не предъявленную.
    """
    fin = case.finances
    if fin is None:
        return Decimal("0")
    parts = [fin.recovered_main, fin.recovered_fines, fin.recovered_state_fee, fin.recovered_rep_expenses]
    total = Decimal("0")
    for p in parts:
        if p is None:
            continue
        total += p if isinstance(p, Decimal) else Decimal(str(p))
    return total


def _unmerge_in_range(ws: Worksheet, start_row: int, end_row: int | None = None) -> None:
    """Снять merge у диапазонов, целиком или частично попадающих в [start_row..end_row]."""
    for cr in list(ws.merged_cells.ranges):
        if cr.max_row < start_row:
            continue
        if end_row is not None and cr.min_row > end_row:
            continue
        ws.unmerge_cells(str(cr))


def _clear_rows(ws: Worksheet, start_row: int, end_row: int, max_col: int) -> None:
    """Очистить значения, заливку и рамки у каждой ячейки в диапазоне."""
    from openpyxl.styles import PatternFill as _PF
    no_fill = _PF(fill_type=None)
    no_border = Border()
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            try:
                cell.fill = no_fill
            except Exception:
                pass
            try:
                cell.border = no_border
            except Exception:
                pass


def _court_block(case: Case) -> str:
    court = _safe_txt(case.court)
    j = _safe_txt(case.judge).strip()
    if not court:
        return ""
    if not j or j in ("—", "–", "-"):
        return court
    return _safe_txt(f"{court}\n{j}".strip())


# ---------------------------------------------------------------------------
# Стили строк-разделителей категорий и шапки авто-сводки
# ---------------------------------------------------------------------------

_CATEGORY_FONT = Font(bold=True, size=11, name="Times New Roman")
_CATEGORY_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
_THIN_SIDE = Side(style="thin", color="000000")
_DATA_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_DATA_ALIGN_TEXT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_ALIGN_NUM = Alignment(horizontal="center", vertical="center", wrap_text=False)
_DATA_FONT = Font(size=10, name="Times New Roman")
_SUMMARY_HEADER_FONT = Font(bold=True, size=11, name="Times New Roman")
_SUMMARY_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TOTAL_FONT = Font(bold=True, size=10, name="Times New Roman")
# Голубая заливка для итоговой ячейки группы (как в эталоне)
from openpyxl.styles import PatternFill as _PatternFill
_TOTAL_HIGHLIGHT_FILL = _PatternFill(fill_type="solid", start_color="DDEBF7", end_color="DDEBF7")

# Единый формат денежных значений: «127 852 355,19» (в RU-локали Excel/Numbers/Sheets).
_MONEY_FORMAT = "#,##0.00"

# Схема итоговых строк по листу:
#   filed_cols  — столбцы предъявленной суммы (F/G/H для истца, F/G/H/I для ответчика…)
#   filed_total — столбец «Итого предъявленная» (=SUM по filed_cols)
#   recov_cols  — столбцы взысканной суммы
#   recov_total — столбец «Итого взысканная»
_GROUP_TOTALS_SCHEMA: dict[str, dict] = {
    # Истец: F/G/H + I (итого предъявл) и M/N/O + P (итого взыск)
    "истец": {
        "filed_cols": [6, 7, 8], "filed_total": 9,
        "recov_cols": [13, 14, 15], "recov_total": 16,
    },
    # Ответчик: F/G/H/I + J (итого предъявл) и N/O/P/Q + R (итого взыск)
    "ответчик": {
        "filed_cols": [6, 7, 8, 9], "filed_total": 10,
        "recov_cols": [14, 15, 16, 17], "recov_total": 18,
    },
    # «3-лицо» имеет 4 предъявленных (G/H/I/J + K-итого) и 4 взысканных (O/P/Q/R + S-итого);
    # делаем по аналогии — отдельная проверка по полям шаблона
    "3-лицо ": {
        "filed_cols": [7, 8, 9, 10], "filed_total": 11,
        "recov_cols": [15, 16, 17, 18], "recov_total": 19,
    },
    "в качестве 3 лица": {
        "filed_cols": [7, 8, 9, 10], "filed_total": 11,
        "recov_cols": [15, 16, 17, 18], "recov_total": 19,
    },
}


def _col_letter(c: int) -> str:
    """Convert 1-based column index to Excel letters (A, B, ..., Z, AA, ...)."""
    s = ""
    while c > 0:
        c, r = divmod(c - 1, 26)
        s = chr(65 + r) + s
    return s


def _sum_col_range(ws: Worksheet, col: int, first_row: int, last_row: int) -> Decimal:
    """Сумма числовых значений в колонке (для pre-computed cached values)."""
    total = Decimal("0")
    for r in range(first_row, last_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is None:
            continue
        if isinstance(v, (int, float, Decimal)):
            total += Decimal(str(v))
        elif isinstance(v, str) and not v.startswith("="):
            try:
                cleaned = v.replace("\xa0", "").replace(" ", "").replace(",", ".")
                total += Decimal(cleaned)
            except Exception:
                pass
    return total


def _write_group_totals(ws: Worksheet, sheet_name: str, first_data_row: int, last_data_row: int, total_row: int) -> None:
    """Записать строку-итог по группе дел: формулы SUM по столбцам предъявленной и взысканной сумм.

    Эталон делает то же самое: F37=SUM(F8:F36), G37=SUM(G8:G36) … I37=SUM(F37:H37) — выделено цветом.

    Здесь записываем pre-computed числовое значение (openpyxl не вычисляет формулы;
    Excel/Numbers пересчитают формулы при открытии, но cached-значение нужно для просмотра без открытия.).
    """
    schema = _GROUP_TOTALS_SCHEMA.get(sheet_name)
    if schema is None or first_data_row > last_data_row:
        return

    def _write_total_cell(col: int, value, highlight: bool = False) -> None:
        cell = ws.cell(row=total_row, column=col)
        if isinstance(cell, MergedCell):
            return
        # Если value — это Decimal/число, конвертируем в float; если строка-формула — пишем как есть.
        if isinstance(value, Decimal):
            cell.value = float(value)
        else:
            cell.value = value
        cell.font = _TOTAL_FONT
        cell.alignment = _DATA_ALIGN_NUM
        cell.border = _DATA_BORDER
        cell.number_format = _MONEY_FORMAT
        if highlight:
            cell.fill = _TOTAL_HIGHLIGHT_FILL

    # Предъявленная: pre-computed numeric sum по каждому столбцу
    filed_total_value = Decimal("0")
    for col in schema["filed_cols"]:
        s = _sum_col_range(ws, col, first_data_row, last_data_row)
        _write_total_cell(col, s)
        filed_total_value += s
    if schema.get("filed_total"):
        _write_total_cell(schema["filed_total"], filed_total_value, highlight=True)

    # Взысканная: аналогично
    recov_total_value = Decimal("0")
    for col in schema["recov_cols"]:
        s = _sum_col_range(ws, col, first_data_row, last_data_row)
        _write_total_cell(col, s)
        recov_total_value += s
    if schema.get("recov_total"):
        _write_total_cell(schema["recov_total"], recov_total_value, highlight=True)

    # Компактная высота итоговой строки (в эталоне строки ИТОГО ~18-22 пункта,
    # без auto-resize'а, чтобы не наследовать высоту от длинных текстовых ячеек дел).
    ws.row_dimensions[total_row].height = 22

# Какие колонки на каждом из ролевых листов считаются «денежными» — к ним применяется _MONEY_FORMAT.
# Номера соответствуют порядку cells-tuple в writers (см. _write_*_row).
_MONEY_COLS_BY_SHEET: dict[str, set[int]] = {
    "истец": {6, 7, 8, 13, 14, 15},
    "ответчик": {6, 7, 8, 9, 14, 15, 16, 17},
    "3-лицо ": {7, 8, 9, 10, 15, 16, 17, 18},
    "в качестве 3 лица": {7, 8, 9, 10, 15, 16, 17, 18},
}


def _apply_category_row(ws: Worksheet, row_idx: int, max_col: int, text: str) -> None:
    cell = ws.cell(row=row_idx, column=1)
    cell.value = text
    cell.font = _CATEGORY_FONT
    cell.alignment = _CATEGORY_ALIGN
    cell.border = _DATA_BORDER
    if max_col >= 2:
        merge_ref = f"A{row_idx}:{ws.cell(row=row_idx, column=max_col).coordinate}"
        ws.merge_cells(merge_ref)
    ws.row_dimensions[row_idx].height = 28


def _apply_data_cell(cell, value, *, numeric: bool) -> None:
    cell.value = value
    cell.font = _DATA_FONT
    cell.alignment = _DATA_ALIGN_NUM if numeric else _DATA_ALIGN_TEXT
    cell.border = _DATA_BORDER


# ---------------------------------------------------------------------------
# Запись строк по ролям
# ---------------------------------------------------------------------------


# Листы, на которых в эталоне колонка «Исполнитель» (B) пустая —
# КТЖ-юрист как ответчик / привлечённое 3-е лицо в оригинале не указывается.
SHEETS_OMIT_EXECUTOR = {"в качестве 3 лица"}


@dataclass
class _RowCtx:
    ws: Worksheet
    row_idx: int
    seq: int
    case: Case
    omit_executor: bool = False


def _write_истец_row(ctx: _RowCtx) -> None:
    case = ctx.case
    fin = case.finances
    lit = case.litigation
    cells: list[tuple[int, object, bool]] = [
        (1, ctx.seq, True),
        (2, _safe_txt(case.assigned_lawyer.full_name if case.assigned_lawyer else ""), False),
        (3, _court_block(case), False),
        (4, _safe_txt(case.branch.name if case.branch else ""), False),
        (5, _safe_txt(case.defendant), False),
        (6, _excel_num(fin.main_debt) if fin else None, True),
        (7, _excel_num(fin.fines) if fin else None, True),
        (8, _excel_num(fin.state_fee) if fin else None, True),
        (9, _safe_txt((lit.claim_summary if lit else "") or ""), False),
        (10, _safe_txt((lit.judgment_first if lit else "") or ""), False),
        (11, _safe_txt((lit.judgment_appeal if lit else "") or ""), False),
        (12, _safe_txt((lit.judgment_cassation if lit else "") or ""), False),
        (13, _excel_num(fin.recovered_main) if fin else None, True),
        (14, _excel_num(fin.recovered_fines) if fin else None, True),
        (15, _excel_num(fin.recovered_state_fee) if fin else None, True),
        (16, _safe_txt((lit.writ_request_note if lit else "") or ""), False),
        (17, _safe_txt((lit.writ_dispatch_note if lit else "") or ""), False),
        (18, _safe_txt((lit.execution_proof_note if lit else "") or ""), False),
        (19, _safe_txt((lit.damage_recovery_note if lit else "") or ""), False),
    ]
    for col, value, numeric in cells:
        _apply_data_cell(ctx.ws.cell(row=ctx.row_idx, column=col), value, numeric=numeric)


def _write_ответчик_row(ctx: _RowCtx) -> None:
    case = ctx.case
    fin = case.finances
    lit = case.litigation
    executor = "" if ctx.omit_executor else _safe_txt(case.assigned_lawyer.full_name if case.assigned_lawyer else "")
    cells: list[tuple[int, object, bool]] = [
        (1, ctx.seq, True),
        (2, executor, False),
        (3, _court_block(case), False),
        (4, _safe_txt(case.branch.name if case.branch else ""), False),
        (5, _safe_txt(case.plaintiff), False),
        (6, _excel_num(fin.main_debt) if fin else None, True),
        (7, _excel_num(fin.fines) if fin else None, True),
        (8, _excel_num(fin.rep_expenses) if fin else None, True),
        (9, _excel_num(fin.state_fee) if fin else None, True),
        (10, _safe_txt((lit.claim_summary if lit else "") or ""), False),
        (11, _safe_txt((lit.judgment_first if lit else "") or ""), False),
        (12, _safe_txt((lit.judgment_appeal if lit else "") or ""), False),
        (13, _safe_txt((lit.judgment_cassation if lit else "") or ""), False),
        # ВЗЫСКАННАЯ сумма: основная / штрафы / представительские / госпошлина
        (14, _excel_num(fin.recovered_main) if fin else None, True),
        (15, _excel_num(fin.recovered_fines) if fin else None, True),
        (16, _excel_num(fin.recovered_rep_expenses) if fin else None, True),
        (17, _excel_num(fin.recovered_state_fee) if fin else None, True),
        (18, _safe_txt((lit.defendant_execution_note if lit else "") or ""), False),
        (19, _safe_txt((lit.damage_recovery_note if lit else "") or ""), False),
    ]
    for col, value, numeric in cells:
        _apply_data_cell(ctx.ws.cell(row=ctx.row_idx, column=col), value, numeric=numeric)


def _write_third_party_row(ctx: _RowCtx) -> None:
    """Общий формат для листов «3-лицо » и «в качестве 3 лица»."""
    case = ctx.case
    fin = case.finances
    lit = case.litigation
    executor = "" if ctx.omit_executor else _safe_txt(case.assigned_lawyer.full_name if case.assigned_lawyer else "")
    cells: list[tuple[int, object, bool]] = [
        (1, ctx.seq, True),
        (2, executor, False),
        (3, _court_block(case), False),
        (4, _safe_txt(case.branch.name if case.branch else ""), False),
        (5, _safe_txt(case.plaintiff), False),
        (6, _safe_txt(case.defendant), False),
        (7, _excel_num(fin.main_debt) if fin else None, True),
        (8, _excel_num(fin.fines) if fin else None, True),
        (9, _excel_num(fin.rep_expenses) if fin else None, True),
        (10, _excel_num(fin.state_fee) if fin else None, True),
        (11, _safe_txt((lit.claim_summary if lit else "") or ""), False),
        (12, _safe_txt((lit.judgment_first if lit else "") or ""), False),
        (13, _safe_txt((lit.judgment_appeal if lit else "") or ""), False),
        (14, _safe_txt((lit.judgment_cassation if lit else "") or ""), False),
        # ВЗЫСКАННАЯ сумма
        (15, _excel_num(fin.recovered_main) if fin else None, True),
        (16, _excel_num(fin.recovered_fines) if fin else None, True),
        (17, _excel_num(fin.recovered_rep_expenses) if fin else None, True),
        (18, _excel_num(fin.recovered_state_fee) if fin else None, True),
        (19, _safe_txt((lit.third_party_note if lit else "") or ""), False),
    ]
    for col, value, numeric in cells:
        _apply_data_cell(ctx.ws.cell(row=ctx.row_idx, column=col), value, numeric=numeric)


_ROW_WRITER_BY_SHEET = {
    "истец": _write_истец_row,
    "ответчик": _write_ответчик_row,
    "3-лицо ": _write_third_party_row,
    "в качестве 3 лица": _write_third_party_row,
}


# ---------------------------------------------------------------------------
# Сводный блок (предъявлено / удовлетворено / медиативные / отказано)
# ---------------------------------------------------------------------------


def _write_summary_block(
    ws: Worksheet,
    start_row: int,
    max_col: int,
    cases: Sequence[Case],
    sheet_name: str | None = None,
    category_ranges: dict[str, tuple[int, int, int]] | None = None,
) -> int:
    """Пишет сводку в формате шаблона КТЖ (предъявлено / удовлетворено / медиативные / отказано).

    Колонки: B — название, C — кол-во, D — % количества, E — сумма (₸), F — % суммы.
    Проценты — формулы Excel (=C{row}/C{filed}*100), чтобы итоги пересчитывались автоматически.

    Если переданы `category_ranges` (см. `_render_role_sheet`), сумма «удовлетворено»
    и «медиативные» вычисляются формулами Excel, ссылающимися на итоговые ячейки
    взысканной суммы (P{totals_row} / R{totals_row}) — как в эталоне (`=P37` / `=P45`).
    «Отказано» — остаточная формула `=E{filed} - E{satisfied} - E{mediation}`.
    """
    # Логика эталона КТЖ (см. формулы G50..G53 в исходном Q1 2026 отчёте):
    # - «Предъявлено» = сумма claim_amount по всем делам категории
    # - «Удовлетворено» = сумма ВЗЫСКАННОЙ суммы (recovered_*) по satisfied+denied (procurement раздел)
    # - «Медиативные»   = сумма ВЗЫСКАННОЙ суммы по mediation
    # - «Отказано»      = «Предъявлено» − «Удовлетворено» − «Медиативные» (остаточный принцип)
    counts = {"filed": 0, "satisfied": 0, "mediation": 0, "denied": 0}
    sums = {"filed": Decimal("0"), "recovered_procurement": Decimal("0"), "recovered_mediation": Decimal("0")}
    voluntary_paid_total = Decimal("0")  # ручные корректировки Малики (добровольно выплаченные суммы)
    for c in cases:
        bucket = _outcome_bucket(c)
        claim = _claim_total(c)
        recovered = _recovered_total(c)
        counts["filed"] += 1
        sums["filed"] += claim
        # paid_amount — добровольные выплаты вне взысканной суммы (например,
        # «добровольно выплаченная сумма ТОО Жарас» в эталонной ПИР). Малика
        # относит такие суммы к разделу «Медиативные» в сводном блоке.
        try:
            paid = Decimal(str(c.finances.paid_amount)) if c.finances and c.finances.paid_amount else Decimal("0")
        except (AttributeError, ValueError):
            paid = Decimal("0")
        voluntary_paid_total += paid
        if bucket == "mediation":
            counts["mediation"] += 1
            sums["recovered_mediation"] += recovered
        else:
            # Все «не-медиативные» дела вместе образуют группу procurement-итогов
            # (как в эталоне R37: SUM(M8:O36) включает и satisfied, и denied).
            sums["recovered_procurement"] += recovered
            if bucket == "satisfied":
                counts["satisfied"] += 1
            elif bucket == "denied":
                counts["denied"] += 1
    sum_filed = sums["filed"]
    sum_satisfied = sums["recovered_procurement"]
    sum_mediation = sums["recovered_mediation"] + voluntary_paid_total
    sum_denied = sum_filed - sum_satisfied - sum_mediation

    header_row = start_row + 1
    headers = ("Категория", "Кол-во", "% кол-ва", "Сумма (₸)", "% суммы")
    for offset, text in enumerate(headers):
        cell = ws.cell(row=header_row, column=2 + offset)
        cell.value = text
        cell.font = _SUMMARY_HEADER_FONT
        cell.alignment = _SUMMARY_HEADER_ALIGN
        cell.border = _DATA_BORDER

    # Записываем pre-computed значения (numeric) — openpyxl не вычисляет формулы;
    # Excel пересчитает при открытии, но и без пересчёта пользователь видит верные числа.
    rows_data = [
        ("предъявлено", counts["filed"], sum_filed, None),
        ("удовлетворено", counts["satisfied"], sum_satisfied, None),
        ("заключены медиативные соглашения", counts["mediation"], sum_mediation, None),
        ("отказано", counts["denied"], sum_denied, None),
    ]
    filed_row = header_row + 1  # «предъявлено» — база для расчёта процентов
    for i, (label, count, total, formula) in enumerate(rows_data, start=1):
        r = header_row + i
        # B: название
        c_label = ws.cell(row=r, column=2)
        c_label.value = label
        c_label.font = _DATA_FONT
        c_label.alignment = _DATA_ALIGN_TEXT
        c_label.border = _DATA_BORDER
        # C: количество
        c_count = ws.cell(row=r, column=3)
        c_count.value = count
        c_count.font = _DATA_FONT
        c_count.alignment = _DATA_ALIGN_NUM
        c_count.border = _DATA_BORDER
        # D: % количества (формула относительно «предъявлено»)
        c_count_pct = ws.cell(row=r, column=4)
        if r == filed_row:
            c_count_pct.value = 100 if counts["filed"] > 0 else None
        else:
            c_count_pct.value = f"=IFERROR(C{r}/C{filed_row}*100,0)"
            c_count_pct.number_format = "0.00"
        c_count_pct.font = _DATA_FONT
        c_count_pct.alignment = _DATA_ALIGN_NUM
        c_count_pct.border = _DATA_BORDER
        # E: сумма — pre-computed numeric value
        c_sum = ws.cell(row=r, column=5)
        c_sum.value = _excel_num(total)
        c_sum.font = _DATA_FONT
        c_sum.alignment = _DATA_ALIGN_NUM
        c_sum.border = _DATA_BORDER
        c_sum.number_format = _MONEY_FORMAT
        # F: % суммы
        c_sum_pct = ws.cell(row=r, column=6)
        if r == filed_row:
            c_sum_pct.value = 100 if sums["filed"] > 0 else None
        else:
            c_sum_pct.value = f"=IFERROR(E{r}/E{filed_row}*100,0)"
            c_sum_pct.number_format = "0.00"
        c_sum_pct.font = _DATA_FONT
        c_sum_pct.alignment = _DATA_ALIGN_NUM
        c_sum_pct.border = _DATA_BORDER
        # Компактная высота для строк сводки (по умолчанию Excel делает строки в Times New Roman ~30pt)
        ws.row_dimensions[r].height = 20
    # И для строки шапки сводки тоже
    ws.row_dimensions[header_row].height = 24

    return header_row + len(rows_data) + 1


# ---------------------------------------------------------------------------
# Основная отрисовка ролевого листа
# ---------------------------------------------------------------------------


def _render_role_sheet(wb: Workbook, sheet_name: str, cases: Iterable[Case]) -> None:
    """Перерисовать ролевой лист: шапку оставляем, тело и сводку строим заново.

    ``FIRST_DATA_ROW`` указывает на первую строку **данных** первой категории —
    непосредственно над ней размещается строка-разделитель ``first_data_row - 1``
    (как в шаблоне КТЖ).
    """
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    max_col = MAIN_MAX_COL[sheet_name]
    header_last = HEADER_LAST_ROW[sheet_name]
    first_data_row = FIRST_DATA_ROW[sheet_name]
    writer = _ROW_WRITER_BY_SHEET[sheet_name]
    cases_list = list(cases)

    clear_until = max(ws.max_row, first_data_row + len(cases_list) + 25)
    _unmerge_in_range(ws, header_last + 1, None)
    _clear_rows(ws, header_last + 1, clear_until, max_col)

    by_category: dict[str, list[Case]] = {}
    for c in cases_list:
        cat = _normalize_category(c, sheet_name)
        by_category.setdefault(cat, []).append(c)

    # Стабильный порядок дел внутри категории — по case_number, затем по filing_date
    for items in by_category.values():
        items.sort(key=lambda x: (x.case_number or "", x.filing_date or date.min))

    # Первая строка-разделитель категории идёт на (first_data_row - 1), как в шаблоне.
    row_idx = first_data_row - 1
    # Запомним диапазоны строк каждой категории для последующего использования в сводке.
    category_ranges: dict[str, tuple[int, int, int]] = {}  # cat → (first_data_row, last_data_row, totals_row)
    for cat, title in CATEGORY_ORDER_BY_SHEET[sheet_name]:
        items = by_category.get(cat, [])
        if not items:
            continue
        _apply_category_row(ws, row_idx, max_col, title)
        row_idx += 1
        omit_exec = sheet_name in SHEETS_OMIT_EXECUTOR
        money_cols = _MONEY_COLS_BY_SHEET.get(sheet_name, set())
        category_first_row = row_idx
        for seq, case in enumerate(items, start=1):
            writer(_RowCtx(ws=ws, row_idx=row_idx, seq=seq, case=case, omit_executor=omit_exec))
            for col in money_cols:
                ws.cell(row=row_idx, column=col).number_format = _MONEY_FORMAT
            row_idx += 1
        category_last_row = row_idx - 1
        # Итоговая строка по группе: формулы SUM по столбцам + выделенный «общий итог»
        totals_row = row_idx
        _write_group_totals(ws, sheet_name, category_first_row, category_last_row, totals_row)
        category_ranges[cat] = (category_first_row, category_last_row, totals_row)
        row_idx += 1

    # Сводка ниже данных (через одну пустую строку)
    if cases_list:
        _write_summary_block(ws, row_idx + 1, max_col, cases_list, sheet_name, category_ranges)


def _horizontalize_main_sheet_headers(wb: Workbook) -> None:
    """После round-trip openpyxl+merge шапка основных листов часто становится «вертикальной».

    В эталонном xlsx заголовки колонок горизонтальны; выравниваем шапку до первой строки данных.
    """
    hdr_align = Alignment(text_rotation=0, horizontal="center", vertical="center", wrap_text=True)
    for sheet_name, last_header in HEADER_LAST_ROW.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        max_col = MAIN_MAX_COL.get(sheet_name, 19)
        for r in range(1, last_header + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                if cell.alignment is None:
                    cell.alignment = hdr_align
                else:
                    cell.alignment = copy(hdr_align)


def _fill_main_sheets(wb: Workbook, cases: list[Case]) -> None:
    plaintiff_cases: list[Case] = []
    defendant_cases: list[Case] = []
    third_party_cases: list[Case] = []
    for c in cases:
        role = (c.party_role or "").strip().lower()
        if role == "plaintiff":
            plaintiff_cases.append(c)
        elif role == "defendant":
            defendant_cases.append(c)
        elif role == "third_party":
            third_party_cases.append(c)

    _render_role_sheet(wb, "истец", plaintiff_cases)
    _render_role_sheet(wb, "ответчик", defendant_cases)
    # third_party дублируется в оба листа — выбран пользователем сознательно.
    _render_role_sheet(wb, "в качестве 3 лица", third_party_cases)
    _render_role_sheet(wb, THIRD_PARTY_MIRROR_SHEET, third_party_cases)

    _horizontalize_main_sheet_headers(wb)


# ---------------------------------------------------------------------------
# Листы исполнительного производства и дебиторки (без категорий)
# ---------------------------------------------------------------------------


def _enforcement_rows_in_period(case: Case, date_from: date, date_to: date) -> list[EnforcementProceeding]:
    return [e for e in case.enforcement_proceedings if date_from <= e.recorded_at <= date_to]


def _debt_rows_in_period(case: Case, date_from: date, date_to: date) -> list[DebtRecoveryEntry]:
    return [
        e for e in case.debt_recovery_entries if e.case_id == case.id and date_from <= e.recorded_at <= date_to
    ]


def _fill_enforcement_sheet(wb: Workbook, cases: list[Case], date_from: date, date_to: date) -> None:
    ws = wb[ENFORCEMENT_SHEET]
    start = ENFORCEMENT_FIRST_ROW
    _unmerge_in_range(ws, start, None)
    _clear_rows(ws, start, max(ws.max_row, start + 50), max_col=14)

    rows: list[EnforcementProceeding] = []
    for c in cases:
        rows.extend(_enforcement_rows_in_period(c, date_from, date_to))
    rows.sort(key=lambda x: (x.recorded_at, str(x.id)))

    enforcement_money_cols = {5, 6, 7, 8, 10, 12}
    for i, e in enumerate(rows, start=1):
        r = start + i - 1
        cells = [
            (1, i, True),
            (2, _safe_txt(e.debtor_name or ""), False),
            (3, _safe_txt(e.debtor_bin or ""), False),
            (4, _safe_txt(e.court_act_summary or ""), False),
            (5, _excel_num(e.amount_total), True),
            (6, _excel_num(e.amount_main), True),
            (7, _excel_num(e.amount_fines), True),
            (8, _excel_num(e.amount_fees), True),
            (9, _safe_txt(e.progress_notes or ""), False),
            (10, _excel_num(e.collected_amount), True),
            (11, _safe_txt(e.collection_doc_ref or ""), False),
            (12, _excel_num(e.balance_remaining), True),
            (13, _safe_txt(e.status_label or ""), False),
        ]
        for col, value, numeric in cells:
            _apply_data_cell(ws.cell(row=r, column=col), value, numeric=numeric)
        for col in enforcement_money_cols:
            ws.cell(row=r, column=col).number_format = _MONEY_FORMAT


def _unhide_report_sheets(wb: Workbook) -> None:
    """Шаблон КТЖ часто помечает листы как hidden; в выгрузке все вкладки должны быть видимы."""
    for name in wb.sheetnames:
        ws = wb[name]
        st = getattr(ws, "sheet_state", "visible") or "visible"
        if st != "visible":
            ws.sheet_state = "visible"


def _fill_debt_sheet(wb: Workbook, cases: list[Case], date_from: date, date_to: date) -> None:
    ws = wb[DEBT_SHEET]
    start = DEBT_FIRST_ROW
    _unmerge_in_range(ws, start, None)
    _clear_rows(ws, start, max(ws.max_row, start + 50), max_col=15)

    rows: list[DebtRecoveryEntry] = []
    for c in cases:
        rows.extend(_debt_rows_in_period(c, date_from, date_to))
    rows.sort(key=lambda x: (x.recorded_at, str(x.id)))

    debt_money_cols = {4, 6, 8}
    for i, d in enumerate(rows, start=1):
        r = start + i - 1
        cells = [
            (1, i, True),
            (2, _safe_txt(d.debtor_name or ""), False),
            (3, _safe_txt(d.debtor_status or ""), False),
            (4, _excel_num(d.debt_amount), True),
            (5, _safe_txt(d.work_summary or ""), False),
            (6, _excel_num(d.paid_amount), True),
            (8, _excel_num(d.written_off_amount), True),
        ]
        for col, value, numeric in cells:
            _apply_data_cell(ws.cell(row=r, column=col), value, numeric=numeric)
        for col in debt_money_cols:
            ws.cell(row=r, column=col).number_format = _MONEY_FORMAT


# ---------------------------------------------------------------------------
# Сборка zip-пакета (фикс для Apple Numbers)
# ---------------------------------------------------------------------------


_UNWANTED_SHEETS = (
    "в качестве 3 лица",
    "исполнительное производство",
    "инф по сниже дебит. задолжненно",
    "Лист1",
)

_TITLE_TEMPLATE = (
    'Результат исковой работы юридической службы АО "Пассажирские перевозки" '
    'и его филиалов, АО "Вагонсервис" {period}'
)

_RU_MONTHS_NOMINATIVE = (
    "январь", "февраль", "март", "апрель", "май", "июнь",
    "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь",
)

_QUARTER_ORDINAL = {1: "1-ый", 2: "2-ой", 3: "3-ий", 4: "4-ый"}


def _format_period_ru(date_from: date, date_to: date) -> str:
    """Превращает [date_from, date_to] в фразу-период для шапки отчёта."""
    from calendar import monthrange

    if date_from > date_to:
        date_from, date_to = date_to, date_from

    last_day_to = monthrange(date_to.year, date_to.month)[1]
    same_year = date_from.year == date_to.year
    starts_month = date_from.day == 1
    ends_month = date_to.day == last_day_to

    # Полный год
    if same_year and date_from.month == 1 and date_from.day == 1 and date_to.month == 12 and date_to.day == 31:
        return f"за {date_from.year} год"

    # Квартал
    if same_year and starts_month and ends_month:
        q_starts = {1: (1, 3), 2: (4, 6), 3: (7, 9), 4: (10, 12)}
        for q, (m1, m2) in q_starts.items():
            if date_from.month == m1 and date_to.month == m2:
                return f"за {_QUARTER_ORDINAL[q]} квартал {date_from.year} года"

    # Один месяц целиком
    if (same_year and starts_month and ends_month and date_from.month == date_to.month):
        return f"за {_RU_MONTHS_NOMINATIVE[date_from.month - 1]} {date_from.year} года"

    # Произвольный период
    return f"за период с {date_from.strftime('%d.%m.%Y')} по {date_to.strftime('%d.%m.%Y')}"


def _set_report_title(wb: Workbook, period_phrase: str) -> None:
    """Записать заголовок отчёта с подставленным периодом в A1 каждого оставшегося листа."""
    title = _TITLE_TEMPLATE.format(period=period_phrase)
    for name in ("истец", "ответчик", "3-лицо "):
        if name in wb.sheetnames:
            wb[name].cell(row=1, column=1).value = title


def _trim_extra_columns(wb) -> None:
    """Удаляет пустые колонки правее реальных данных на основных листах.

    В исходном шаблоне после колонки 19 («Информация о возмещении ущерба
    виновными лицами») остались «фантомные» ячейки с границами без заголовков
    и значений — они проявляются в выгрузке как пустые клетки справа.
    """
    for sheet_name, last_col in MAIN_MAX_COL.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        extra = ws.max_column - last_col
        if extra > 0:
            ws.delete_cols(last_col + 1, extra)


def build_pir_workbook_bytes(cases: list[Case], date_from: date, date_to: date) -> bytes:
    if not TEMPLATE_PATH.is_file():
        raise FileNotFoundError(f"PIR template missing: {TEMPLATE_PATH}")
    wb = load_workbook(TEMPLATE_PATH)
    # Заказчик попросил оставить только три листа: «истец», «ответчик», «3-лицо ».
    for name in _UNWANTED_SHEETS:
        if name in wb.sheetnames:
            del wb[name]
    _unhide_report_sheets(wb)
    _fill_main_sheets(wb, cases)
    _trim_extra_columns(wb)
    _set_report_title(wb, _format_period_ru(date_from, date_to))
    # Заставим Excel пересчитать все формулы при первом открытии (иначе ячейки
    # с формулами будут пустыми, потому что openpyxl сам формулы не вычисляет).
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass
    buf = BytesIO()
    wb.save(buf)
    return merge_openpyxl_into_template_package(TEMPLATE_PATH, buf.getvalue())
