"""Compare generated PIR workbook with the original template (headers + structure)."""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _norm(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, float):
        if v == int(v):
            return int(v)
        return round(v, 6)
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def cell_values_equal(a: Any, b: Any) -> bool:
    return _norm(a) == _norm(b)


def compare_region(
    ws_orig: Worksheet,
    ws_gen: Worksheet,
    row1: int,
    row2: int,
    col1: int,
    col2: int,
) -> list[tuple[int, int, Any, Any]]:
    mismatches: list[tuple[int, int, Any, Any]] = []
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            vo = ws_orig.cell(r, c).value
            vg = ws_gen.cell(r, c).value
            if not cell_values_equal(vo, vg):
                mismatches.append((r, c, vo, vg))
    return mismatches


@dataclass
class PirParityReport:
    template_path: str
    sheetnames_match: bool
    template_sheets: list[str] = field(default_factory=list)
    generated_sheets: list[str] = field(default_factory=list)
    header_mismatches: dict[str, list[tuple[int, int, Any, Any]]] = field(default_factory=dict)
    merged_ranges_orig: dict[str, int] = field(default_factory=dict)
    merged_ranges_gen: dict[str, int] = field(default_factory=dict)
    data_area_nonempty_cells_gen: dict[str, int] = field(default_factory=dict)
    summary: str = ""

    def build_summary(self) -> str:
        lines = [
            "=== Сверка выгрузки ПИР с оригинальным шаблоном ===",
            f"Шаблон: {self.template_path}",
            f"Состав листов совпадает: {'да' if self.sheetnames_match else 'нет'}",
        ]
        if not self.sheetnames_match:
            lines.append(f"  шаблон: {self.template_sheets}")
            lines.append(f"  выгрузка: {self.generated_sheets}")
        for sh, mm in self.header_mismatches.items():
            lines.append(f"Шапка «{sh}»: расхождений {len(mm)}")
            for r, c, vo, vg in mm[:12]:
                lines.append(f"  R{r}C{c}: шаблон={vo!r} vs выгрузка={vg!r}")
            if len(mm) > 12:
                lines.append(f"  … ещё {len(mm) - 12} ячеек")
        lines.append("Количество merged-диапазонов (openpyxl) по ключевым листам:")
        keys = sorted(set(self.merged_ranges_orig) | set(self.merged_ranges_gen))
        for k in keys:
            lines.append(f"  {k}: шаблон={self.merged_ranges_orig.get(k, 0)} выгрузка={self.merged_ranges_gen.get(k, 0)}")
        lines.append("Ненулевых/непустых ячеек в зоне данных (после первой строки данных):")
        for k, n in sorted(self.data_area_nonempty_cells_gen.items()):
            lines.append(f"  {k}: {n}")
        self.summary = "\n".join(lines)
        return self.summary


def analyze_pir_vs_template(template_path: str, generated_bytes: bytes) -> PirParityReport:
    wb_o = load_workbook(template_path, read_only=False, data_only=True)
    wb_g = load_workbook(BytesIO(generated_bytes), read_only=False, data_only=True)

    names_o = wb_o.sheetnames
    names_g = wb_g.sheetnames
    report = PirParityReport(
        template_path=template_path,
        sheetnames_match=names_o == names_g,
        template_sheets=list(names_o),
        generated_sheets=list(names_g),
    )

    # Header regions (everything strictly above the first category divider on PIR sheets).
    # Сами разделители категорий и сводный блок генерируются динамически — их сюда не включаем.
    header_regions: dict[str, tuple[int, int, int, int]] = {
        "истец": (1, 6, 1, 19),
        "ответчик": (1, 5, 1, 19),
        "3-лицо ": (1, 6, 1, 19),
        "в качестве 3 лица": (1, 6, 1, 19),
        "исполнительное производство": (1, 5, 1, 14),
        "инф по сниже дебит. задолжненно": (1, 6, 1, 15),
    }

    for sh, (r1, r2, c1, c2) in header_regions.items():
        if sh not in wb_o.sheetnames or sh not in wb_g.sheetnames:
            continue
        mm = compare_region(wb_o[sh], wb_g[sh], r1, r2, c1, c2)
        if mm:
            report.header_mismatches[sh] = mm

    for sh in header_regions:
        if sh in wb_o.sheetnames:
            report.merged_ranges_orig[sh] = len(wb_o[sh].merged_cells.ranges)
        if sh in wb_g.sheetnames:
            report.merged_ranges_gen[sh] = len(wb_g[sh].merged_cells.ranges)

    # Count non-empty cells in data band (heuristic: from known first divider/data row to +120).
    # Для основных листов берём строку первой категории (на 1 выше первой строки данных),
    # чтобы пустая выгрузка давала 0, а заполненная — учитывала и заголовки разделов.
    first_data = {
        "истец": 7,
        "ответчик": 6,
        "3-лицо ": 7,
        "в качестве 3 лица": 7,
        "исполнительное производство": 6,
        "инф по сниже дебит. задолжненно": 7,
    }
    for sh, start in first_data.items():
        if sh not in wb_g.sheetnames:
            continue
        ws = wb_g[sh]
        n = 0
        end = min(ws.max_row, start + 120)
        for r in range(start, end + 1):
            for c in range(1, 23):
                v = ws.cell(r, c).value
                if v is not None and str(v).strip() != "":
                    n += 1
        report.data_area_nonempty_cells_gen[sh] = n

    report.build_summary()
    return report
