"""Unit tests for PIR Excel export and parity with the official template."""

from __future__ import annotations

import uuid
import zipfile
from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from app.models import (
    Branch,
    Case,
    CaseFinance,
    CaseLitigation,
    DebtRecoveryEntry,
    EnforcementProceeding,
    User,
)
from app.domain.pir_excel_fill import (
    DEBT_FIRST_ROW,
    ENFORCEMENT_FIRST_ROW,
    FIRST_DATA_ROW,
    MAIN_MAX_COL,
    TEMPLATE_PATH,
    build_pir_workbook_bytes,
)

from tests.pir_compare import analyze_pir_vs_template, cell_values_equal


def _branch_and_lawyer() -> tuple[Branch, User]:
    bid = uuid.UUID("aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa")
    uid = uuid.UUID("bbbbbbbb-bbbb-bbbb-bbbb-bbbbbbbbbbbb")
    br = Branch(id=bid, name="ЦЛЮ / тест", city="Астана")
    lawyer = User(
        id=uid,
        email="test-lawyer@example.kz",
        password_hash="x",
        full_name="Тестов Ю.И.",
        role="branch_lawyer",
        branch_id=bid,
    )
    return br, lawyer


def _case_plaintiff(*, with_lit: bool = True, dispute_category: str = "procurement") -> Case:
    br, lawyer = _branch_and_lawyer()
    cid = uuid.UUID("cccccccc-cccc-cccc-cccc-cccccccccccc")
    today = date(2025, 6, 15)
    fin = CaseFinance(
        case_id=cid,
        claim_amount=Decimal("100000"),
        main_debt=Decimal("80000"),
        state_fee=Decimal("5000"),
        fines=Decimal("15000"),
        rep_expenses=Decimal("0"),
        other_costs=Decimal("0"),
        paid_amount=Decimal("10000"),
        recovered_main=Decimal("70000"),
        recovered_fines=Decimal("1000"),
        recovered_state_fee=Decimal("3000"),
        recovered_rep_expenses=Decimal("0"),
    )
    case = Case(
        id=cid,
        case_number="Т-2025-PIR-1",
        court="СМЭС тестовой области",
        court_instance="first",
        case_type="civil",
        status="active",
        outcome="pending",
        party_role="plaintiff",
        opponent_type="juridical",
        plaintiff="АО «Тест»",
        defendant='ТОО «Контрагент»',
        company="АО «Тест»",
        company_bin="123456789012",
        city="Астана",
        judge="Судья Тестов Т.Т.",
        filing_date=today,
        next_hearing=None,
        payment_deadline=None,
        last_updated=today,
        days_overdue=0,
        risk_level="low",
        is_archived=False,
        dispute_category=dispute_category,
        branch_id=br.id,
        assigned_lawyer_id=lawyer.id,
        branch=br,
        assigned_lawyer=lawyer,
        finances=fin,
        enforcement_proceedings=[],
        debt_recovery_entries=[],
    )
    if with_lit:
        from datetime import datetime, timezone

        now = datetime.now(timezone.utc)
        case.litigation = CaseLitigation(
            case_id=cid,
            claim_summary="О взыскании задолженности (тест).",
            judgment_first="Решение: тест.",
            judgment_appeal="",
            judgment_cassation="",
            damage_recovery_note="",
            writ_request_note="Заявление о выписке ИЛ от 01.01.2025",
            writ_dispatch_note="ЧСИ, исп. лист №1",
            execution_proof_note="Платёжное поручение №9",
            defendant_execution_note="",
            third_party_note="",
            created_at=now,
            updated_at=now,
        )
    return case


def _case_defendant() -> Case:
    """Дело-ответчик с различимыми взысканными суммами для проверки колонок 14-18."""
    br, lawyer = _branch_and_lawyer()
    cid = uuid.UUID("dddddddd-dddd-dddd-dddd-dddddddddddd")
    today = date(2025, 7, 1)
    fin = CaseFinance(
        case_id=cid,
        claim_amount=Decimal("500000"),
        main_debt=Decimal("400000"),
        state_fee=Decimal("12000"),
        fines=Decimal("80000"),
        rep_expenses=Decimal("8000"),
        other_costs=Decimal("0"),
        paid_amount=Decimal("0"),
        recovered_main=Decimal("100000"),
        recovered_fines=Decimal("20000"),
        recovered_state_fee=Decimal("3000"),
        recovered_rep_expenses=Decimal("4000"),
    )
    from datetime import datetime, timezone

    now = datetime.now(timezone.utc)
    return Case(
        id=cid,
        case_number="Т-2025-DEF-1",
        court="СМЭС теста",
        court_instance="first",
        case_type="civil",
        status="active",
        outcome="partially_satisfied",
        party_role="defendant",
        opponent_type="juridical",
        plaintiff="ТОО «Истец»",
        defendant='АО «Тест»',
        company="АО «Тест»",
        company_bin="123456789012",
        city="Астана",
        judge="Судья Алимов",
        filing_date=today,
        next_hearing=None,
        payment_deadline=None,
        last_updated=today,
        days_overdue=0,
        risk_level="medium",
        is_archived=False,
        dispute_category="transportation",
        branch_id=br.id,
        assigned_lawyer_id=lawyer.id,
        branch=br,
        assigned_lawyer=lawyer,
        finances=fin,
        enforcement_proceedings=[],
        debt_recovery_entries=[],
        litigation=CaseLitigation(
            case_id=cid,
            claim_summary="О взыскании по перевозке.",
            judgment_first="Решение от 01.06.2025: удовлетворено частично.",
            judgment_appeal="",
            judgment_cassation="",
            damage_recovery_note="",
            writ_request_note="",
            writ_dispatch_note="",
            execution_proof_note="",
            defendant_execution_note="пл.поручение №77 от 15.06.2025",
            third_party_note="",
            created_at=now,
            updated_at=now,
        ),
    )


def _case_third_party() -> Case:
    """Дело с КТЖ-ГП в качестве третьего лица."""
    br, lawyer = _branch_and_lawyer()
    cid = uuid.UUID("eeeeeeee-eeee-eeee-eeee-eeeeeeeeeeee")
    today = date(2025, 8, 1)
    fin = CaseFinance(
        case_id=cid,
        claim_amount=Decimal("0"),
        main_debt=Decimal("0"),
        state_fee=Decimal("0"),
        fines=Decimal("0"),
        rep_expenses=Decimal("0"),
        other_costs=Decimal("0"),
        paid_amount=Decimal("0"),
        recovered_main=Decimal("0"),
        recovered_fines=Decimal("0"),
        recovered_state_fee=Decimal("0"),
        recovered_rep_expenses=Decimal("0"),
    )
    from datetime import datetime, timezone

    now = datetime.now(timezone.utc)
    return Case(
        id=cid,
        case_number="Т-2025-3P-1",
        court="СМАС теста",
        court_instance="first",
        case_type="administrative",
        status="active",
        outcome="pending",
        party_role="third_party",
        opponent_type="juridical",
        plaintiff="ТОО «Поставщик»",
        defendant="ТОО «Самрук-Казына Контракт»",
        company="АО «КТЖ»",
        company_bin="123456789012",
        city="Алматы",
        judge="",
        filing_date=today,
        next_hearing=None,
        payment_deadline=None,
        last_updated=today,
        days_overdue=0,
        risk_level="low",
        is_archived=False,
        dispute_category="other",
        branch_id=br.id,
        assigned_lawyer_id=lawyer.id,
        branch=br,
        assigned_lawyer=lawyer,
        finances=fin,
        enforcement_proceedings=[],
        debt_recovery_entries=[],
        litigation=CaseLitigation(
            case_id=cid,
            claim_summary="Об оспаривании.",
            judgment_first="",
            judgment_appeal="",
            judgment_cassation="",
            damage_recovery_note="",
            writ_request_note="",
            writ_dispatch_note="",
            execution_proof_note="",
            defendant_execution_note="",
            third_party_note="Участие в качестве заинтересованного лица",
            created_at=now,
            updated_at=now,
        ),
    )


def test_template_file_exists():
    assert TEMPLATE_PATH.is_file(), f"Ожидается шаблон: {TEMPLATE_PATH}"


def test_export_unhides_auxiliary_sheets():
    """В шаблоне листы ИП и дебиторки скрыты; в выгрузке они должны быть видимы."""
    out = build_pir_workbook_bytes([], date(2025, 1, 1), date(2025, 12, 31))
    wb = load_workbook(BytesIO(out))
    try:
        assert wb["исполнительное производство"].sheet_state == "visible"
        assert wb["инф по сниже дебит. задолжненно"].sheet_state == "visible"
        assert wb["в качестве 3 лица"].sheet_state == "visible"
        assert wb["Лист1"].sheet_state == "visible"
    finally:
        wb.close()


def test_generated_workbook_restores_template_app_xml_for_numbers():
    """openpyxl would otherwise emit a minimal docProps/app.xml that Apple Numbers often rejects."""
    out = build_pir_workbook_bytes([], date(2025, 1, 1), date(2025, 12, 31))
    with zipfile.ZipFile(BytesIO(out), "r") as zf:
        app = zf.read("docProps/app.xml").decode("utf-8")
    assert "Openpyxl" not in app
    assert "HeadingPairs" in app and "TitlesOfParts" in app
    assert "Microsoft Excel" in app


def test_pir_package_portable_without_legacy_drawings():
    """Huge template xl/drawings (especially drawing2.xml) break Apple Numbers; export must drop them."""
    out = build_pir_workbook_bytes([], date(2025, 1, 1), date(2025, 12, 31))
    assert len(out) < 2_000_000, f"expected compact xlsx without embedded giant drawings, got {len(out)}"
    with zipfile.ZipFile(BytesIO(out), "r") as zf:
        names = zf.namelist()
        assert not any(n.startswith("xl/drawings/") for n in names)
        assert "xl/sharedStrings.xml" not in names
        ct = zf.read("[Content_Types].xml").decode("utf-8")
        assert "/xl/sharedStrings.xml" not in ct
        assert zf.testzip() is None
        for n in names:
            if n.startswith("xl/worksheets/_rels/") and n.endswith(".rels"):
                body = zf.read(n).decode("utf-8")
                assert "drawing" not in body.lower(), n


def test_main_role_sheets_header_text_not_rotated():
    """В эталонном ПИР заголовки колонок горизонтальны; после выгрузки не должно быть вертикального текста (90° / stacked)."""
    out = build_pir_workbook_bytes([], date(2025, 1, 1), date(2025, 12, 31))
    wb = load_workbook(BytesIO(out))
    try:
        for sheet_name in ("истец", "ответчик", "3-лицо ", "в качестве 3 лица"):
            ws = wb[sheet_name]
            last_h = FIRST_DATA_ROW[sheet_name] - 1
            max_col = MAIN_MAX_COL[sheet_name]
            for r in range(1, last_h + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    if isinstance(cell, MergedCell):
                        continue
                    al = cell.alignment
                    rot = getattr(al, "text_rotation", None) if al else None
                    assert rot in (0, None), f"{sheet_name} R{r}C{c}: text_rotation={rot!r}"
    finally:
        wb.close()


def test_build_empty_cases_headers_match_original():
    """Пустая выгрузка: шапки ключевых листов совпадают с шаблоном, демо-строки очищены."""
    out = build_pir_workbook_bytes([], date(2025, 1, 1), date(2025, 12, 31))
    rep = analyze_pir_vs_template(str(TEMPLATE_PATH), out)
    assert rep.sheetnames_match, (rep.template_sheets, rep.generated_sheets)
    assert not rep.header_mismatches, rep.header_mismatches
    # После выгрузки без дел в зоне данных основных листов не должно остаться значений из образца
    assert rep.data_area_nonempty_cells_gen.get("истец", 0) == 0
    assert rep.data_area_nonempty_cells_gen.get("ответчик", 0) == 0
    assert rep.data_area_nonempty_cells_gen.get("3-лицо ", 0) == 0


def test_build_plaintiff_row_values():
    case = _case_plaintiff()
    out = build_pir_workbook_bytes([case], date(2025, 1, 1), date(2025, 12, 31))
    wb = load_workbook(BytesIO(out), read_only=True, data_only=True)
    ws = wb["истец"]
    r = FIRST_DATA_ROW["истец"]
    assert cell_values_equal(ws.cell(r, 1).value, 1)
    assert ws.cell(r, 2).value == "Тестов Ю.И."
    assert "СМЭС тестовой области" in (ws.cell(r, 3).value or "")
    assert "Судья Тестов" in (ws.cell(r, 3).value or "")
    assert ws.cell(r, 4).value == "ЦЛЮ / тест"
    assert ws.cell(r, 5).value == 'ТОО «Контрагент»'
    assert float(ws.cell(r, 6).value or 0) == 80000.0
    assert float(ws.cell(r, 13).value or 0) == 70000.0
    assert float(ws.cell(r, 14).value or 0) == 1000.0
    assert float(ws.cell(r, 15).value or 0) == 3000.0
    assert "Заявление о выписке" in (ws.cell(r, 16).value or "")
    assert "ЧСИ" in (ws.cell(r, 17).value or "")
    assert "Платёжное поручение" in (ws.cell(r, 18).value or "")
    assert "О взыскании задолженности" in (ws.cell(r, 9).value or "")


def test_enforcement_and_debt_rows_in_period():
    case = _case_plaintiff()
    eid = uuid.uuid4()
    did = uuid.uuid4()
    rd = date(2025, 6, 10)
    case.enforcement_proceedings = [
        EnforcementProceeding(
            id=eid,
            case_id=case.id,
            debtor_name="ТОО Должник",
            debtor_bin="940440000385",
            court_act_summary="ИЛ №123",
            amount_total=Decimal("5000"),
            amount_main=Decimal("4000"),
            amount_fines=Decimal("500"),
            amount_fees=Decimal("500"),
            progress_notes="ход",
            collected_amount=Decimal("0"),
            collection_doc_ref="",
            balance_remaining=Decimal("5000"),
            status_label="В работе",
            recorded_at=rd,
        )
    ]
    case.debt_recovery_entries = [
        DebtRecoveryEntry(
            id=did,
            case_id=case.id,
            counterparty_bin=None,
            debtor_name="Дебитор X",
            debtor_status="действующий",
            debt_amount=Decimal("9000"),
            paid_amount=Decimal("1000"),
            written_off_amount=Decimal("0"),
            work_summary="Претензия",
            recorded_at=rd,
        )
    ]
    out = build_pir_workbook_bytes([case], date(2025, 6, 1), date(2025, 6, 30))
    wb = load_workbook(BytesIO(out), read_only=True, data_only=True)
    wse = wb["исполнительное производство"]
    r = ENFORCEMENT_FIRST_ROW
    assert wse.cell(r, 2).value == "ТОО Должник"
    assert str(wse.cell(r, 3).value) == "940440000385"
    assert float(wse.cell(r, 5).value or 0) == 5000.0

    wsd = wb["инф по сниже дебит. задолжненно"]
    rd_row = DEBT_FIRST_ROW
    assert wsd.cell(rd_row, 2).value == "Дебитор X"
    assert float(wsd.cell(rd_row, 4).value or 0) == 9000.0


def test_parity_report_against_template_synthetic(capfd: pytest.CaptureFixture[str]):
    """Финальная сверка: структура как у шаблона + шапки совпадают; печать отчёта в лог pytest."""
    case = _case_plaintiff()
    case.enforcement_proceedings = []
    case.debt_recovery_entries = []
    out = build_pir_workbook_bytes([case], date(2025, 1, 1), date(2025, 12, 31))
    rep = analyze_pir_vs_template(str(TEMPLATE_PATH), out)
    assert rep.sheetnames_match
    assert not rep.header_mismatches, rep.header_mismatches
    # Один истец → одна непустая строка данных на листе «истец»
    assert rep.data_area_nonempty_cells_gen.get("истец", 0) >= 8
    print("\n" + rep.summary)
    capfd.readouterr()  # flush if needed


def test_defendant_recovered_columns_and_execution_note():
    """На листе «ответчик»: cols 14/15/16/17 — взысканные (осн./штр./предст./госп.), col 18 — инф. об исполнении."""
    case = _case_defendant()
    out = build_pir_workbook_bytes([case], date(2025, 1, 1), date(2025, 12, 31))
    wb = load_workbook(BytesIO(out), read_only=True, data_only=True)
    try:
        ws = wb["ответчик"]
        # Категория transportation → разделитель в row 6, данные с row 7
        r = FIRST_DATA_ROW["ответчик"]
        assert ws.cell(r, 1).value == 1, f"Ожидали seq=1, получили {ws.cell(r, 1).value!r}"
        # Колонки взысканной суммы
        assert float(ws.cell(r, 14).value or 0) == 100000.0, "recovered_main"
        assert float(ws.cell(r, 15).value or 0) == 20000.0, "recovered_fines"
        assert float(ws.cell(r, 16).value or 0) == 4000.0, "recovered_rep_expenses"
        assert float(ws.cell(r, 17).value or 0) == 3000.0, "recovered_state_fee"
        # Информация об исполнении (col 18) и возмещение (col 19)
        assert "пл.поручение №77" in str(ws.cell(r, 18).value or "")
    finally:
        wb.close()


def test_defendant_categories_are_grouped_with_dividers():
    """Каждая категория получает строку-разделитель, нумерация дел внутри категории — с 1."""
    a = _case_defendant()
    a.id = uuid.UUID("11111111-1111-1111-1111-111111111111")
    a.case_number = "DEF-A"
    a.dispute_category = "procurement"
    a.finances.case_id = a.id
    a.litigation.case_id = a.id

    b = _case_defendant()
    b.id = uuid.UUID("22222222-2222-2222-2222-222222222222")
    b.case_number = "DEF-B"
    b.dispute_category = "labor"
    b.finances.case_id = b.id
    b.litigation.case_id = b.id

    c = _case_defendant()
    c.id = uuid.UUID("33333333-3333-3333-3333-333333333333")
    c.case_number = "DEF-C"
    c.dispute_category = "labor"
    c.finances.case_id = c.id
    c.litigation.case_id = c.id

    out = build_pir_workbook_bytes([a, b, c], date(2025, 1, 1), date(2025, 12, 31))
    wb = load_workbook(BytesIO(out), read_only=True, data_only=True)
    try:
        ws = wb["ответчик"]
        # Шаблонный порядок: procurement → transportation → labor → other → mediation
        # Строки: 6=procurement-divider, 7=DEF-A, 8=labor-divider, 9=DEF-B, 10=DEF-C
        assert "Иски, связанные" in str(ws.cell(6, 1).value or "")
        assert ws.cell(7, 1).value == 1  # первый procurement
        assert "Трудовые споры" in str(ws.cell(8, 1).value or "")
        assert ws.cell(9, 1).value == 1, f"labor seq must restart at 1, got {ws.cell(9, 1).value!r}"
        assert ws.cell(10, 1).value == 2
    finally:
        wb.close()


def test_third_party_role_goes_to_correct_sheet_and_mirror():
    """party_role=third_party → лист «в качестве 3 лица» + дублирование в «3-лицо » (для совместимости с шаблоном)."""
    case = _case_third_party()
    out = build_pir_workbook_bytes([case], date(2025, 1, 1), date(2025, 12, 31))
    wb = load_workbook(BytesIO(out), read_only=True, data_only=True)
    try:
        primary = wb["в качестве 3 лица"]
        mirror = wb["3-лицо "]
        for ws in (primary, mirror):
            r = FIRST_DATA_ROW["в качестве 3 лица"]
            assert ws.cell(r, 1).value == 1, f"{ws.title}: expected seq=1, got {ws.cell(r, 1).value!r}"
            assert ws.cell(r, 5).value == "ТОО «Поставщик»"
            assert ws.cell(r, 6).value == "ТОО «Самрук-Казына Контракт»"
            # Колонка 19 — примечание
            assert "Участие в качестве заинтересованного лица" in str(ws.cell(r, 19).value or "")
    finally:
        wb.close()


def test_plaintiff_transportation_falls_back_to_other():
    """Категория transportation на листе «истец» отсутствует — дело должно попасть в «Иные споры»."""
    case = _case_plaintiff(dispute_category="transportation")
    out = build_pir_workbook_bytes([case], date(2025, 1, 1), date(2025, 12, 31))
    wb = load_workbook(BytesIO(out), read_only=True, data_only=True)
    try:
        ws = wb["истец"]
        # Должны увидеть только разделитель «Иные споры», procurement/labor/mediation — пропущены
        # Порядок: procurement / labor / other / mediation. transport на «истец» нет.
        # Поскольку only "other" имеет элементы, на листе будет один разделитель «Иные споры»
        # и одна строка данных подряд.
        divider_row = FIRST_DATA_ROW["истец"] - 1
        assert "Иные споры" in str(ws.cell(divider_row, 1).value or ""), (
            f"row {divider_row} = {ws.cell(divider_row, 1).value!r}"
        )
        data_row = FIRST_DATA_ROW["истец"]
        assert ws.cell(data_row, 1).value == 1
    finally:
        wb.close()
